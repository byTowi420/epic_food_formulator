from typing import Any, Dict, List
from pathlib import Path
import json
import re
import unicodedata
import os
from fractions import Fraction
import math

from PySide6.QtCore import QObject, QThread, Qt, QItemSelectionModel, QCoreApplication, QTimer, QEvent
from PySide6.QtGui import QIcon, QPixmap, QPainter, QColor, QFont, QKeySequence, QShortcut, QBrush
from PySide6.QtWidgets import (
    QComboBox,
    QApplication,
    QFileDialog,
    QHBoxLayout,
    QGridLayout,
    QGroupBox,
    QInputDialog,
    QLabel,
    QLineEdit,
    QCheckBox,
    QMainWindow,
    QMessageBox,
    QDialog,
    QDialogButtonBox,
    QListWidget,
    QListWidgetItem,
    QPlainTextEdit,
    QPushButton,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QSizePolicy,
    QHeaderView,
    QSplitter,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment

import logging

from services.usda_api import USDAApiError, get_food_details, search_foods, has_cached_food
from services.nutrient_normalizer import (
    augment_fat_nutrients,
    canonical_alias_name,
    canonical_unit,
    normalize_nutrients,
)
from ui.workers import ApiWorker, ImportWorker, AddWorker

logging.basicConfig(
    filename="app_debug.log",
    level=logging.DEBUG,
    format="%(asctime)s [%(threadName)s] %(levelname)s %(message)s",
)


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()

        self.base_window_title = "Food Formulator - Proto"
        self.setWindowTitle(self.base_window_title)
        self.import_max_attempts = 4
        self.import_read_timeout = 8.0
        self.resize(900, 600)
        self._threads: list[QThread] = []
        self._workers: list[QObject] = []
        self._current_import_worker: ImportWorker | None = None
        self._current_add_worker: AddWorker | None = None
        self._prefetching_fdc_ids: set[int] = set()
        self.formulation_items: List[Dict] = []
        self.quantity_mode: str = "g"
        self.amount_g_column_index = 2
        self.percent_column_index = 3
        self.lock_column_index = 4
        self.nutrient_export_flags: Dict[str, bool] = {}
        self.last_path = self._load_last_path()
        self.search_page = 1
        self.search_page_size = 25
        self.search_fetch_page_size = 200
        self.search_max_pages = 5
        self.search_results: List[Dict[str, Any]] = []
        self.last_query = ""
        self.last_include_brands = False
        self._last_results_count = 0
        self.data_type_priority = {
            "Foundation": 0,
            "SR Legacy": 1,
            "Survey": 2,
            "Survey (FNDDS)": 2,
            "Experimental": 3,
            "Branded": 4,
        }
        self._reference_order_map: Dict[str, Dict[str, Any]] = {}
        self._nutrient_catalog: list[tuple[str, list[str]]] = self._build_nutrient_catalog()
        self._nutrient_order_map: dict[str, int] = {}
        self._nutrient_category_map: dict[str, str] = {}
        for idx, (_, names) in enumerate(self._nutrient_catalog):
            for offset, name in enumerate(names):
                self._nutrient_order_map[name.strip().lower()] = idx * 1000 + offset
                self._nutrient_category_map[name.strip().lower()] = self._nutrient_catalog[idx][0]

        self.label_base_nutrients = self._build_base_label_nutrients()
        self.household_measure_options = self._build_household_measure_options()
        self.household_capacity_map = {
            name: capacity for name, capacity in self.household_measure_options
        }
        self._auto_updating_household_amount = False
        self.label_manual_overrides: dict[str, float] = {}
        self._last_totals: Dict[str, Dict[str, Any]] = {}
        self.label_no_significant: list[str] = []
        self._label_display_nutrients: list[Dict[str, Any]] = []
        self.label_manual_hint_color = QColor(204, 255, 204)
        self.label_no_sig_order = [
            "Energia",
            "Carbohidratos",
            "Proteinas",
            "Grasas totales",
            "Grasas saturadas",
            "Grasas trans",
            "Fibra alimentaria",
            "Sodio",
        ]
        self.label_nutrient_usda_map = {
            "Energia": "Energy (kcal)",
            "Carbohidratos": "Carbohydrate, by difference (g)",
            "Proteinas": "Protein (g)",
            "Grasas totales": "Total lipid (fat) (g)",
            "Grasas saturadas": "Fatty acids, total saturated (g)",
            "Grasas trans": "Fatty acids, total trans (g)",
            "Fibra alimentaria": "Fiber, total dietary (g)",
            "Sodio": "Sodium, Na (mg)",
        }
        self.label_no_significant_thresholds = {
            "Energia": {"unit": "kcal", "max": 4.0, "kj_max": 17.0},
            "Carbohidratos": {"unit": "g", "max": 0.5},
            "Proteinas": {"unit": "g", "max": 0.5},
            "Grasas totales": {"unit": "g", "max": 0.5},
            "Grasas saturadas": {"unit": "g", "max": 0.2},
            "Grasas trans": {"unit": "g", "max": 0.2},
            "Fibra alimentaria": {"unit": "g", "max": 0.5},
            "Sodio": {"unit": "mg", "max": 5.0},
        }
        self.label_no_significant_display_map = {"Energia": "Valor energético"}

        self._build_ui()

    def _set_window_progress(self, progress: str | None = None) -> None:
        """Update the window title with progress info or reset it."""
        title = self.base_window_title
        if progress:
            title = f"{title} | {progress}"
        self.setWindowTitle(title)

    def _build_ui(self) -> None:
        central = QWidget(self)
        self.setCentralWidget(central)

        main_layout = QVBoxLayout(central)

        self.tabs = QTabWidget()
        self.search_tab = QWidget()
        self.formulation_tab = QWidget()
        self.label_tab = QWidget()
        self.tabs.addTab(self.search_tab, "Búsqueda")
        self.tabs.addTab(self.formulation_tab, "Formulación")
        self.tabs.addTab(self.label_tab, "Etiqueta")

        main_layout.addWidget(self.tabs)

        self._build_search_tab_ui()
        self._build_formulation_tab_ui()
        self._build_label_tab_ui()

    def _build_search_tab_ui(self) -> None:
        layout = QVBoxLayout(self.search_tab)
        layout.setSpacing(6)

        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText(
            "Buscar alimento (ej: apple, rice, cheese)..."
        )
        self.search_button = QPushButton("Buscar")

        search_layout.addWidget(self.search_input)
        search_layout.addWidget(self.search_button)

        self.include_brands_checkbox = QCheckBox("Incluir Marcas")
        self.prev_page_button = QPushButton("<")
        self.prev_page_button.setFixedWidth(32)
        self.next_page_button = QPushButton(">")
        self.next_page_button.setFixedWidth(32)
        self.prev_page_button.setEnabled(False)
        self.next_page_button.setEnabled(False)

        # Controles legacy ocultos (buscar por FDC y botón de agregar)
        self.fdc_id_input = QLineEdit()
        self.fdc_id_input.hide()
        self.fdc_id_button = QPushButton("Cargar FDC ID")
        self.fdc_id_button.hide()
        self.add_button = QPushButton("Agregar seleccionado a formulación")
        self.add_button.hide()

        self.status_label = QLabel("")
        self.status_label.setAlignment(Qt.AlignLeft)
        self.status_label.setStyleSheet("color: gray;")
        self.status_label.setTextInteractionFlags(
            Qt.TextSelectableByMouse | Qt.TextSelectableByKeyboard
        )
        self.status_label.setWordWrap(True)
        self.status_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.status_label.setMaximumHeight(
            int(self.status_label.fontMetrics().height() * 2.4)
        )

        layout.addLayout(search_layout)
        status_controls_layout = QHBoxLayout()
        status_controls_layout.setContentsMargins(0, 0, 0, 0)
        status_controls_layout.addWidget(self.status_label, 1)
        status_controls_layout.addStretch()
        status_controls_layout.addWidget(self.include_brands_checkbox)
        status_controls_layout.addWidget(self.prev_page_button)
        status_controls_layout.addWidget(self.next_page_button)
        layout.addLayout(status_controls_layout)

        # Tabla de resultados (arriba)
        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(
            ["FDC ID", "Descripción", "Marca / Origen", "Tipo de dato"]
        )
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table.horizontalHeader().setStretchLastSection(True)

        self.details_table = QTableWidget(0, 3)
        self.details_table.setHorizontalHeaderLabels(
            ["Nutriente", "Cantidad", "Unidad"]
        )
        self.details_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.details_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.details_table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.details_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.details_table.horizontalHeader().setStretchLastSection(True)

        bottom_layout = QHBoxLayout()

        left_panel = QVBoxLayout()
        left_panel.addWidget(QLabel("Ingredientes en formulación"))
        self.formulation_preview = QTableWidget(0, 2)
        self.formulation_preview.setHorizontalHeaderLabels(["FDC ID", "Ingrediente"])
        self.formulation_preview.setEditTriggers(QTableWidget.NoEditTriggers)
        self.formulation_preview.setSelectionBehavior(QTableWidget.SelectRows)
        self.formulation_preview.setSelectionMode(QTableWidget.ExtendedSelection)
        self.formulation_preview.horizontalHeader().setStretchLastSection(True)
        left_panel.addWidget(self.formulation_preview)

        self.remove_preview_button = QPushButton("Eliminar ingrediente seleccionado")
        left_panel.addWidget(self.remove_preview_button)

        right_panel = QVBoxLayout()
        right_panel.addWidget(QLabel("Nutrientes del ingrediente seleccionado"))
        right_panel.addWidget(self.details_table)

        bottom_layout.addLayout(left_panel, 1)
        bottom_layout.addLayout(right_panel, 1)

        bottom_widget = QWidget()
        bottom_widget.setLayout(bottom_layout)

        splitter = QSplitter(Qt.Vertical)
        splitter.addWidget(self.table)
        splitter.addWidget(bottom_widget)
        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 2)
        splitter.setSizes([400, 500])

        layout.addWidget(splitter)

        self.search_button.clicked.connect(self.on_search_clicked)
        self.search_input.returnPressed.connect(self.on_search_clicked)
        self.table.cellDoubleClicked.connect(self.on_result_double_clicked)
        self.table.itemSelectionChanged.connect(self.on_search_selection_changed)
        self.remove_preview_button.clicked.connect(self.on_remove_preview_clicked)
        self.formulation_preview.cellDoubleClicked.connect(
            self.on_formulation_preview_double_clicked
        )
        self.formulation_preview.itemSelectionChanged.connect(
            self.on_preview_selection_changed
        )
        self.prev_page_button.clicked.connect(self.on_prev_page_clicked)
        self.next_page_button.clicked.connect(self.on_next_page_clicked)
        self.include_brands_checkbox.stateChanged.connect(
            self.on_include_brands_toggled
        )
        self._set_default_column_widths()

    def _build_formulation_tab_ui(self) -> None:
        layout = QVBoxLayout(self.formulation_tab)

        header_layout = QHBoxLayout()
        self.export_state_button = QPushButton("Exportar")
        self.import_state_button = QPushButton("Importar")
        header_layout.addWidget(self.export_state_button)
        header_layout.addWidget(self.import_state_button)
        header_layout.addStretch()
        self.formula_name_input = QLineEdit()
        self.formula_name_input.setPlaceholderText(
            "Nombre Fórmula Ej.: Pan dulce con chocolate"
        )
        self.formula_name_input.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(self.formula_name_input, 1)
        header_layout.addStretch()
        header_layout.addWidget(QLabel("Unidad de formulación:"))
        self.quantity_mode_selector = QComboBox()
        self.quantity_mode_selector.addItems(["Gramos (g)", "Porcentaje (%)"])
        header_layout.addWidget(self.quantity_mode_selector)
        layout.addLayout(header_layout)

        self.formulation_table = QTableWidget(0, 6)
        self.formulation_table.setHorizontalHeaderLabels(
            [
                "FDC ID",
                "Ingrediente",
                "Cantidad (g)",
                "Cantidad (%)",
                "Fijar %",
                "Marca / Origen",
            ]
        )
        self.formulation_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.formulation_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.formulation_table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.formulation_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.formulation_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.formulation_table)

        buttons_layout = QHBoxLayout()
        self.edit_quantity_button = QPushButton("Editar cantidad seleccionada")
        self.remove_formulation_button = QPushButton(
            "Eliminar ingrediente seleccionado"
        )
        buttons_layout.addWidget(self.edit_quantity_button)
        buttons_layout.addWidget(self.remove_formulation_button)
        buttons_layout.addStretch()
        layout.addLayout(buttons_layout)

        layout.addWidget(
            QLabel(
                "Totales nutricionales (asumiendo valores por 100 g del alimento origen)"
            )
        )
        self.totals_table = QTableWidget(0, 4)
        self.totals_table.setHorizontalHeaderLabels(
            ["Nutriente", "Total", "Unidad", "Exportar"]
        )
        export_header = QTableWidgetItem("Exportar")
        export_header.setIcon(self._create_question_icon())
        export_header.setToolTip(
            "Los nutrientes seleccionados seran exportados al excel"
        )
        self.totals_table.setHorizontalHeaderItem(3, export_header)
        self.totals_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.totals_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.totals_table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.totals_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.totals_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.totals_table)

        export_buttons_layout = QHBoxLayout()
        self.toggle_export_button = QPushButton("Deseleccionar todos")
        self.export_excel_button = QPushButton("Exportar a Excel")
        export_buttons_layout.addStretch()
        export_buttons_layout.addWidget(self.toggle_export_button)
        export_buttons_layout.addWidget(self.export_excel_button)
        export_buttons_layout.addStretch()
        layout.addLayout(export_buttons_layout)

        self.remove_formulation_button.clicked.connect(
            self.on_remove_formulation_clicked
        )
        self.export_excel_button.clicked.connect(self.on_export_to_excel_clicked)
        self.quantity_mode_selector.currentIndexChanged.connect(
            self.on_quantity_mode_changed
        )
        self.formulation_table.cellDoubleClicked.connect(
            self.on_formulation_cell_double_clicked
        )
        self.edit_quantity_button.clicked.connect(self.on_edit_quantity_clicked)
        self.formulation_table.itemChanged.connect(self.on_lock_toggled_from_table)
        self.totals_table.itemChanged.connect(self.on_totals_checkbox_changed)
        self.toggle_export_button.clicked.connect(self.on_toggle_export_clicked)
        self.export_state_button.clicked.connect(self.on_export_state_clicked)
        self.import_state_button.clicked.connect(self.on_import_state_clicked)

        # Copy shortcuts (Ctrl+C) for all tables
        for table in (
            self.table,
            self.details_table,
            self.formulation_preview,
            self.formulation_table,
            self.totals_table,
        ):
            self._attach_copy_shortcut(table)

        self._set_default_column_widths(formulation=True)

    def _build_label_tab_ui(self) -> None:
        layout = QVBoxLayout(self.label_tab)
        layout.setSpacing(10)

        top_layout = QHBoxLayout()
        top_layout.setSpacing(12)

        left_group = QGroupBox("Rotulado Nutricional")
        left_form = QGridLayout()
        left_form.setContentsMargins(10, 10, 10, 10)
        left_form.setHorizontalSpacing(8)
        left_form.setVerticalSpacing(6)
        left_group.setLayout(left_form)

        left_form.addWidget(QLabel("Tamaño Porción:"), 0, 0)
        self.portion_size_input = QSpinBox()
        self.portion_size_input.setRange(1, 100000)
        self.portion_size_input.setValue(100)
        self.portion_unit_combo = QComboBox()
        self.portion_unit_combo.addItems(["g", "ml"])
        portion_widget = QWidget()
        portion_layout = QHBoxLayout(portion_widget)
        portion_layout.setContentsMargins(0, 0, 0, 0)
        portion_layout.setSpacing(6)
        portion_layout.addWidget(self.portion_size_input)
        portion_layout.addWidget(self.portion_unit_combo)
        left_form.addWidget(portion_widget, 0, 1, 1, 2)

        left_form.addWidget(QLabel("Medida Casera:"), 1, 0)
        self.household_amount_input = QLineEdit()
        self.household_amount_input.setPlaceholderText("1/2, 2/3, 1 1/2")
        self.household_amount_input.setText("1/2")
        self.household_unit_combo = QComboBox()
        for name, _ in self.household_measure_options:
            self.household_unit_combo.addItem(name)
        unidad_index = self.household_unit_combo.findText("Unidad")
        if unidad_index >= 0:
            self.household_unit_combo.setCurrentIndex(unidad_index)
        measure_widget = QWidget()
        measure_layout = QHBoxLayout(measure_widget)
        measure_layout.setContentsMargins(0, 0, 0, 0)
        measure_layout.setSpacing(6)
        measure_layout.addWidget(self.household_amount_input)
        measure_layout.addWidget(self.household_unit_combo)
        left_form.addWidget(measure_widget, 1, 1, 1, 2)

        self.custom_household_unit_input = QLineEdit()
        self.custom_household_unit_input.setPlaceholderText(
            "Ej.: Envase, Barrita, Paquete"
        )
        self.custom_household_unit_input.setVisible(False)
        left_form.addWidget(self.custom_household_unit_input, 2, 1, 1, 2)

        left_form.addWidget(QLabel("Capacidad o dimensión:"), 3, 0)
        self.household_capacity_label = QLabel("-")
        self.household_capacity_label.setStyleSheet("color: gray;")
        left_form.addWidget(self.household_capacity_label, 3, 1, 1, 2)

        self.breakdown_carb_checkbox = QCheckBox("Desglose Carbohidratos")
        self.breakdown_carb_checkbox.setEnabled(False)
        self.breakdown_fat_checkbox = QCheckBox("Desglose Grasas")
        self.breakdown_fat_checkbox.setEnabled(False)
        left_form.addWidget(self.breakdown_carb_checkbox, 4, 0, 1, 2)
        left_form.addWidget(self.breakdown_fat_checkbox, 4, 2)

        left_form.addWidget(QLabel("Sin aportes significativos:"), 5, 0)
        self.no_significant_display = QLineEdit()
        self.no_significant_display.setReadOnly(True)
        self.no_significant_display.setPlaceholderText("Seleccione nutrientes elegibles")
        self.no_significant_display.setCursor(Qt.PointingHandCursor)
        left_form.addWidget(self.no_significant_display, 5, 1, 1, 2)

        left_form.addWidget(QLabel("Nutrientes Adicionales:"), 6, 0)
        self.additional_nutrients_input = QLineEdit()
        self.additional_nutrients_input.setPlaceholderText("Seleccione Nutrientes")
        self.additional_nutrients_input.setEnabled(False)
        left_form.addWidget(self.additional_nutrients_input, 6, 1, 1, 2)

        self.label_placeholder_note = QLabel(
            "Espacio reservado para botones de desglose y futuras acciones."
        )
        self.label_placeholder_note.setStyleSheet("color: gray; font-style: italic;")
        left_form.addWidget(self.label_placeholder_note, 7, 0, 1, 3)
        left_form.setRowStretch(8, 1)

        right_group = QGroupBox("Formato Vertical")
        right_layout = QVBoxLayout()
        right_group.setLayout(right_layout)

        self.manual_note_label = QLabel("Verde: valor manual (Doble clic para editar)")
        self.manual_note_label.setStyleSheet("color: #2e7d32; font-style: italic;")
        note_row = QHBoxLayout()
        note_row.addStretch()
        note_row.addWidget(self.manual_note_label)
        right_layout.addLayout(note_row)

        self.label_table_widget = QTableWidget()
        self._setup_label_table_widget()
        right_layout.addWidget(self.label_table_widget)

        export_layout = QHBoxLayout()
        export_layout.addStretch()
        export_layout.addWidget(QLabel("Exportar tabla como png:"))
        self.export_label_no_bg_button = QPushButton("Sin Fondo")
        self.export_label_with_bg_button = QPushButton("Con Fondo")
        export_layout.addWidget(self.export_label_no_bg_button)
        export_layout.addWidget(self.export_label_with_bg_button)
        export_layout.addStretch()
        right_layout.addLayout(export_layout)

        top_layout.addWidget(left_group, 1)
        top_layout.addWidget(right_group, 1)
        layout.addLayout(top_layout, 3)

        linear_group = QGroupBox("Formato Lineal")
        linear_layout = QVBoxLayout()
        linear_group.setLayout(linear_layout)
        self.linear_format_preview = QPlainTextEdit()
        self.linear_format_preview.setReadOnly(True)
        self.linear_format_preview.setStyleSheet(
            "background-color: #f5f5f5; color: #333;"
        )
        self.linear_format_preview.setMinimumHeight(140)
        linear_layout.addWidget(self.linear_format_preview)
        layout.addWidget(linear_group, 1)

        self.portion_size_input.valueChanged.connect(self._on_portion_value_changed)
        self.portion_unit_combo.currentTextChanged.connect(
            self._on_portion_unit_changed
        )
        self.household_unit_combo.currentTextChanged.connect(
            self._on_household_unit_changed
        )
        self.household_amount_input.textChanged.connect(
            self._on_household_amount_changed
        )
        self.custom_household_unit_input.textChanged.connect(
            self._on_household_unit_changed
        )
        self.export_label_no_bg_button.clicked.connect(
            lambda: self._on_export_label_table_clicked(with_background=False)
        )
        self.export_label_with_bg_button.clicked.connect(
            lambda: self._on_export_label_table_clicked(with_background=True)
        )
        self.label_table_widget.cellDoubleClicked.connect(
            self._on_label_table_cell_double_clicked
        )
        self._attach_copy_shortcut(self.label_table_widget)
        self.no_significant_display.installEventFilter(self)

        self._update_capacity_label()
        self._auto_fill_household_measure()
        self._update_label_preview(force_recalc_totals=True)

    def _set_default_column_widths(self, formulation: bool = False) -> None:
        """
        Set sensible initial column widths while keeping them resizable.
        """
        if not formulation:
            for table in (
                self.table,
                self.details_table,
                self.formulation_preview,
            ):
                header = table.horizontalHeader()
                header.setSectionResizeMode(QHeaderView.Interactive)

            self.table.setColumnWidth(0, 75)   # FDC ID
            self.table.setColumnWidth(1, 340)  # Descripcion
            self.table.setColumnWidth(2, 200)  # Marca / Origen
            self.table.setColumnWidth(3, 120)  # Tipo de dato

            self.details_table.setColumnWidth(0, 200)  # Nutriente
            self.details_table.setColumnWidth(1, 90)   # Cantidad
            self.details_table.setColumnWidth(2, 70)   # Unidad

            self.formulation_preview.setColumnWidth(0, 70)  # FDC ID
            self.formulation_preview.setColumnWidth(1, 290)  # Ingrediente
            return

        for table in (self.formulation_table, self.totals_table):
            header = table.horizontalHeader()
            header.setSectionResizeMode(QHeaderView.Interactive)

        self.formulation_table.setColumnWidth(0, 75)   # FDC ID
        self.formulation_table.setColumnWidth(1, 330)  # Ingrediente
        self.formulation_table.setColumnWidth(2, 95)   # Cantidad (g)
        self.formulation_table.setColumnWidth(3, 85)   # Cantidad (%)
        self.formulation_table.setColumnWidth(4, 65)   # Fijar %
        self.formulation_table.setColumnWidth(5, 150)  # Marca / Origen

        self.totals_table.setColumnWidth(0, 210)  # Nutriente
        self.totals_table.setColumnWidth(1, 85)   # Total
        self.totals_table.setColumnWidth(2, 60)   # Unidad
        self.totals_table.setColumnWidth(3, 80)   # Exportar

    def _setup_label_table_widget(self) -> None:
        table = self.label_table_widget
        self.label_table_title_row = 0
        self.label_table_portion_row = 1
        self.label_table_header_row = 2
        self.label_table_nutrient_start_row = 3
        self.label_table_footer_row = (
            self.label_table_nutrient_start_row + len(self.label_base_nutrients)
        )
        table.setColumnCount(3)
        table.setRowCount(self.label_table_footer_row + 2)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        table.setSelectionMode(QTableWidget.ExtendedSelection)
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.horizontalHeader().setVisible(False)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setMinimumHeight(360)
        table.setStyleSheet("gridline-color: #c0c0c0;")

    def _build_base_label_nutrients(self) -> list[dict[str, Any]]:
        return [
            {
                "name": "Energia",
                "type": "energy",
                "kcal": 0.0,
                "kj": 0.0,
                "vd": 13.0,
                "vd_reference": 263.0,
            },
            {
                "name": "Carbohidratos",
                "unit": "g",
                "amount": 0.0,
                "vd": 7.0,
                "vd_reference": 20.0,
            },
            {
                "name": "Proteinas",
                "unit": "g",
                "amount": 0.0,
                "vd": 16.0,
                "vd_reference": 12.0,
            },
            {
                "name": "Grasas totales",
                "unit": "g",
                "amount": 0.0,
                "vd": 27.0,
                "vd_reference": 15.0,
            },
            {
                "name": "Grasas saturadas",
                "unit": "g",
                "amount": 0.0,
                "vd": 23.0,
                "vd_reference": 5.0,
            },
            {
                "name": "Grasas trans",
                "unit": "g",
                "amount": 0.0,
                "vd": None,
                "vd_reference": 0.0,
            },
            {
                "name": "Fibra alimentaria",
                "unit": "g",
                "amount": 0.0,
                "vd": 12.0,
                "vd_reference": 3.0,
            },
            {
                "name": "Sodio",
                "unit": "mg",
                "amount": 0.0,
                "vd": 5.0,
                "vd_reference": 120.0,
            },
        ]

    def _build_household_measure_options(self) -> list[tuple[str, int | None]]:
        return [
            ("Taza de té", 200),
            ("Vaso", 200),
            ("Cuchara de sopa", 10),
            ("Cuchara de té", 5),
            ("Plato hondo", 250),
            ("Unidad", None),
            ("Otro", None),
        ]

    def _current_portion_factor(self) -> float:
        return float(self.portion_size_input.value() or 0) / 100.0

    def _format_fraction_amount(self, value: float) -> str:
        if value <= 0:
            return ""
        frac = Fraction(value).limit_denominator(12)
        whole, remainder = divmod(frac.numerator, frac.denominator)
        if remainder == 0:
            return str(whole)
        if whole == 0:
            return f"{remainder}/{frac.denominator}"
        return f"{whole} {remainder}/{frac.denominator}"

    def _update_capacity_label(self) -> None:
        unit_name = self.household_unit_combo.currentText()
        capacity = self.household_capacity_map.get(unit_name)
        if capacity:
            label_text = f"{capacity} ml"
        else:
            label_text = "-"
        if unit_name == "Otro" and self.custom_household_unit_input.isVisible():
            label_text = "Definir capacidad manualmente"
        self.household_capacity_label.setText(label_text)

    def _auto_fill_household_measure(self) -> None:
        if self.portion_unit_combo.currentText() != "ml":
            return
        unit_name = self.household_unit_combo.currentText()
        capacity = self.household_capacity_map.get(unit_name)
        if not capacity:
            return
        portion_value = float(self.portion_size_input.value() or 0)
        if portion_value <= 0:
            return
        ratio = portion_value / float(capacity)
        text = self._format_fraction_amount(ratio)
        self._auto_updating_household_amount = True
        try:
            self.household_amount_input.setText(text or "")
        finally:
            self._auto_updating_household_amount = False

    def _on_portion_value_changed(self, _: int) -> None:
        if self.portion_unit_combo.currentText() == "ml":
            self._auto_fill_household_measure()
        self._update_label_preview()

    def _on_portion_unit_changed(self, _: str) -> None:
        if self.portion_unit_combo.currentText() == "ml":
            self._auto_fill_household_measure()
        self._update_capacity_label()
        self._update_label_preview()

    def _on_household_unit_changed(self, _: str) -> None:
        is_custom = self.household_unit_combo.currentText() == "Otro"
        self.custom_household_unit_input.setVisible(is_custom)
        self._update_capacity_label()
        if self.portion_unit_combo.currentText() == "ml":
            self._auto_fill_household_measure()
        self._update_label_preview()

    def _on_household_amount_changed(self, _: str) -> None:
        if self._auto_updating_household_amount:
            return
        self._update_label_preview()

    def _current_household_unit_label(self) -> str:
        if self.household_unit_combo.currentText() == "Otro":
            custom = self.custom_household_unit_input.text().strip()
            return custom or "Unidad"
        return self.household_unit_combo.currentText()

    def _portion_description_for_table(self) -> str:
        measure_amount = self.household_amount_input.text().strip()
        measure_unit = self._current_household_unit_label()
        portion_unit = self.portion_unit_combo.currentText()
        portion_value = self.portion_size_input.value()
        measure_display = measure_unit if not measure_amount else f"{measure_amount} {measure_unit}"
        return f"Porción {portion_value} {portion_unit} ({measure_display})"

    def _format_number_for_unit(self, value: float, unit: str) -> str:
        if unit == "mg":
            return f"{value:.0f} mg"
        if unit == "g":
            if abs(value) < 10:
                return f"{value:.1f} g"
            return f"{value:.0f} g"
        if value >= 10:
            return f"{value:.0f} {unit}"
        if value >= 1:
            return f"{value:.1f} {unit}"
        return f"{value:.2f} {unit}"

    def _format_nutrient_amount(self, nutrient: Dict[str, Any], factor: float) -> str:
        if nutrient.get("type") == "energy":
            kcal_val = nutrient.get("kcal", 0.0) * factor
            kj_val = nutrient.get("kj", 0.0) * factor
            kcal_text = f"{kcal_val:.0f}" if kcal_val >= 10 else f"{kcal_val:.1f}"
            kj_text = f"{kj_val:.0f}" if kj_val >= 10 else f"{kj_val:.1f}"
            return f"{kcal_text} kcal = {kj_text} kJ"
        amount = nutrient.get("amount", 0.0) * factor
        unit = nutrient.get("unit", "")
        return self._format_number_for_unit(amount, unit)

    def _format_vd_value(self, nutrient: Dict[str, Any], factor: float, effective_amount: float | None = None) -> str:  # type: ignore[override]
        if nutrient.get("vd") is None:
            return "-"
        if nutrient.get("type") == "energy":
            base_amount = nutrient.get("vd_reference", nutrient.get("kcal", 0.0))
            eff_amount = effective_amount if effective_amount is not None else nutrient.get("kcal", 0.0)
        else:
            base_amount = nutrient.get("vd_reference", nutrient.get("amount", 0.0))
            eff_amount = effective_amount if effective_amount is not None else nutrient.get("amount", 0.0)
        portion_amount = eff_amount * factor
        if base_amount and base_amount > 0:
            vd_val = nutrient.get("vd", 0.0) * (portion_amount / base_amount)
        else:
            vd_val = nutrient.get("vd", 0.0) * factor
        if vd_val >= 10:
            return f"{vd_val:.0f}%"
        return f"{vd_val:.1f}%"

    def _format_manual_amount(self, nutrient: Dict[str, Any], manual_amount: float) -> str:
        if nutrient.get("type") == "energy":
            kcal_val = manual_amount
            kj_val = manual_amount * 4.184
            kcal_text = f"{kcal_val:.0f}" if kcal_val >= 10 else f"{kcal_val:.1f}"
            kj_text = f"{kj_val:.0f}" if kj_val >= 10 else f"{kj_val:.1f}"
            return f"{kcal_text} kcal = {kj_text} kJ"
        unit = nutrient.get("unit", "")
        return self._format_number_for_unit(manual_amount, unit)

    def _format_manual_vd(self, nutrient: Dict[str, Any], manual_amount: float) -> str:
        vd_ref = nutrient.get("vd")
        if vd_ref is None:
            return "-"
        base_amount = nutrient.get("kcal") if nutrient.get("type") == "energy" else nutrient.get("amount")
        if not base_amount:
            return "-"
        vd_val = vd_ref * (manual_amount / base_amount)
        if vd_val >= 10:
            return f"{vd_val:.0f}%"
        return f"{vd_val:.1f}%"

    def _parse_user_float(self, text: str) -> float | None:
        clean = text.strip().replace(",", ".")
        if not clean:
            return None
        try:
            return float(clean)
        except ValueError:
            return None

    def _on_label_table_cell_double_clicked(self, row: int, _: int) -> None:
        if (
            row < self.label_table_nutrient_start_row
            or row >= self.label_table_nutrient_start_row + len(self._label_display_nutrients)
        ):
            return
        idx = row - self.label_table_nutrient_start_row
        nutrient = self._label_display_nutrients[idx]
        self._prompt_manual_value_for_nutrient(nutrient)

    def _prompt_manual_value_for_nutrient(self, nutrient: Dict[str, Any]) -> None:
        name = nutrient.get("name", "")
        unit = "kcal" if nutrient.get("type") == "energy" else nutrient.get("unit", "")
        current = self.label_manual_overrides.get(name)
        default_text = "" if current is None else str(current)
        text, ok = QInputDialog.getText(
            self,
            "Valor manual",
            f"Ingrese la cantidad por porción para {name} ({unit}).\n"
            "Deje vacío para volver al cálculo automático.",
            text=default_text,
        )
        if not ok:
            return
        if text.strip() == "":
            self.label_manual_overrides.pop(name, None)
            self._update_label_preview()
            return
        value = self._parse_user_float(text)
        if value is None:
            QMessageBox.warning(self, "Valor inválido", "Ingresa un número válido (ej.: 12.5).")
            return
        self.label_manual_overrides[name] = value
        self._update_label_preview()

    def _eligible_no_significant(self) -> list[str]:
        eligible: list[str] = []
        factor = self._current_portion_factor()
        def portion_amount(name: str) -> float:
            for nutrient in self.label_base_nutrients:
                if nutrient.get("name") != name:
                    continue
                eff = self._effective_label_nutrient(nutrient)
                if eff.get("type") == "energy":
                    return (eff.get("kcal") or 0.0) * factor
                return (eff.get("amount") or 0.0) * factor
            return 0.0
        for nutrient in self.label_base_nutrients:
            eff = self._effective_label_nutrient(nutrient)
            name = eff.get("name", nutrient.get("name", ""))
            thresh = self.label_no_significant_thresholds.get(name)
            if not thresh:
                continue
            if eff.get("type") == "energy":
                kcal_portion = (eff.get("kcal") or 0.0) * factor
                kj_portion = (eff.get("kj") or 0.0) * factor
                if kcal_portion <= thresh.get("max", 0) or kj_portion < thresh.get("kj_max", 0):
                    eligible.append(name)
                continue
            amount_portion = (eff.get("amount") or 0.0) * factor
            unit = eff.get("unit", nutrient.get("unit", "")).lower()
            max_allowed = thresh.get("max", 0.0)
            if name == "Grasas totales":
                sat = portion_amount("Grasas saturadas")
                trans = portion_amount("Grasas trans")
                if (
                    amount_portion <= max_allowed + 1e-9
                    and sat <= self.label_no_significant_thresholds["Grasas saturadas"]["max"] + 1e-9
                    and trans <= self.label_no_significant_thresholds["Grasas trans"]["max"] + 1e-9
                ):
                    eligible.append(name)
            elif amount_portion <= max_allowed + 1e-9:
                eligible.append(name)
        return eligible

    def _update_no_significant_controls(self) -> None:
        eligible = self._eligible_no_significant()
        self.label_no_significant = self._sort_no_significant_list(
            [name for name in self.label_no_significant if name in eligible]
        )
        display_names = [
            self.label_no_significant_display_map.get(name, name)
            for name in self.label_no_significant
        ]
        self.no_significant_display.setText(", ".join(display_names))
        has_options = bool(eligible)
        self.no_significant_display.setEnabled(has_options)
        self.no_significant_display.setToolTip(
            ""
            if has_options
            else "No hay nutrientes elegibles con la porción y valores actuales."
        )

    def _on_select_no_significant_clicked(self) -> None:
        eligible = self._eligible_no_significant()
        if not eligible:
            QMessageBox.information(
                self,
                "Sin opciones",
                "No hay nutrientes elegibles para 'sin aportes significativos' con la porción actual.",
            )
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Sin aportes significativos")
        layout = QVBoxLayout(dialog)
        list_widget = QListWidget(dialog)
        prev_state_role = Qt.UserRole + 100
        for name in eligible:
            item = QListWidgetItem(self.label_no_significant_display_map.get(name, name))
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked if name in self.label_no_significant else Qt.Unchecked)
            item.setData(Qt.UserRole, name)
            item.setData(prev_state_role, item.checkState())
            list_widget.addItem(item)
        def _remember_state(item: QListWidgetItem) -> None:
            item.setData(prev_state_role, item.checkState())
        def _toggle_if_unchanged(item: QListWidgetItem) -> None:
            prev = item.data(prev_state_role)
            if prev is None:
                prev = item.checkState()
            if item.checkState() == prev:
                item.setCheckState(
                    Qt.Unchecked if item.checkState() == Qt.Checked else Qt.Checked
                )
            item.setData(prev_state_role, item.checkState())
        list_widget.itemPressed.connect(_remember_state)
        list_widget.itemClicked.connect(_toggle_if_unchanged)
        layout.addWidget(list_widget)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, parent=dialog)
        layout.addWidget(buttons)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)

        if dialog.exec() == QDialog.Accepted:
            selected: list[str] = []
            for idx in range(list_widget.count()):
                item = list_widget.item(idx)
                if item.checkState() == Qt.Checked:
                    selected.append(item.data(Qt.UserRole))
            if "Grasas totales" in selected:
                for dep in ("Grasas saturadas", "Grasas trans"):
                    if dep not in selected and dep in eligible:
                        selected.append(dep)
            self.label_no_significant = self._sort_no_significant_list(selected)
            self._update_label_preview()

    def eventFilter(self, obj: QObject, event) -> bool:
        if obj is self.no_significant_display and event.type() == QEvent.MouseButtonPress:
            if self.no_significant_display.isEnabled():
                self._on_select_no_significant_clicked()
            return True
        return super().eventFilter(obj, event)

    def _human_join(self, items: list[str]) -> str:
        if not items:
            return ""
        if len(items) == 1:
            return items[0]
        return ", ".join(items[:-1]) + " y " + items[-1]

    def _sort_no_significant_list(self, names: list[str]) -> list[str]:
        order_index = {name: idx for idx, name in enumerate(self.label_no_sig_order)}
        return sorted(
            names,
            key=lambda n: (
                order_index.get(n, len(order_index)),
                n.lower(),
            ),
        )

    def _parse_label_mapping(self, label_name: str) -> tuple[str, str]:
        mapped = self.label_nutrient_usda_map.get(label_name, "")
        if "(" in mapped and mapped.endswith(")"):
            base, unit_part = mapped.split("(", 1)
            unit = unit_part.rstrip(")").strip()
            return canonical_alias_name(base.strip()), canonical_unit(unit)
        return canonical_alias_name(mapped.strip()), ""

    def _find_total_entry(self, canonical_name: str, unit: str) -> Dict[str, Any] | None:
        if not self._last_totals:
            self._last_totals = self._calculate_totals()
        target = canonical_alias_name(canonical_name).lower()
        target_unit = canonical_unit(unit).lower()
        target_key = re.sub(r"[^a-z0-9]", "", target)
        entries = list(self._last_totals.values())

        def _match_entry(entry: Dict[str, Any]) -> bool:
            entry_name = canonical_alias_name(entry.get("name", "")).lower()
            entry_unit = canonical_unit(entry.get("unit", "")).lower()
            entry_key = re.sub(r"[^a-z0-9]", "", entry_name)
            name_match = (
                entry_name == target
                or entry_name.startswith(target)
                or target in entry_name
                or entry_key == target_key
                or entry_key.startswith(target_key)
                or target_key in entry_key
            )
            unit_match = (not target_unit) or entry_unit == target_unit
            return name_match and unit_match

        for entry in entries:
            if _match_entry(entry):
                return entry

        # Fallback: raw substring match (e.g., "total lipid (fat)")
        raw_target = canonical_name.lower()
        for entry in entries:
            raw_name = (entry.get("name") or "").lower()
            if raw_target in raw_name:
                if not target_unit or canonical_unit(entry.get("unit", "")).lower() == target_unit:
                    return entry
        return None

    def _factor_for_energy(self, name: str) -> float | None:
        factor_map: list[tuple[str, float]] = [
            ("alcohol", 7.0),
            ("ethanol", 7.0),
            ("protein", 4.0),
            ("carbohydrate", 4.0),
            ("carbohydrate, by difference", 4.0),
            ("polydextrose", 1.0),
            ("polyol", 2.4),
            ("sugar alcohol", 2.4),
            ("organic acid", 3.0),
            ("total lipid", 9.0),  # Solo lipidos totales, no Fat (NLEA)
        ]
        lower = name.lower()
        for key, factor in factor_map:
            if key in lower:
                return factor
        return None

    def _compute_energy_label_values(self) -> Dict[str, float] | None:
        if not self._last_totals:
            self._last_totals = self._calculate_totals()
        totals = self._last_totals or {}
        factor = self._current_portion_factor()
        if factor <= 0:
            factor = 1.0

        # Build a skip set for manual overrides (so we replace totals with manual)
        manual_names = {
            name for name in self.label_manual_overrides.keys() if name != "Energia"
        }

        kcal_portion = 0.0
        seen_keys: set[str] = set()

        # First pass: totals contributions (excluding manual overrides)
        for entry in totals.values():
            name = entry.get("name", "") or ""
            if name in manual_names:
                continue
            key = f"{canonical_alias_name(name).lower()}|{canonical_unit(entry.get('unit', '')).lower()}"
            if key in seen_keys:
                continue
            seen_keys.add(key)
            factor_energy = self._factor_for_energy(name)
            if factor_energy is None:
                continue
            unit = (entry.get("unit", "") or "").lower()
            amount = float(entry.get("amount", 0.0) or 0.0)
            amount_g = amount / 1000.0 if unit == "mg" else amount
            # totals are per 100 g producto final; convert to porción
            amount_portion = amount_g * self._current_portion_factor()
            kcal_portion += amount_portion * factor_energy

        # Second pass: manual overrides from la etiqueta base (por porción)
        for base in self.label_base_nutrients:
            name = base.get("name", "")
            if name == "Energia":
                continue
            if name not in manual_names:
                continue
            manual_amount = float(self.label_manual_overrides.get(name, 0.0) or 0.0)
            factor_energy = self._factor_for_energy(
                self.label_nutrient_usda_map.get(name, name)
            )
            if factor_energy is None:
                continue
            unit = (base.get("unit", "") or "").lower()
            amount_g = manual_amount / 1000.0 if unit == "mg" else manual_amount
            kcal_portion += amount_g * factor_energy

        if math.isclose(kcal_portion, 0.0, abs_tol=1e-6):
            return None

        kcal_per_100 = kcal_portion / factor
        return {"kcal": kcal_per_100, "kj": kcal_per_100 * 4.184}

    def _label_amount_from_totals(self, nutrient: Dict[str, Any]) -> Dict[str, float] | None:
        name = nutrient.get("name", "")
        mapped_name, mapped_unit = self._parse_label_mapping(name)
        if not mapped_name:
            return None
        if nutrient.get("type") == "energy":
            computed = self._compute_energy_label_values()
            if computed:
                return computed
            kcal_entry = self._find_total_entry(mapped_name, "kcal")
            kj_entry = self._find_total_entry(mapped_name, "kJ")
            if not kcal_entry and not kj_entry:
                return None
            kcal_amount = kcal_entry.get("amount") if kcal_entry else None
            if kcal_amount is None and kj_entry:
                kcal_amount = kj_entry.get("amount", 0.0) / 4.184
            kj_amount = kj_entry.get("amount") if kj_entry else None
            if kj_amount is None and kcal_entry:
                kj_amount = kcal_entry.get("amount", 0.0) * 4.184
            return {"kcal": float(kcal_amount or 0.0), "kj": float(kj_amount or 0.0)}
        entry = self._find_total_entry(mapped_name, mapped_unit)
        if not entry and name == "Grasas totales":
            # Fallback: cualquier nutriente que contenga "total fat" o "total lipid"
            for e in (self._last_totals or {}).values():
                raw_name = (e.get("name") or "").lower()
                if "total lipid" in raw_name:
                    entry = e
                    break
        if not entry:
            return None
        return {"amount": float(entry.get("amount", 0.0))}

    def _effective_label_nutrient(self, nutrient: Dict[str, Any]) -> Dict[str, Any]:
        name = nutrient.get("name", "")
        manual_amount = self.label_manual_overrides.get(name)
        if nutrient.get("type") == "energy":
            manual_amount = None
        totals_amount = self._label_amount_from_totals(nutrient)

        effective = dict(nutrient)
        effective["vd_reference"] = nutrient.get("vd_reference") or (
            nutrient.get("kcal", nutrient.get("amount", 0.0))
            if nutrient.get("type") == "energy"
            else nutrient.get("amount", 0.0)
        )

        if manual_amount is not None:
            if nutrient.get("type") == "energy":
                effective["kcal"] = manual_amount
                effective["kj"] = manual_amount * 4.184
                effective["amount"] = manual_amount
            else:
                effective["amount"] = manual_amount
            effective["manual"] = True
            return effective

        if totals_amount:
            if nutrient.get("type") == "energy":
                effective["kcal"] = totals_amount.get("kcal", nutrient.get("kcal", 0.0))
                effective["kj"] = totals_amount.get("kj", nutrient.get("kj", 0.0))
                effective["amount"] = effective["kcal"]
            else:
                effective["amount"] = totals_amount.get("amount", nutrient.get("amount", 0.0))
            effective["manual"] = False
            effective["from_totals"] = True
            return effective

        effective["manual"] = False
        effective["from_totals"] = False
        return effective

    def _update_label_table_preview(self) -> None:
        table = self.label_table_widget
        factor = self._current_portion_factor()
        table.clearSpans()

        display_nutrients: list[Dict[str, Any]] = []
        for nutrient in self.label_base_nutrients:
            if nutrient.get("name", "") in self.label_no_significant:
                continue
            display_nutrients.append(nutrient)
        self._label_display_nutrients = display_nutrients

        total_rows = 3 + len(display_nutrients) + (1 if self.label_no_significant else 0) + 1
        table.setRowCount(total_rows)

        title_item = QTableWidgetItem("INFORMACIÓN NUTRICIONAL")
        title_font = title_item.font()
        title_font.setBold(True)
        title_font.setPointSize(title_font.pointSize() + 1)
        title_item.setFont(title_font)
        title_item.setTextAlignment(Qt.AlignCenter)
        table.setItem(0, 0, title_item)

        portion_item = QTableWidgetItem(self._portion_description_for_table())
        portion_item.setTextAlignment(Qt.AlignCenter)
        table.setItem(1, 0, portion_item)

        header_font = QFont()
        header_font.setBold(True)
        header_item_amount = QTableWidgetItem("Cantidad por porción")
        header_item_amount.setFont(header_font)
        header_item_amount.setTextAlignment(Qt.AlignCenter)
        header_item_vd = QTableWidgetItem("% VD(*)")
        header_item_vd.setFont(header_font)
        header_item_vd.setTextAlignment(Qt.AlignCenter)
        table.setItem(2, 1, header_item_amount)
        table.setItem(2, 2, header_item_vd)

        for idx, nutrient in enumerate(display_nutrients):
            row = 3 + idx
            effective = self._effective_label_nutrient(nutrient)
            name = effective.get("name", nutrient.get("name", ""))
            name_item = QTableWidgetItem(name)

            if effective.get("manual"):
                manual_amount = self.label_manual_overrides.get(name, 0.0)
                amount_text = self._format_manual_amount(effective, manual_amount)
                vd_text = self._format_manual_vd(effective, manual_amount)
            else:
                amount_text = self._format_nutrient_amount(effective, factor)
                eff_amount = (
                    effective.get("kcal", 0.0)
                    if effective.get("type") == "energy"
                    else effective.get("amount", 0.0)
                )
                vd_text = self._format_vd_value(effective, factor, eff_amount)

            amount_item = QTableWidgetItem(amount_text)
            vd_item = QTableWidgetItem(vd_text)
            name_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            amount_item.setTextAlignment(Qt.AlignCenter)
            vd_item.setTextAlignment(Qt.AlignCenter)
            if effective.get("manual"):
                brush = QBrush(self.label_manual_hint_color)
                for itm in (name_item, amount_item, vd_item):
                    itm.setBackground(brush)
            table.setItem(row, 0, name_item)
            table.setItem(row, 1, amount_item)
            table.setItem(row, 2, vd_item)

        note_row = 3 + len(display_nutrients)
        if self.label_no_significant:
            names = [
                self.label_no_significant_display_map.get(name, name)
                for name in self._sort_no_significant_list(self.label_no_significant)
            ]
            note_text = (
                f"No aporta cantidades significativas de {self._human_join(names)}."
            )
            note_item = QTableWidgetItem(note_text)
            note_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            table.setItem(note_row, 0, note_item)
            table.setSpan(note_row, 0, 1, 3)
            footer_row = note_row + 1
        else:
            footer_row = note_row

        footer_text = (
            "*% Valores Diarios con base a una dieta de 2000 kcal u 8400 kJ. "
            "Sus valores diarios pueden ser mayores o menores dependiendo de sus necesidades energéticas."
        )
        footer_item = QTableWidgetItem(footer_text)
        footer_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        table.setItem(footer_row, 0, footer_item)
        table.setSpan(0, 0, 1, 3)
        table.setSpan(1, 0, 1, 3)
        table.setSpan(footer_row, 0, 1, 3)
        self.label_table_footer_row = footer_row
        table.resizeRowsToContents()

    def _update_linear_preview(self) -> None:
        factor = self._current_portion_factor()
        portion_desc = self._portion_description_for_table()
        parts: list[str] = []
        for nutrient in self.label_base_nutrients:
            if nutrient.get("name", "") in self.label_no_significant:
                continue
            effective = self._effective_label_nutrient(nutrient)
            if effective.get("manual"):
                manual_amount = self.label_manual_overrides.get(
                    effective.get("name", ""), 0.0
                )
                amount = self._format_manual_amount(effective, manual_amount)
                vd = self._format_manual_vd(effective, manual_amount)
            else:
                amount = self._format_nutrient_amount(effective, factor)
                eff_amount = (
                    effective.get("kcal", 0.0)
                    if effective.get("type") == "energy"
                    else effective.get("amount", 0.0)
                )
                vd = self._format_vd_value(effective, factor, eff_amount)
            vd_suffix = "" if vd in ("", "-") else f" ({vd} VD*)"
            parts.append(f"{nutrient.get('name', '')} {amount}{vd_suffix}")

        note_text = ""
        if self.label_no_significant:
            names = [
                self.label_no_significant_display_map.get(name, name)
                for name in self._sort_no_significant_list(self.label_no_significant)
            ]
            note_text = f" No aporta cantidades significativas de {self._human_join(names)}."

        base_text = (
            "Información Nutricional: "
            f"{portion_desc}. "
            + "; ".join(parts)
            + ";"
            + note_text
            + " % Valores Diarios con base a una dieta de 2000 kcal u 8400 kJ. "
            "Sus valores diarios pueden ser mayores o menores dependiendo de sus necesidades energéticas."
        )
        self.linear_format_preview.setPlainText(base_text)

    def _update_label_preview(self, force_recalc_totals: bool = False) -> None:
        if QThread.currentThread() is not self.thread():
            QTimer.singleShot(0, lambda: self._update_label_preview(force_recalc_totals))
            return
        if force_recalc_totals or not self._last_totals:
            self._last_totals = self._calculate_totals()
        self._update_no_significant_controls()
        self._update_label_table_preview()
        self._update_linear_preview()

    def _on_export_label_table_clicked(self, with_background: bool) -> None:
        fondo_text = "con fondo" if with_background else "sin fondo"
        QMessageBox.information(
            self,
            "Exportar tabla",
            f"La exportación a PNG ({fondo_text}) quedará disponible en el siguiente paso. "
            "Por ahora se preparó el diseño base solicitado.",
        )

    def _attach_copy_shortcut(self, table: QTableWidget) -> None:
        """Attach Ctrl+C to copy the current selection of a table as TSV to clipboard."""
        shortcut = QShortcut(QKeySequence.Copy, table)
        shortcut.setContext(Qt.WidgetWithChildrenShortcut)
        shortcut.activated.connect(lambda t=table: self._copy_table_selection(t))

    def _copy_table_selection(self, table: QTableWidget) -> None:
        """Copy selected cells/rows to clipboard (TSV) with headers."""
        sel_model = table.selectionModel()
        if not sel_model or not sel_model.hasSelection():
            return
        ranges = table.selectedRanges()
        if not ranges:
            return
        selected_range = ranges[0]
        rows = range(selected_range.topRow(), selected_range.bottomRow() + 1)
        cols = range(selected_range.leftColumn(), selected_range.rightColumn() + 1)

        headers: list[str] = []
        for col in cols:
            header_item = table.horizontalHeaderItem(col)
            headers.append(header_item.text() if header_item else "")
        lines = ["\t".join(headers)]

        for row in rows:
            row_vals: list[str] = []
            for col in cols:
                item = table.item(row, col)
                row_vals.append("" if item is None else item.text())
            lines.append("\t".join(row_vals))

        QApplication.clipboard().setText("\n".join(lines))

    def on_search_clicked(self) -> None:
        query = self.search_input.text().strip()
        if not query:
            self.status_label.setText("Ingresa un termino de busqueda.")
            return

        self.search_page = 1
        self.last_query = query
        self.last_include_brands = self.include_brands_checkbox.isChecked()
        self._start_search()

    def on_prev_page_clicked(self) -> None:
        if self.search_page <= 1:
            return
        self.search_page -= 1
        self._show_current_search_page()
        self._update_paging_buttons(len(self.search_results))

    def on_next_page_clicked(self) -> None:
        total = len(self.search_results)
        if self.search_page * self.search_page_size >= total:
            return
        self.search_page += 1
        self._show_current_search_page()
        self._update_paging_buttons(total)

    def on_include_brands_toggled(self) -> None:
        # Re-lanzar búsqueda con el filtro actualizado si ya hay texto
        if not self.search_input.text().strip():
            return
        self.search_page = 1
        self.last_query = self.search_input.text().strip()
        self.last_include_brands = self.include_brands_checkbox.isChecked()
        self._start_search()

    def _start_search(self) -> None:
        if not self.last_query:
            self.status_label.setText("Ingresa un termino de busqueda.")
            return

        self.search_button.setEnabled(False)
        self.prev_page_button.setEnabled(False)
        self.next_page_button.setEnabled(False)
        self.search_results = []
        self._last_results_count = 0
        self._populate_table([])  # clear while loading

        data_types = self._data_types_for_search()
        self.status_label.setText(
            f"Buscando en FoodData Central... (pagina {self.search_page})"
        )

        self._run_in_thread(
            fn=self._fetch_all_pages,
            args=(self.last_query, data_types),
            on_success=self._on_search_success,
            on_error=self._on_search_error,
        )

    def _data_types_for_search(self) -> List[str] | None:
        if self.last_include_brands:
            # None => sin filtro de tipos (trae todos)
            return None
        return ["Foundation", "SR Legacy"]

    def _fetch_all_pages(self, query: str, data_types: List[str] | None) -> List[Dict[str, Any]]:
        all_results: List[Dict[str, Any]] = []
        page = 1
        while page <= self.search_max_pages:
            batch = search_foods(
                query,
                page_size=self.search_fetch_page_size,
                data_types=data_types,
                page_number=page,
            )
            if not batch:
                break
            all_results.extend(batch)
            if len(batch) < self.search_fetch_page_size:
                break
            page += 1

        # Fallback: if no search results and the query looks like an FDC ID, try direct lookup.
        stripped = query.strip()
        if not all_results and stripped.isdigit():
            try:
                details = get_food_details(int(stripped), detail_format="abridged")
            except Exception:
                return all_results
            all_results.append(
                {
                    "fdcId": details.get("fdcId"),
                    "description": details.get("description", ""),
                    "brandOwner": details.get("brandOwner", "") or "",
                    "dataType": details.get("dataType", "") or "",
                }
            )
        return all_results

    def _sort_search_results(self, foods: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        def priority(data_type: str) -> int:
            return self.data_type_priority.get(
                data_type, self.data_type_priority.get(data_type.strip(), len(self.data_type_priority))
            )

        return sorted(
            foods,
            key=lambda f: (
                priority(f.get("dataType", "") or ""),
                (f.get("description", "") or "").lower(),
            ),
        )

    def _filter_results_by_query(self, foods: List[Dict[str, Any]], query: str) -> List[Dict[str, Any]]:
        tokens = [t for t in query.lower().split() if t]
        if not tokens:
            return foods

        filtered: List[Dict[str, Any]] = []
        for f in foods:
            haystack = f"{f.get('description', '')} {f.get('brandOwner', '')} {f.get('fdcId', '')}".lower()
            if all(tok in haystack for tok in tokens):
                filtered.append(f)
        return filtered

    def _prefetch_fdc_id(self, fdc_id: Any) -> None:
        """Warm USDA cache for a given FDC ID in background (no UI impact)."""
        try:
            fdc_int = int(fdc_id)
        except Exception:
            return
        if fdc_int in self._prefetching_fdc_ids:
            return
        self._prefetching_fdc_ids.add(fdc_int)
        logging.debug(f"Prefetching fdc_id={fdc_int}")

        def _on_done(_: object) -> None:
            self._prefetching_fdc_ids.discard(fdc_int)
            logging.debug(f"Prefetch done fdc_id={fdc_int}")

        self._run_in_thread(
            fn=lambda fid=fdc_int: get_food_details(
                fid,
                timeout=(3.05, 6.0),
                detail_format="abridged",
            ),
            args=(),
            on_success=_on_done,
            on_error=_on_done,
        )

    def _show_current_search_page(self) -> None:
        start = (self.search_page - 1) * self.search_page_size
        end = start + self.search_page_size
        slice_results = self.search_results[start:end]
        self._populate_table(slice_results, base_index=start)
        self._prefetch_visible_results(slice_results)
        total_pages = max(1, (len(self.search_results) + self.search_page_size - 1) // self.search_page_size)
        self.status_label.setText(
            f"Se encontraron {len(self.search_results)} resultados (pagina {self.search_page}/{total_pages})."
        )

    def _prefetch_visible_results(self, foods, limit: int = 2) -> None:
        """Proactively fetch details for the first results on screen to warm the cache."""
        count = 0
        for food in foods:
            fdc_id = food.get("fdcId")
            if fdc_id is None:
                continue
            self._prefetch_fdc_id(fdc_id)
            count += 1
            if count >= limit:
                break

    def _update_paging_buttons(self, count: int) -> None:
        has_query = bool(self.last_query)
        self.prev_page_button.setEnabled(has_query and self.search_page > 1)
        self.next_page_button.setEnabled(
            has_query and (self.search_page * self.search_page_size < count)
        )

    def on_fdc_search_clicked(self) -> None:
        fdc_id_text = self.fdc_id_input.text().strip()
        if not fdc_id_text.isdigit():
            self.status_label.setText("Ingresa un FDC ID numérico.")
            return

        self.status_label.setText(f"Cargando detalles de {fdc_id_text}...")
        self.fdc_id_button.setEnabled(False)
        self._run_in_thread(
            fn=get_food_details,
            args=(int(fdc_id_text),),
            on_success=self._on_details_success,
            on_error=self._on_details_error,
        )

    def on_result_double_clicked(self, row: int, _: int) -> None:
        """Double click on a search row -> add to formulation with quantity."""
        logging.debug(f"UI double click row={row}")
        self._add_row_to_formulation(row)

    def on_search_selection_changed(self) -> None:
        """Prefetch details for the selected search result to speed up adding."""
        indexes = self.table.selectionModel().selectedRows()
        if not indexes:
            return
        row = indexes[0].row()
        fdc_item = self.table.item(row, 0)
        if not fdc_item:
            return
        fdc_id_text = fdc_item.text().strip()
        self._prefetch_fdc_id(fdc_id_text)

    def on_add_selected_clicked(self) -> None:
        self._add_row_to_formulation()

    def on_remove_preview_clicked(self) -> None:
        self._remove_selected_from_formulation(self.formulation_preview)

    def on_remove_formulation_clicked(self) -> None:
        self._remove_selected_from_formulation(self.formulation_table)

    def on_formulation_preview_double_clicked(self, row: int, column: int) -> None:
        """Allow quick edit from the preview table."""
        if not self._can_edit_column(column):
            return
        self._edit_quantity_for_row(row)

    def on_preview_selection_changed(self) -> None:
        """Update nutrients panel when preview selection changes."""
        self._show_nutrients_for_selected_preview()

    def on_formulation_cell_double_clicked(self, row: int, column: int) -> None:
        """Double click on formulation row -> edit its quantity."""
        if not self._can_edit_column(column):
            return
        self._edit_quantity_for_row(row)

    def on_lock_toggled_from_table(self, item: QTableWidgetItem) -> None:
        """Handle lock/unlock toggles coming from any formulation table."""
        if item.column() != self.lock_column_index:
            return
        if self.quantity_mode == "g":
            return
        table = item.tableWidget()
        if table not in (self.formulation_table, self.formulation_preview):
            return
        row = item.row()
        if row < 0 or row >= len(self.formulation_items):
            return

        desired_locked = item.checkState() == Qt.Checked
        if desired_locked and self._locked_count(exclude_row=row) >= len(
            self.formulation_items
        ) - 1:
            # Avoid all items locked: keep one free.
            table.blockSignals(True)
            item.setCheckState(Qt.Unchecked)
            table.blockSignals(False)
            self.status_label.setText("Debe quedar al menos un ingrediente sin fijar.")
            return

        self.formulation_items[row]["locked"] = desired_locked
        self._refresh_formulation_views()

    def on_edit_quantity_clicked(self) -> None:
        indexes = self.formulation_table.selectionModel().selectedRows()
        if not indexes:
            self.status_label.setText("Selecciona un ingrediente para editar.")
            return
        self._edit_quantity_for_row(indexes[0].row())

    def on_quantity_mode_changed(self) -> None:
        """Switch between grams and percent modes for quantities."""
        self.quantity_mode = "g" if self.quantity_mode_selector.currentIndex() == 0 else "%"
        self._refresh_formulation_views()
        mode_text = "gramos (g)" if self.quantity_mode == "g" else "porcentaje (%)"
        self.status_label.setText(f"Modo de cantidad cambiado a {mode_text}.")

    def on_export_to_excel_clicked(self) -> None:
        """Export current formulation and totals to an Excel file."""
        if not self.formulation_items:
            QMessageBox.information(
                self,
                "Exportar a Excel",
                "No hay ingredientes en la formulacion para exportar.",
            )
            return

        default_name = f"{self._safe_base_name()}.xlsx"
        initial_path = (
            str(Path(self.last_path or "").with_name(default_name))
            if self.last_path
            else default_name
        )
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar formulacion a Excel",
            initial_path,
            "Archivos de Excel (*.xlsx)",
        )
        if not path:
            return

        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        self._save_last_path(path)
        try:
            self._export_formulation_to_excel(path)
        except Exception as exc:  # noqa: BLE001 - surface the error to the user
            QMessageBox.critical(
                self,
                "Error al exportar",
                f"No se pudo exportar el archivo:\n{exc}",
            )
        else:
            QMessageBox.information(
                self,
                "Exportado",
                f"Archivo guardado en:\n{path}",
            )

    def on_export_state_clicked(self) -> None:
        """Export formulation state (ingredientes + cantidades + flags) a JSON."""
        if not self.formulation_items:
            QMessageBox.information(
                self,
                "Exportar formulación",
                "No hay ingredientes en la formulación para exportar.",
            )
            return

        default_name = f"{self._safe_base_name()}.json"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar formulación",
            str(Path(self.last_path or "").with_name(default_name)) if self.last_path else default_name,
            "Archivos JSON (*.json)",
        )
        if not path:
            return
        if not path.lower().endswith(".json"):
            path += ".json"

        self._save_last_path(path)
        payload_items: list[Dict[str, Any]] = []
        for item in self.formulation_items:
            payload_items.append(
                {
                    "fdc_id": item.get("fdc_id"),
                    "description": item.get("description", ""),
                    "brand": item.get("brand", ""),
                    "data_type": item.get("data_type", ""),
                    "amount_g": float(item.get("amount_g", 0.0) or 0.0),
                    "locked": bool(item.get("locked", False)),
                }
            )

        payload = {
            "quantity_mode": self.quantity_mode,
            "items": payload_items,
            "nutrient_export_flags": self.nutrient_export_flags,
            "formula_name": self.formula_name_input.text(),
            "version": 2,
        }
        try:
            Path(path).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(
                self,
                "Error al exportar",
                f"No se pudo exportar la formulación:\n{exc}",
            )
            return

        self.status_label.setText(f"Formulación exportada en {path}")

    def on_import_state_clicked(self) -> None:
        """Import formulation state from JSON or Excel and refresh UI."""
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Importar formulación",
            self.last_path or "",
            "Archivos JSON (*.json);;Archivos Excel (*.xlsx)",
        )
        if not path:
            return
        self._save_last_path(path)

        ext = Path(path).suffix.lower()
        if ext == ".json":
            parsed = self._load_state_from_json(path)
        elif ext in (".xlsx", ".xls"):
            parsed = self._load_state_from_excel(path)
        else:
            QMessageBox.warning(
                self,
                "Formato no soportado",
                "Selecciona un archivo .json o .xlsx",
            )
            return

        if not parsed:
            return

        base_items, meta = parsed
        meta.setdefault("path", path)
        self._start_import_hydration(base_items, meta)

    def _load_state_from_json(self, path: str) -> tuple[list[Dict[str, Any]], Dict[str, Any]] | None:
        try:
            data = json.loads(Path(path).read_text(encoding="utf-8"))
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(
                self,
                "Error al importar",
                f"No se pudo leer el archivo:\n{exc}",
            )
            return None

        if not isinstance(data, dict):
            QMessageBox.warning(
                self,
                "Formato inválido",
                "El archivo no contiene una formulación válida.",
            )
            return None

        items = data.get("items") or []
        if not isinstance(items, list) or not items:
            QMessageBox.warning(
                self,
                "Formato inválido",
                "El archivo no contiene ingredientes válidos.",
            )
            return None

        base_items: list[Dict[str, Any]] = []
        for item in items:
            if not isinstance(item, dict):
                continue
            try:
                fdc_int = int(item.get("fdc_id") or item.get("fdcId"))
            except Exception:
                QMessageBox.warning(
                    self,
                    "FDC ID inválido",
                    f"FDC ID no numérico: {item.get('fdc_id') or item.get('fdcId')}",
                )
                return None
            base_items.append(
                {
                    "fdc_id": fdc_int,
                    "amount_g": float(item.get("amount_g", 0.0) or 0.0),
                    "locked": bool(item.get("locked", False)),
                    "description": item.get("description", ""),
                    "brand": item.get("brand", ""),
                    "data_type": item.get("data_type", ""),
                }
            )

        flags = data.get("nutrient_export_flags")
        nutrient_flags = (
            {k: bool(v) for k, v in flags.items()} if isinstance(flags, dict) else {}
        )

        mode = data.get("quantity_mode", "g")
        quantity_mode = "g" if mode != "%" else "%"

        meta = {
            "nutrient_export_flags": nutrient_flags,
            "quantity_mode": quantity_mode,
            "formula_name": data.get("formula_name", ""),
            "path": path,
            "respect_existing_formula_name": False,
        }
        return base_items, meta

    def _load_state_from_excel(self, path: str) -> tuple[list[Dict[str, Any]], Dict[str, Any]] | None:
        def _read(sheet: str | int, header_row: int) -> pd.DataFrame:
            return pd.read_excel(path, sheet_name=sheet, header=header_row)

        df: pd.DataFrame | None = None
        # Prefer sheet "Ingredientes" with headers on second row (row index 1)
        for sheet in ("Ingredientes", 0):
            for header_row in (1, 0):
                try:
                    tmp = _read(sheet, header_row)
                    if not tmp.empty:
                        df = tmp
                        break
                except Exception:
                    continue
            if df is not None:
                break

        if df is None or df.empty:
            QMessageBox.warning(
                self,
                "Sin datos",
                "El archivo no tiene filas para importar.",
            )
            return None

        # Normalize columns for matching
        cols_norm: Dict[str, str] = {self._normalize_label(c): c for c in df.columns}
        fdc_candidates = [
            "fdc id",
            "fdc_id",
            "fdcid",
            "fdc",
        ]
        amount_candidates = [
            "cantidad (g)",
            "cantidad g",
            "cantidad",
            "cantidad gramos",
            "cantidad en gramos",
            "amount g",
            "amount_g",
            "g",
            "grams",
        ]

        fdc_col = next((cols_norm[c] for c in fdc_candidates if c in cols_norm), None)
        amount_col = next(
            (cols_norm[c] for c in amount_candidates if c in cols_norm), None
        )
        if not fdc_col or not amount_col:
            QMessageBox.warning(
                self,
                "Columnas faltantes",
                "Se requieren columnas FDC ID y Cantidad (g).",
            )
            return None

        base_items: list[Dict[str, Any]] = []
        for _, row in df.iterrows():
            fdc_val = row.get(fdc_col)
            amt_val = row.get(amount_col)
            if pd.isna(fdc_val):
                continue
            try:
                fdc_int = int(fdc_val)
            except Exception:
                continue
            try:
                amt = float(amt_val) if not pd.isna(amt_val) else 0.0
            except Exception:
                amt = 0.0
            base_items.append(
                {
                    "fdc_id": fdc_int,
                    "amount_g": amt,
                    "locked": False,
                }
            )

        if not base_items:
            QMessageBox.warning(
                self,
                "Sin ingredientes",
                "No se encontraron filas válidas con FDC ID y Cantidad (g).",
            )
            return None

        meta = {
            "nutrient_export_flags": {},
            "quantity_mode": "g",
            "formula_name": self.formula_name_input.text() or Path(path).stem,
            "path": path,
            "respect_existing_formula_name": True,
        }
        return base_items, meta

    def _start_import_hydration(
        self, base_items: list[Dict[str, Any]], meta: Dict[str, Any]
    ) -> None:
        if not base_items:
            QMessageBox.warning(
                self,
                "Sin ingredientes",
                "El archivo no contiene ingredientes para importar.",
            )
            return

        self.import_state_button.setEnabled(False)
        self.export_state_button.setEnabled(False)
        self.status_label.setText("Importando ingredientes...")
        self._set_window_progress("Importando ingredientes")

        self._pending_import_meta = meta

        thread = QThread(self)
        worker = ImportWorker(
            base_items,
            max_attempts=self.import_max_attempts,
            read_timeout=self.import_read_timeout,
        )
        worker.moveToThread(thread)
        self._workers.append(worker)
        self._threads.append(thread)
        self._current_import_worker = worker

        thread.started.connect(worker.run)
        worker.progress.connect(self._on_import_progress)
        worker.finished.connect(self._on_import_finished)
        worker.error.connect(self._on_import_error)
        worker.finished.connect(thread.quit)
        worker.error.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        worker.error.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)

        def _cleanup() -> None:
            if thread in self._threads:
                self._threads.remove(thread)
            if worker in self._workers:
                self._workers.remove(worker)
            if self._current_import_worker is worker:
                self._current_import_worker = None

        thread.finished.connect(_cleanup)
        thread.start()

    def _on_import_progress(self, message: str) -> None:
        self._set_window_progress(message)
        self.status_label.setText(f"Importando ingredientes: {message}")

    def _on_import_finished(self, payload: list[Dict[str, Any]]) -> None:
        meta = getattr(self, "_pending_import_meta", {}) or {}
        self._pending_import_meta = {}
        if QThread.currentThread() is not self.thread():
            QTimer.singleShot(0, lambda p=payload, m=meta: self._on_import_finished(p))
            return
        self._reset_import_ui_state()
        hydrated: list[Dict[str, Any]] = []
        for entry in payload:
            base = entry.get("base") or {}
            details = entry.get("details") or {}
            fdc_id = base.get("fdc_id") or details.get("fdcId")
            try:
                fdc_id_int = int(fdc_id) if fdc_id is not None else None
            except Exception:
                fdc_id_int = fdc_id

            nutrients = augment_fat_nutrients(details.get("foodNutrients", []) or [])
            self._update_reference_from_details(details)
            hydrated.append(
                {
                    "fdc_id": fdc_id_int,
                    "description": details.get("description", "") or base.get("description", ""),
                    "brand": details.get("brandOwner", "") or base.get("brand", ""),
                    "data_type": details.get("dataType", "") or base.get("data_type", ""),
                    "amount_g": float(base.get("amount_g", 0.0) or 0.0),
                    "nutrients": normalize_nutrients(
                        nutrients, details.get("dataType")
                    ),
                    "locked": bool(base.get("locked", False)),
                }
            )

        self.formulation_items = hydrated
        self.nutrient_export_flags = meta.get("nutrient_export_flags", {})
        mode = meta.get("quantity_mode", "g")
        self.quantity_mode = "g" if mode != "%" else "%"
        self.quantity_mode_selector.blockSignals(True)
        self.quantity_mode_selector.setCurrentIndex(
            0 if self.quantity_mode == "g" else 1
        )
        self.quantity_mode_selector.blockSignals(False)

        formula_name = meta.get("formula_name", "")
        respect_existing = meta.get("respect_existing_formula_name", False)
        if not respect_existing or not self.formula_name_input.text().strip():
            self.formula_name_input.setText(formula_name)

        self._refresh_formulation_views()
        source = meta.get("path", "archivo")
        self.status_label.setText(f"Formulación importada desde {source}")

    def _on_import_error(self, message: str) -> None:
        self._reset_import_ui_state()
        self.status_label.setText("Error al importar ingredientes.")
        QMessageBox.critical(self, "Error al cargar ingrediente", message)

    def _reset_import_ui_state(self) -> None:
        self.import_state_button.setEnabled(True)
        self.export_state_button.setEnabled(True)
        self._set_window_progress(None)
        self._current_import_worker = None

    def _populate_table(self, foods, base_index: int = 0) -> None:
        self.table.setRowCount(0)

        for row_idx, food in enumerate(foods):
            self.table.insertRow(row_idx)

            fdc_id = str(food.get("fdcId", ""))
            description = food.get("description", "") or ""
            brand = food.get("brandOwner", "") or ""
            data_type = food.get("dataType", "") or ""

            self.table.setItem(row_idx, 0, QTableWidgetItem(fdc_id))
            self.table.setItem(row_idx, 1, QTableWidgetItem(description))
            self.table.setItem(row_idx, 2, QTableWidgetItem(brand))
            self.table.setItem(row_idx, 3, QTableWidgetItem(data_type))
            self.table.setVerticalHeaderItem(
                row_idx, QTableWidgetItem(str(base_index + row_idx + 1))
            )

    def _populate_details_table(self, nutrients) -> None:
        nutrients = self._sort_nutrients_for_display(augment_fat_nutrients(nutrients or []))
        self.details_table.setRowCount(0)

        row_idx = 0
        for n in nutrients:
            if n.get("amount") is None:
                continue
            self.details_table.insertRow(row_idx)
            nut = n.get("nutrient") or {}
            name = nut.get("name", "") or ""
            unit = nut.get("unitName", "") or ""
            amount = n.get("amount")
            amount_text = "" if amount is None else str(amount)

            self.details_table.setItem(row_idx, 0, QTableWidgetItem(name))
            self.details_table.setItem(row_idx, 1, QTableWidgetItem(amount_text))
            self.details_table.setItem(row_idx, 2, QTableWidgetItem(unit))
            row_idx += 1

    def _total_weight(self) -> float:
        """Total weight of current formulation in grams."""
        return sum((item.get("amount_g", 0) or 0) for item in self.formulation_items)

    def _locked_count(self, exclude_row: int | None = None) -> int:
        """How many items are locked, optionally ignoring one row."""
        count = 0
        for idx, item in enumerate(self.formulation_items):
            if exclude_row is not None and idx == exclude_row:
                continue
            if item.get("locked"):
                count += 1
        return count

    def _amount_to_percent(self, amount_g: float, total: float | None = None) -> float:
        total_weight = self._total_weight() if total is None else total
        if total_weight <= 0:
            return 0.0
        return (amount_g / total_weight) * 100.0

    def _update_quantity_headers(self) -> None:
        self.formulation_preview.setHorizontalHeaderLabels(
            ["FDC ID", "Ingrediente"]
        )
        self.formulation_table.setHorizontalHeaderLabels(
            [
                "FDC ID",
                "Ingrediente",
                "Cantidad (g)",
                "Cantidad (%)",
                "Fijar %",
                "Marca / Origen",
            ]
        )

    def _set_item_enabled(self, item: QTableWidgetItem | None, enabled: bool) -> None:
        if item is None:
            return
        flags = item.flags()
        if enabled:
            flags |= Qt.ItemIsEnabled
        else:
            flags &= ~Qt.ItemIsEnabled
        item.setFlags(flags)
        item.setBackground(QColor("#f0f0f0") if not enabled else QColor("white"))

    def _apply_column_state(self, table: QTableWidget, row: int) -> None:
        grams_enabled = self.quantity_mode == "g"
        percent_enabled = self.quantity_mode == "%"

        self._set_item_enabled(
            table.item(row, self.amount_g_column_index), grams_enabled
        )
        self._set_item_enabled(
            table.item(row, self.percent_column_index), percent_enabled
        )
        self._set_item_enabled(
            table.item(row, self.lock_column_index), percent_enabled
        )

    def _can_edit_column(self, column: int | None) -> bool:
        if column is None:
            return True
        if self.quantity_mode == "g":
            return column == self.amount_g_column_index
        return column == self.percent_column_index

    def _nutrients_by_header(
        self, nutrients: List[Dict[str, Any]], header_by_key: Dict[str, str]
    ) -> Dict[str, float]:
        """
        Build a mapping of template header -> nutrient amount (per 100 g),
        aligned to a precomputed header key map to avoid duplicate columns.
        """
        out: Dict[str, float] = {}
        best_priority: Dict[str, int] = {}
        allowed_keys = set(header_by_key.keys())
        alias_priority = {
            "carbohydrate, by difference": 2,
            "carbohydrate, by summation": 1,
            "carbohydrate by summation": 1,
            "sugars, total": 2,
            "total sugars": 1,
        }

        for entry in nutrients:
            amount = entry.get("amount")
            if amount is None:
                continue
            nut = entry.get("nutrient") or {}
            header_key, name, unit = self._header_key(nut)
            if not header_key or header_key not in allowed_keys:
                continue
            header = header_by_key[header_key]
            priority = alias_priority.get((nut.get("name") or "").strip().lower(), 0)
            current_best = best_priority.get(header, -1)
            if priority < current_best:
                continue

            best_priority[header] = priority
            out[header] = amount

        return out

    def _category_for_nutrient(self, name: str, nutrient: Dict[str, Any] | None = None) -> str:
        """Resolve a nutrient category using the static catalog then reference hints."""
        lower = (name or "").strip().lower()
        if lower in self._nutrient_category_map:
            return self._nutrient_category_map[lower]

        amino_acids = {
            "tryptophan",
            "threonine",
            "isoleucine",
            "leucine",
            "lysine",
            "methionine",
            "phenylalanine",
            "tyrosine",
            "valine",
            "arginine",
            "histidine",
            "alanine",
            "aspartic acid",
            "glutamic acid",
            "glycine",
            "proline",
            "serine",
            "hydroxyproline",
            "cysteine",
            "cystine",
        }
        organic_acids = {
            "citric acid",
            "malic acid",
            "oxalic acid",
            "quinic acid",
        }
        oligosaccharides = {"raffinose", "stachyose", "verbascose"}
        isoflavones = {"daidzein", "genistein", "daidzin", "genistin", "glycitin"}

        vitamin_like = (
            lower.startswith("vitamin ")
            or "tocopherol" in lower
            or "tocotrienol" in lower
            or "carotene" in lower
            or "lycopene" in lower
            or "lutein" in lower
            or "zeaxanthin" in lower
            or "retinol" in lower
            or "folate" in lower
            or "folic acid" in lower
            or "betaine" in lower
            or "choline" in lower
            or "caffeine" in lower
            or "theobromine" in lower
        )
        if vitamin_like:
            return "Vitamins and Other Components"

        if lower in amino_acids:
            return "Amino acids"
        if (
            "fatty acids" in lower
            or lower.startswith(("sfa ", "mufa ", "pufa "))
            or lower in {"cholesterol", "total lipid (fat)", "total fat (nlea)"}
        ):
            return "Lipids"
        if "sterol" in lower:
            return "Phytosterols"
        if lower in organic_acids or (lower.endswith("acid") and lower not in amino_acids):
            return "Organic acids"
        if lower in oligosaccharides:
            return "Oligosaccharides"
        if lower in isoflavones:
            return "Isoflavones"

        if nutrient:
            ref = self._reference_info(nutrient)
            if ref.get("category"):
                return ref["category"]
        return "Nutrientes"

    def _collect_nutrient_columns(self) -> tuple[list[str], Dict[str, str], Dict[str, str]]:
        """
        Collect ordered nutrient headers and their categories.
        Uses the static catalog to force a default ordering/grouping and only keeps
        nutrients that appear with a value in any ingredient.
        """
        candidates: Dict[str, Dict[str, Any]] = {}
        categories_seen_order: Dict[str, int] = {}
        preferred_order = [cat for cat, _ in self._nutrient_catalog]
        preferred_count = len(preferred_order)

        for item in self.formulation_items:
            data_priority = self.data_type_priority.get(
                (item.get("data_type") or "").strip(), len(self.data_type_priority)
            )
            for entry in self._sort_nutrients_for_display(item.get("nutrients", [])):
                nut = entry.get("nutrient") or {}
                amount = entry.get("amount")
                if amount is None:
                    continue
                header_key, canonical_name, canonical_unit = self._header_key(nut)
                if header_key and not self.nutrient_export_flags.get(header_key, True):
                    continue

                if not header_key or not canonical_name:
                    continue

                category = self._category_for_nutrient(canonical_name, nut)
                if category not in categories_seen_order:
                    categories_seen_order[category] = len(categories_seen_order)

                order = self._nutrient_order_map.get(canonical_name.strip().lower())
                if order is None:
                    order = self._nutrient_order(nut, len(candidates))
                # Keep kcal ahead of kJ when both present
                unit_lower = (canonical_unit or "").strip().lower()
                if canonical_name.strip().lower() == "energy":
                    if unit_lower == "kcal":
                        order = order - 0.1 if isinstance(order, (int, float)) else order
                    elif unit_lower == "kj":
                        order = order + 0.1 if isinstance(order, (int, float)) else order

                header = (
                    f"{canonical_name} ({canonical_unit})"
                    if canonical_unit
                    else canonical_name
                )

                existing = candidates.get(header_key)
                if existing is None or (
                    data_priority < existing["data_priority"]
                    or (
                        data_priority == existing["data_priority"]
                        and order < existing["order"]
                    )
                    or (
                        data_priority == existing["data_priority"]
                        and order == existing["order"]
                        and header < existing["header"]
                    )
                ):
                    candidates[header_key] = {
                        "header_key": header_key,
                        "header": header,
                        "category": category,
                        "order": order,
                        "data_priority": data_priority,
                    }

        def category_rank(cat: str) -> int:
            if cat in preferred_order:
                return preferred_order.index(cat)
            return preferred_count + categories_seen_order.get(cat, preferred_count)

        sorted_candidates = sorted(
            candidates.values(),
            key=lambda c: (
                category_rank(c["category"]),
                c["order"],
                c["header"].lower(),
            ),
        )

        ordered_headers: list[str] = [c["header"] for c in sorted_candidates]
        categories: Dict[str, str] = {c["header"]: c["category"] for c in sorted_candidates}
        header_key_map: Dict[str, str] = {c["header"]: c["header_key"] for c in sorted_candidates}

        return ordered_headers, categories, header_key_map

    def _load_last_path(self) -> str:
        try:
            data = json.loads(Path("last_path.json").read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return data.get("last_path", "")
        except Exception:
            return ""
        return ""

    def _save_last_path(self, path: str) -> None:
        try:
            Path("last_path.json").write_text(
                json.dumps({"last_path": str(Path(path).expanduser())}, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            self.last_path = str(Path(path).expanduser())
        except Exception:
            pass

    def _ensure_normalized_items(self) -> None:
        """Normalize all formulation_items in-place (fat + energy)."""
        for idx, item in enumerate(self.formulation_items):
            original = item.get("nutrients", []) or []
            normalized = normalize_nutrients(original, item.get("data_type"))
            if normalized != original:
                # preserve reference to allow downstream updates
                item["nutrients"] = normalized

    def _split_header_unit(self, header: str) -> tuple[str, str]:
        if header.endswith(")") and " (" in header:
            name, unit = header.rsplit(" (", 1)
            return name, unit[:-1]
        return header, ""

    def _hydrate_items(self, items: list[Dict[str, Any]]) -> list[Dict[str, Any]] | None:
        """Fetch USDA details for items to populate description/nutrients."""
        hydrated: list[Dict[str, Any]] = []
        total = len(items)
        for item in items:
            fdc_id = item.get("fdc_id") or item.get("fdcId")
            if fdc_id is None:
                QMessageBox.warning(
                    self,
                    "FDC ID faltante",
                    "Uno de los ingredientes no tiene FDC ID.",
                )
                return None
            try:
                fdc_id_int = int(fdc_id)
            except ValueError:
                QMessageBox.warning(
                    self,
                    "FDC ID inválido",
                    f"FDC ID no numérico: {fdc_id}",
                )
                return None

            self.status_label.setText(
                f"Importando ingrediente {len(hydrated)+1}/{total} (FDC {fdc_id_int})..."
            )
            QCoreApplication.processEvents()
            try:
                details = get_food_details(fdc_id_int)
            except Exception as exc:  # noqa: BLE001 - surface to user
                QMessageBox.critical(
                    self,
                    "Error al cargar ingrediente",
                    f"No se pudo cargar el FDC {fdc_id_int}:\n{exc}",
                )
                return None
            self._update_reference_from_details(details)

            hydrated.append(
                {
                    "fdc_id": fdc_id_int,
                    "description": details.get("description", "")
                    or item.get("description", ""),
                    "brand": details.get("brandOwner", "") or item.get("brand", ""),
                    "data_type": details.get("dataType", "") or item.get("data_type", ""),
                    "amount_g": float(item.get("amount_g", 0.0) or 0.0),
                    "nutrients": normalize_nutrients(
                        details.get("foodNutrients", []) or [], details.get("dataType")
                    ),
                    "locked": bool(item.get("locked", False)),
                }
            )

        return hydrated

    def _normalize_label(self, label: str) -> str:
        """Normalize column labels for loose matching (casefold + strip accents)."""
        if label is None:
            return ""
        text = str(label)
        text = unicodedata.normalize("NFKD", text)
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        text = text.replace("_", " ").replace("-", " ")
        return re.sub(r"\s+", " ", text).strip().lower()

    def _populate_formulation_tables(self) -> None:
        """Refresh formulation tables with current items."""
        logging.debug(f"_populate_formulation_tables rows={len(self.formulation_items)}")
        self._update_quantity_headers()
        total_weight = self._total_weight()

        # Left preview (ID + Ingrediente)
        self.formulation_preview.blockSignals(True)
        self.formulation_preview.setRowCount(len(self.formulation_items))
        for idx, item in enumerate(self.formulation_items):
            self.formulation_items[idx].setdefault("locked", False)
            self.formulation_preview.setItem(
                idx, 0, QTableWidgetItem(str(item.get("fdc_id", "")))
            )
            self.formulation_preview.setItem(
                idx, 1, QTableWidgetItem(item.get("description", ""))
            )
        self.formulation_preview.blockSignals(False)

        # Main formulation table
        self.formulation_table.blockSignals(True)
        self.formulation_table.setRowCount(len(self.formulation_items))
        for idx, item in enumerate(self.formulation_items):
            amount_g = float(item.get("amount_g", 0.0) or 0.0)
            percent = self._amount_to_percent(amount_g, total_weight)

            cells: list[QTableWidgetItem] = [
                QTableWidgetItem(str(item.get("fdc_id", ""))),
                QTableWidgetItem(item.get("description", "")),
                QTableWidgetItem(f"{amount_g:.1f}"),
                QTableWidgetItem(f"{percent:.2f}"),
            ]

            lock_item = QTableWidgetItem("")
            lock_item.setFlags(
                Qt.ItemIsSelectable | Qt.ItemIsUserCheckable | Qt.ItemIsEnabled
            )
            lock_item.setCheckState(
                Qt.Checked if item.get("locked") else Qt.Unchecked
            )
            cells.append(lock_item)

            cells.append(QTableWidgetItem(item.get("brand", "")))

            for col, cell in enumerate(cells):
                self.formulation_table.setItem(idx, col, cell)

            self._apply_column_state(self.formulation_table, idx)
        self.formulation_table.blockSignals(False)
        logging.debug("_populate_formulation_tables done")

    def _populate_totals_table(self) -> None:
        logging.debug("_populate_totals_table start")
        totals = self._calculate_totals()
        self._last_totals = totals

        category_order = [cat for cat, _ in self._nutrient_catalog]

        def _cat_rank(name: str) -> int:
            cat = self._category_for_nutrient(name)
            if cat in category_order:
                return category_order.index(cat)
            return len(category_order) + 1

        def _order_val(name: str) -> float:
            return float(self._nutrient_order_map.get(name.strip().lower(), float("inf")))

        sorted_totals = sorted(
            totals.items(),
            key=lambda item: (
                _cat_rank(item[1].get("name", "")),
                _order_val(item[1].get("name", "")),
                item[1].get("name", "").lower(),
            ),
        )

        self.totals_table.blockSignals(True)
        self.totals_table.setRowCount(len(sorted_totals))
        new_flags: Dict[str, bool] = {}

        for row_idx, (nut_key, entry) in enumerate(sorted_totals):
            name_item = QTableWidgetItem(entry["name"])
            name_item.setData(Qt.UserRole, nut_key)
            self.totals_table.setItem(row_idx, 0, name_item)

            self.totals_table.setItem(
                row_idx, 1, QTableWidgetItem(f"{entry['amount']:.2f}")
            )
            self.totals_table.setItem(row_idx, 2, QTableWidgetItem(entry["unit"]))

            export_item = QTableWidgetItem("")
            export_item.setFlags(
                Qt.ItemIsUserCheckable | Qt.ItemIsEnabled | Qt.ItemIsSelectable
            )
            current_checked = self.nutrient_export_flags.get(nut_key, True)
            export_item.setCheckState(Qt.Checked if current_checked else Qt.Unchecked)
            export_item.setData(Qt.UserRole, nut_key)
            self.totals_table.setItem(row_idx, 3, export_item)
            new_flags[nut_key] = current_checked

        self.nutrient_export_flags = new_flags
        self.totals_table.blockSignals(False)
        self._update_toggle_export_button()
        logging.debug("_populate_totals_table done")

    def _update_toggle_export_button(self) -> None:
        if not hasattr(self, "toggle_export_button"):
            return
        all_checked = self.nutrient_export_flags and all(
            self.nutrient_export_flags.values()
        )
        text = "Deseleccionar todos" if all_checked else "Seleccionar todos"
        self.toggle_export_button.setText(text)

    def _create_question_icon(self) -> QIcon:
        """Create a small yellow square icon with a '?' centered."""
        size = 16
        pixmap = QPixmap(size, size)
        pixmap.fill(Qt.transparent)
        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.fillRect(0, 0, size, size, QColor(255, 236, 179))
        painter.setPen(Qt.black)
        font: QFont = painter.font()
        font.setBold(True)
        font.setPointSize(10)
        painter.setFont(font)
        painter.drawText(pixmap.rect(), Qt.AlignCenter, "?")
        painter.end()
        return QIcon(pixmap)

    def _refresh_formulation_views(self) -> None:
        logging.debug(f"_refresh_formulation_views count={len(self.formulation_items)}")
        if QThread.currentThread() is not self.thread():
            QTimer.singleShot(0, self._refresh_formulation_views)
            return
        self._ensure_normalized_items()
        self._populate_formulation_tables()
        self._populate_totals_table()
        self._update_label_preview(force_recalc_totals=True)
        logging.debug("_refresh_formulation_views done")
        self._ensure_preview_selection()

    def _select_preview_row(self, row: int) -> None:
        if row < 0 or row >= self.formulation_preview.rowCount():
            return
        sel_model = self.formulation_preview.selectionModel()
        if sel_model is None:
            return
        sel_model.clearSelection()
        index = self.formulation_preview.model().index(row, 0)
        sel_model.select(
            index, QItemSelectionModel.ClearAndSelect | QItemSelectionModel.Rows
        )
        self.formulation_preview.setCurrentIndex(index)

    def _ensure_preview_selection(self) -> None:
        if not self.formulation_items:
            self.details_table.setRowCount(0)
            return
        self._ensure_normalized_items()
        sel_model = self.formulation_preview.selectionModel()
        has_sel = sel_model and sel_model.hasSelection()
        if not has_sel:
            self._select_preview_row(len(self.formulation_items) - 1)
        self._show_nutrients_for_selected_preview()

    def _show_nutrients_for_row(self, row: int) -> None:
        if row < 0 or row >= len(self.formulation_items):
            self.details_table.setRowCount(0)
            return
        nutrients = self.formulation_items[row].get("nutrients", []) or []
        self._populate_details_table(nutrients)

    def _show_nutrients_for_selected_preview(self) -> None:
        sel_model = self.formulation_preview.selectionModel()
        if not sel_model or not sel_model.hasSelection():
            self._show_nutrients_for_row(-1)
            return
        row = sel_model.selectedRows()[0].row()
        self._show_nutrients_for_row(row)

    def _export_formulation_to_excel(self, filepath: str) -> None:
        """Build Excel from scratch with nutrient categories as in USDA view."""
        self._ensure_normalized_items()
        wb = Workbook()
        ws = wb.active
        ws.title = "Ingredientes"
        totals_sheet = wb.create_sheet("Totales")

        base_headers = [
            "FDC ID",
            "Ingrediente",
            "Marca / Origen",
            "Tipo de dato",
            "Cantidad (g)",
            "Cantidad (%)",
        ]

        nutrient_headers, header_categories, header_key_map = self._collect_nutrient_columns()
        header_by_key = {v: k for k, v in header_key_map.items()}

        header_fill = PatternFill("solid", fgColor="D9D9D9")
        total_fill = PatternFill("solid", fgColor="FFF2CC")
        category_fills = {
            "Proximates": PatternFill("solid", fgColor="DAEEF3"),
            "Carbohydrates": PatternFill("solid", fgColor="E6B8B7"),
            "Minerals": PatternFill("solid", fgColor="C4D79B"),
            "Vitamins and Other Components": PatternFill("solid", fgColor="FFF2CC"),
            "Lipids": PatternFill("solid", fgColor="D9E1F2"),
            "Amino acids": PatternFill("solid", fgColor="E4DFEC"),
        }
        center = Alignment(horizontal="center", vertical="center")

        # Row 1: group titles (base + nutrient categories)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(base_headers))
        ws.cell(row=1, column=1, value="Detalles de formulación").alignment = center

        if nutrient_headers:
            start_col = len(base_headers) + 1
            col = start_col
            while col < start_col + len(nutrient_headers):
                idx = col - start_col
                category = header_categories.get(nutrient_headers[idx], "Nutrientes")
                run_start = col
                while (
                    col < start_col + len(nutrient_headers)
                    and header_categories.get(nutrient_headers[col - start_col], "Nutrientes")
                    == category
                ):
                    col += 1
                run_end = col - 1
                ws.merge_cells(start_row=1, start_column=run_start, end_row=1, end_column=run_end)
                cat_cell = ws.cell(row=1, column=run_start, value=category)
                cat_cell.alignment = center
                cat_cell.fill = category_fills.get(category, header_fill)

        # Row 2: headers
        headers = base_headers + nutrient_headers
        for col_idx, name in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=name)
            cell.fill = category_fills.get(
                header_categories.get(name, ""), header_fill
            ) if col_idx > len(base_headers) else header_fill
            cell.alignment = center

        start_row = 3
        grams_col = base_headers.index("Cantidad (g)") + 1
        percent_col = base_headers.index("Cantidad (%)") + 1
        data_rows = len(self.formulation_items)
        end_row = start_row + data_rows - 1

        # Write ingredient rows
        for idx, item in enumerate(self.formulation_items):
            row = start_row + idx
            values = [
                item.get("fdc_id", ""),
                item.get("description", ""),
                item.get("brand", ""),
                item.get("data_type", ""),
                float(item.get("amount_g", 0.0) or 0.0),
                None,  # placeholder for percent formula
            ]
            for col_idx, val in enumerate(values, start=1):
                ws.cell(row=row, column=col_idx, value=val)

            gram_cell = f"{get_column_letter(grams_col)}{row}"
            total_range = f"${get_column_letter(grams_col)}${start_row}:${get_column_letter(grams_col)}${end_row}"
            ws.cell(
                row=row,
                column=percent_col,
                value=f"={gram_cell}/SUM({total_range})",
            ).number_format = "0.00%"

            nut_map = self._nutrients_by_header(item.get("nutrients", []), header_by_key)
            for offset, header in enumerate(nutrient_headers, start=len(base_headers) + 1):
                if header in nut_map:
                    ws.cell(row=row, column=offset, value=nut_map[header])

        total_row = end_row + 1
        ws.cell(row=total_row, column=1, value="Total")
        ws.cell(row=total_row, column=2, value="Formulado")
        ws.cell(row=total_row, column=3, value="Formulado")
        ws.cell(row=total_row, column=4, value="Formulado")

        gram_total_cell = ws.cell(
            row=total_row,
            column=grams_col,
            value=f"=SUBTOTAL(9,{get_column_letter(grams_col)}{start_row}:{get_column_letter(grams_col)}{end_row})",
        )
        gram_total_cell.fill = total_fill
        percent_total_cell = ws.cell(row=total_row, column=percent_col, value="100%")
        percent_total_cell.number_format = "0.00%"
        percent_total_cell.fill = total_fill

        for offset, header in enumerate(nutrient_headers, start=len(base_headers) + 1):
            col_letter = get_column_letter(offset)
            formula = (
                f"=SUMPRODUCT(${get_column_letter(percent_col)}${start_row}:${get_column_letter(percent_col)}${end_row},"
                f"${col_letter}${start_row}:${col_letter}${end_row})"
            )
            cell = ws.cell(row=total_row, column=offset, value=formula)
            cell.fill = total_fill

        # Styles for data rows
        for row in range(start_row, total_row + 1):
            ws.cell(row=row, column=grams_col).number_format = "0.0"

        # Freeze panes to keep headers/base columns visible
        freeze_col = len(base_headers) + 1 if nutrient_headers else 1
        ws.freeze_panes = f"{get_column_letter(freeze_col)}3"

        # Adjust widths
        widths = {
            "A": 12,
            "B": 35,
            "C": 18,
            "D": 14,
            "E": 12,
            "F": 12,
        }
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

        # Totales sheet: simple reference to totals row
        totals_headers = ["Nutriente", "Total", "Unidad"]
        for col_idx, name in enumerate(totals_headers, start=1):
            cell = totals_sheet.cell(row=1, column=col_idx, value=name)
            cell.fill = header_fill
            cell.alignment = center

        for idx, header in enumerate(nutrient_headers, start=1):
            name_part, unit = self._split_header_unit(header)
            totals_sheet.cell(row=idx + 1, column=1, value=name_part)
            totals_sheet.cell(row=idx + 1, column=3, value=unit or "")
            source_cell = f"Ingredientes!{get_column_letter(len(base_headers) + idx)}{total_row}"
            totals_sheet.cell(row=idx + 1, column=2, value=f"={source_cell}")

        wb.save(filepath)

    def _calculate_totals(self) -> Dict[str, Dict[str, Any]]:
        """
        Sum nutrients across all ingredients and normalize to 100 g de producto final.
        Assumes nutrient amounts from the API are per 100 g of ingredient.
        Groups by nutrient id/number to merge Foundation + SR Legacy entries.
        """
        logging.debug(f"_calculate_totals start items={len(self.formulation_items)}")
        self._ensure_normalized_items()
        totals: Dict[str, Dict[str, Any]] = {}
        total_weight = self._total_weight()
        for item in self.formulation_items:
            qty = item.get("amount_g", 0) or 0
            for nutrient in self._sort_nutrients_for_display(item.get("nutrients", [])):
                amount = nutrient.get("amount")
                if amount is None:
                    continue
                nut = nutrient.get("nutrient") or {}
                header_key, canonical_name, canonical_unit = self._header_key(nut)
                if not header_key:
                    continue
                entry = totals.setdefault(
                    header_key,
                    {
                        "name": canonical_name or nut.get("name", ""),
                        "unit": canonical_unit or "",
                        "amount": 0.0,
                        "order": self._nutrient_order(nut, len(totals)),
                    },
                )
                if canonical_name and not entry["name"]:
                    entry["name"] = canonical_name
                if canonical_unit and not entry["unit"]:
                    entry["unit"] = canonical_unit
                inferred_unit = self._infer_unit(nut)
                if inferred_unit and not entry["unit"]:
                    entry["unit"] = inferred_unit
                entry["order"] = min(
                    entry.get("order", float("inf")),
                    self._nutrient_order(nut, len(totals)),
                )
                entry["amount"] += amount * qty / 100.0

        if total_weight > 0:
            factor = 100.0 / total_weight
            for entry in totals.values():
                entry["amount"] *= factor
        logging.debug(f"_calculate_totals done nutrients={len(totals)} total_weight={total_weight}")
        return totals

    def _nutrient_key(self, nutrient: Dict[str, Any]) -> str:
        """
        Build a consistent key for nutrients, preferring id then number then name.
        This avoids duplicates when some records lack unitName (Foundation vs SR).
        """
        name_lower = (nutrient.get("name") or "").strip().lower()
        unit_lower = (nutrient.get("unitName") or "").strip().lower()
        # Special-case Energy to keep kcal/kJ separated when ids/numbers are missing.
        if name_lower == "energy" and unit_lower:
            return f"energy:{unit_lower}"
        # Special-case Water to merge branded-calculated (no id) with USDA water (id present).
        if name_lower == "water":
            return f"water|{unit_lower}"
        if "id" in nutrient and nutrient["id"] is not None:
            return f"id:{nutrient['id']}"
        if nutrient.get("number"):
            return f"num:{nutrient['number']}"
        name = name_lower
        return f"name:{name}" if name else ""

    def _reference_info(self, nutrient: Dict[str, Any]) -> Dict[str, Any]:
        """Return cached rank/category info for a nutrient key if available."""
        key = self._nutrient_key(nutrient)
        return self._reference_order_map.get(key, {})

    def _build_nutrient_catalog(self) -> list[tuple[str, list[str]]]:
        """Static catalog to enforce ordering/categories when using abridged."""
        return [
            (
                "Proximates",
                [
                    "Water",
                    "Energy",
                    "Nitrogen",
                    "Protein",
                    "Total fat (NLEA)",
                    "Total lipid (fat)",
                    "Ash",
                    "Carbohydrate, by difference",
                ],
            ),
            (
                "Carbohydrates",
                [
                    "Fiber, total dietary",
                    "Fiber, soluble",
                    "Fiber, insoluble",
                    "Total dietary fiber (AOAC 2011.25)",
                    "High Molecular Weight Dietary Fiber (HMWDF)",
                    "Low Molecular Weight Dietary Fiber (LMWDF)",
                    "Sugars, Total",
                    "Sucrose",
                    "Glucose",
                    "Fructose",
                    "Lactose",
                    "Maltose",
                    "Galactose",
                    "Starch",
                    "Resistant starch",
                    "Sugars, added",
                ],
            ),
            (
                "Minerals",
                [
                    "Calcium, Ca",
                    "Iron, Fe",
                    "Magnesium, Mg",
                    "Phosphorus, P",
                    "Potassium, K",
                    "Sodium, Na",
                    "Zinc, Zn",
                    "Copper, Cu",
                    "Manganese, Mn",
                    "Iodine, I",
                    "Selenium, Se",
                    "Molybdenum, Mo",
                    "Fluoride, F",
                ],
            ),
            (
                "Vitamins and Other Components",
                [
                    "Thiamin",
                    "Riboflavin",
                    "Niacin",
                    "Vitamin B-6",
                    "Folate, total",
                    "Folic acid",
                    "Folate, DFE",
                    "Choline, total",
                    "Choline, free",
                    "Choline, from phosphocholine",
                    "Choline, from phosphatidyl choline",
                    "Choline, from glycerophosphocholine",
                    "Choline, from sphingomyelin",
                    "Betaine",
                    "Vitamin B-12",
                    "Vitamin B-12, added",
                    "Vitamin A, RAE",
                    "Retinol",
                    "Carotene, beta",
                    "cis-beta-Carotene",
                    "trans-beta-Carotene",
                    "Carotene, alpha",
                    "Carotene, gamma",
                    "Cryptoxanthin, beta",
                    "Cryptoxanthin, alpha",
                    "Vitamin A, IU",
                    "Lycopene",
                    "cis-Lycopene",
                    "trans-Lycopene",
                    "Lutein + zeaxanthin",
                    "cis-Lutein/Zeaxanthin",
                    "Lutein",
                    "Zeaxanthin",
                    "Phytoene",
                    "Phytofluene",
                    "Vitamin D (D2 + D3), International Units",
                    "Vitamin D (D2 + D3)",
                    "Vitamin D2 (ergocalciferol)",
                    "Vitamin D3 (cholecalciferol)",
                    "25-hydroxycholecalciferol",
                    "Vitamin K (phylloquinone)",
                    "Vitamin K (Dihydrophylloquinone)",
                    "Vitamin K (Menaquinone-4)",
                    "Vitamin E (alpha-tocopherol)",
                    "Vitamin E, added",
                    "Tocopherol, beta",
                    "Tocopherol, gamma",
                    "Tocopherol, delta",
                    "Tocotrienol, alpha",
                    "Tocotrienol, beta",
                    "Tocotrienol, gamma",
                    "Tocotrienol, delta",
                    "Vitamin C, total ascorbic acid",
                    "Pantothenic acid",
                    "Biotin",
                    "Caffeine",
                    "Theobromine",
                ],
            ),
            (
                "Lipids",
                [
                    "Fatty acids, total saturated",
                    "SFA 4:0",
                    "SFA 5:0",
                    "SFA 6:0",
                    "SFA 7:0",
                    "SFA 8:0",
                    "SFA 9:0",
                    "SFA 10:0",
                    "SFA 11:0",
                    "SFA 12:0",
                    "SFA 13:0",
                    "SFA 14:0",
                    "SFA 15:0",
                    "SFA 16:0",
                    "SFA 17:0",
                    "SFA 18:0",
                    "SFA 20:0",
                    "SFA 21:0",
                    "SFA 22:0",
                    "SFA 23:0",
                    "SFA 24:0",
                    "Fatty acids, total monounsaturated",
                    "MUFA 12:1",
                    "MUFA 14:1",
                    "MUFA 14:1 c",
                    "MUFA 15:1",
                    "MUFA 16:1",
                    "MUFA 16:1 c",
                    "MUFA 17:1",
                    "MUFA 17:1 c",
                    "MUFA 18:1",
                    "MUFA 18:1 c",
                    "MUFA 20:1",
                    "MUFA 20:1 c",
                    "MUFA 22:1",
                    "MUFA 22:1 c",
                    "MUFA 22:1 n-9",
                    "MUFA 22:1 n-11",
                    "MUFA 24:1 c",
                    "Fatty acids, total polyunsaturated",
                    "PUFA 18:2",
                    "PUFA 18:2 c",
                    "PUFA 18:2 n-6 c,c",
                    "PUFA 18:2 CLAs",
                    "PUFA 18:2 i",
                    "PUFA 18:3",
                    "PUFA 18:3 c",
                    "PUFA 18:3 n-3 c,c,c (ALA)",
                    "PUFA 18:3 n-6 c,c,c",
                    "PUFA 18:4",
                    "PUFA 20:2 c",
                    "PUFA 20:2 n-6 c,c",
                    "PUFA 20:3",
                    "PUFA 20:3 c",
                    "PUFA 20:3 n-3",
                    "PUFA 20:3 n-6",
                    "PUFA 20:3 n-9",
                    "PUFA 20:4",
                    "PUFA 20:4c",
                    "PUFA 20:5c",
                    "PUFA 20:5 n-3 (EPA)",
                    "PUFA 22:2",
                    "PUFA 22:3",
                    "PUFA 22:4",
                    "PUFA 22:5 c",
                    "PUFA 22:5 n-3 (DPA)",
                    "PUFA 22:6 c",
                    "PUFA 22:6 n-3 (DHA)",
                    "Fatty acids, total trans",
                    "Fatty acids, total trans-monoenoic",
                    "Fatty acids, total trans-dienoic",
                    "Fatty acids, total trans-polyenoic",
                    "TFA 14:1 t",
                    "TFA 16:1 t",
                    "TFA 18:1 t",
                    "TFA 18:2 t",
                    "TFA 18:2 t,t",
                    "TFA 18:2 t not further defined",
                    "TFA 18:3 t",
                    "TFA 20:1 t",
                    "TFA 22:1 t",
                    "Cholesterol",
                ],
            ),
            (
                "Amino acids",
                [
                    "Tryptophan",
                    "Threonine",
                    "Isoleucine",
                    "Leucine",
                    "Lysine",
                    "Methionine",
                    "Phenylalanine",
                    "Tyrosine",
                    "Valine",
                    "Arginine",
                    "Histidine",
                    "Alanine",
                    "Aspartic acid",
                    "Glutamic acid",
                    "Glycine",
                    "Proline",
                    "Serine",
                    "Hydroxyproline",
                    "Cysteine",
                ],
            ),
            (
                "Phytosterols",
                [
                    "Phytosterols",
                    "Beta-sitosterol",
                    "Brassicasterol",
                    "Campesterol",
                    "Campestanol",
                    "Delta-5-avenasterol",
                    "Phytosterols, other",
                    "Stigmasterol",
                    "Beta-sitostanol",
                ],
            ),
            ("Organic acids", ["Citric acid", "Malic acid", "Oxalic acid", "Quinic acid"]),
            ("Oligosaccharides", ["Verbascose", "Raffinose", "Stachyose"]),
            ("Isoflavones", ["Daidzin", "Genistin", "Glycitin", "Daidzein", "Genistein"]),
        ]

    def _update_reference_from_details(self, details: Dict[str, Any]) -> None:
        """Update reference rank/category map from a full USDA response."""
        nutrients = details.get("foodNutrients", []) or []
        if not nutrients:
            return
        current_category: str | None = None
        for entry in nutrients:
            nut = entry.get("nutrient") or {}
            key = self._nutrient_key(nut)
            if not key:
                continue
            if entry.get("amount") is None:
                # category row
                current_category = (nut.get("name") or "").strip() or current_category
                self._reference_order_map.setdefault(
                    key,
                    {
                        "rank": nut.get("rank"),
                        "category": current_category,
                        "unit": nut.get("unitName"),
                    },
                )
                continue
            self._reference_order_map[key] = {
                "rank": nut.get("rank"),
                "category": current_category,
                "unit": nut.get("unitName"),
            }

    def _sort_nutrients_for_display(self, nutrients: list[Dict[str, Any]]) -> list[Dict[str, Any]]:
        """Return nutrients ordered by USDA rank or reference map."""
        if not nutrients:
            return []
        indexed = []
        for idx, entry in enumerate(nutrients):
            nut = entry.get("nutrient") or {}
            order = self._nutrient_order(nut, idx + 10000)
            indexed.append((order, idx, entry))
        indexed.sort(key=lambda t: (t[0], t[1]))
        return [item[2] for item in indexed]

    def _header_key(self, nutrient: Dict[str, Any]) -> tuple[str, str, str]:
        """Return a stable header key plus canonical name and unit for a nutrient."""
        name = canonical_alias_name(nutrient.get("name", "") or "")
        unit = canonical_unit(nutrient.get("unitName") or self._infer_unit(nutrient) or "")
        unit_part = unit.strip().lower()
        name_part = name.strip().lower()
        if name_part:
            header_key = f"{name_part}|{unit_part}"
        else:
            base_key = self._nutrient_key(nutrient)
            if not base_key:
                return "", name, unit
            header_key = f"{base_key}|{unit_part}"
        return header_key, name, unit

    def _infer_unit(self, nutrient: Dict[str, Any]) -> str:
        """Try to fill missing unit from nutrient metadata or heuristic defaults."""
        unit = nutrient.get("unitName")
        if unit:
            return unit

        number = str(nutrient.get("number") or "").strip()
        name = (nutrient.get("name") or "").lower()

        default_units_by_number = {
            # Proximates / macros
            "255": "g",  # Water
            "203": "g",  # Protein
            "204": "g",  # Total lipid (fat)
            "298": "g",  # Total fat (NLEA)
            "202": "g",  # Nitrogen
            "207": "g",  # Ash
            "205": "g",  # Carbohydrate, by difference
            "291": "g",  # Fiber, total dietary
            "269": "g",  # Sugars, total
            "268": "kJ",  # Energy (kJ)
            "208": "kcal",  # Energy (kcal)
            "951": "g",  # Proximates
            "956": "g",  # Carbohydrates
        }
        if number in default_units_by_number:
            return default_units_by_number[number]

        if "energy" in name and "kcal" in name:
            return "kcal"
        if "energy" in name and "kj" in name:
            return "kJ"

        macro_hints = [
            "water",
            "protein",
            "lipid",
            "fat",
            "ash",
            "carbohydrate",
            "fiber",
            "sugar",
            "starch",
            "nitrogen",
            "fatty acids",
            "sfa",
            "mufa",
            "pufa",
        ]
        if any(hint in name for hint in macro_hints) or ":" in name:
            return "g"

        amino_acids = [
            "alanine",
            "arginine",
            "aspartic acid",
            "cystine",
            "cysteine",
            "hydroxyproline",
            "glutamic acid",
            "glycine",
            "histidine",
            "isoleucine",
            "leucine",
            "lysine",
            "methionine",
            "phenylalanine",
            "proline",
            "serine",
            "threonine",
            "tryptophan",
            "tyrosine",
            "valine",
        ]
        if name in amino_acids:
            return "g"

        simple_sugars = [
            "sucrose",
            "glucose",
            "fructose",
            "lactose",
            "maltose",
            "galactose",
        ]
        if name in simple_sugars:
            return "g"

        if name == "alcohol, ethyl":
            return "g"

        return ""

    def _nutrient_order(self, nutrient: Dict[str, Any], fallback: int) -> float:
        """Return USDA rank if available; otherwise keep insertion order fallback."""
        rank = nutrient.get("rank")
        if rank is None:
            ref = self._reference_info(nutrient)
            rank = ref.get("rank")
        if rank is None:
            name_lower = (nutrient.get("name") or "").strip().lower()
            rank = self._nutrient_order_map.get(name_lower)
        try:
            return float(rank)
        except (TypeError, ValueError):
            return float(fallback)

    def _add_row_to_formulation(self, row: int | None = None) -> None:
        logging.debug(f"_add_row_to_formulation row={row}")
        if row is None:
            indexes = self.table.selectionModel().selectedRows()
            if not indexes:
                self.status_label.setText("Selecciona un alimento para agregar.")
                return
            row = indexes[0].row()

        fdc_item = self.table.item(row, 0)
        if not fdc_item:
            return
        fdc_id_text = fdc_item.text().strip()
        if not fdc_id_text or not fdc_id_text.isdigit():
            self.status_label.setText("FDC ID inválido.")
            return

        mode, value = self._prompt_quantity()
        logging.debug(f"_prompt_quantity returned mode={mode} value={value}")
        if mode is None:
            return

        display_amount = (
            self._format_amount_for_status(value, include_new=True)
            if mode == "g"
            else f"{value:.2f} %"
        )
        self.status_label.setText(
            f"Agregando {fdc_id_text} ({display_amount})..."
        )
        self.add_button.setEnabled(False)
        self._start_add_fetch(int(fdc_id_text), mode, value)

    def _start_add_fetch(self, fdc_id: int, mode: str, value: float) -> None:
        logging.debug(f"_start_add_fetch fdc_id={fdc_id} mode={mode} value={value}")
        self._set_window_progress(f"1/1 ID #{fdc_id}")
        thread = QThread(self)
        worker = AddWorker(
            fdc_id,
            max_attempts=self.import_max_attempts,
            read_timeout=self.import_read_timeout,
            mode=mode,
            value=value,
        )
        worker.moveToThread(thread)
        self._workers.append(worker)
        self._threads.append(thread)
        self._current_add_worker = worker

        thread.started.connect(worker.run)
        worker.progress.connect(self._on_add_progress)
        worker.finished.connect(self._on_add_finished)
        worker.error.connect(self._on_add_error)
        worker.finished.connect(thread.quit)
        worker.error.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        worker.error.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)

        def _cleanup() -> None:
            if thread in self._threads:
                self._threads.remove(thread)
            if worker in self._workers:
                self._workers.remove(worker)
            if self._current_add_worker is worker:
                self._current_add_worker = None

        thread.finished.connect(_cleanup)
        thread.start()

    def _on_add_progress(self, message: str) -> None:
        self._set_window_progress(message)
        self.status_label.setText(f"Agregando ingrediente: {message}")

    def _on_add_finished(self, details, mode: str, value: float) -> None:
        logging.debug(
            f"_on_add_finished fdc_id={details.get('fdcId', '?')} "
            f"mode={mode} value={value} nutrients={len(details.get('foodNutrients', []) or [])}"
        )
        self._reset_add_ui_state()
        self._on_add_details_loaded(details, mode, value)

    def _reset_add_ui_state(self) -> None:
        self.add_button.setEnabled(True)
        self._set_window_progress(None)
        if self._current_add_worker:
            self._current_add_worker = None

    def _format_amount_for_status(self, amount_g: float, include_new: bool = False) -> str:
        """Return a user-facing label for a quantity using the active mode."""
        if self.quantity_mode == "g":
            return f"{amount_g:.1f} g"

        total = self._total_weight()
        if include_new:
            total += amount_g
        percent = 0.0 if total <= 0 else (amount_g / total) * 100.0
        return f"{percent:.2f} %"

    def _safe_base_name(self, fallback: str = "formulacion") -> str:
        """Return a filesystem-safe base name using the formula input or fallback."""
        name = (self.formula_name_input.text() or "").strip()
        clean = re.sub(r'[\\/:*?"<>|]+', "_", name).strip(". ")
        return clean or fallback

    def _prompt_quantity(
        self, default_amount: float | None = None, editing_index: int | None = None
    ) -> tuple[str | None, float]:
        if self.quantity_mode == "g":
            start_value = default_amount if default_amount is not None else 100.0
            title = "Cantidad"
            label = "Cantidad del ingrediente (g):"
            min_value, max_value, decimals = 0.1, 1_000_000.0, 1
        else:
            start_value = (
                self._amount_to_percent(default_amount or 0.0)
                if default_amount is not None
                else 10.0
            )
            title = "Porcentaje"
            label = "Porcentaje del ingrediente sobre la formulación (%):"
            min_value, max_value, decimals = 0.01, 100.0, 2

        raw_value, ok = QInputDialog.getDouble(
            self,
            title,
            label,
            start_value,
            min_value,
            max_value,
            decimals,
        )
        if not ok:
            return None, 0.0

        return ("g", raw_value) if self.quantity_mode == "g" else ("percent", raw_value)

    def _edit_quantity_for_row(self, row: int) -> None:
        """Prompt the user to update the quantity of a specific row."""
        if row < 0 or row >= len(self.formulation_items):
            self.status_label.setText("Fila seleccionada inválida.")
            return

        item = self.formulation_items[row]
        default_amount = item.get("amount_g", 0.0)
        mode, value = self._prompt_quantity(default_amount, editing_index=row)
        if mode is None:
            return

        if mode == "g":
            item["amount_g"] = value
        else:
            if not self._apply_percent_edit(row, value):
                return

        self._refresh_formulation_views()
        if mode == "g":
            msg_value = self._format_amount_for_status(item.get("amount_g", 0.0))
        else:
            msg_value = f"{value:.2f} %"
        self.status_label.setText(
            f"Actualizado {item.get('fdc_id', '')} a {msg_value}"
        )

    def _apply_percent_edit(self, target_idx: int, target_percent: float) -> bool:
        """Redistribute quantities so locked percentages stay fixed."""
        if target_percent < 0 or target_percent > 100:
            QMessageBox.warning(self, "Porcentaje inválido", "Ingresa un valor entre 0 y 100.")
            return False

        if target_idx < 0 or target_idx >= len(self.formulation_items):
            self.status_label.setText("Fila seleccionada inválida.")
            return False

        total_weight = self._total_weight()
        base_total = total_weight if total_weight > 0 else 100.0
        if base_total <= 0:
            QMessageBox.warning(
                self,
                "Porcentaje inválido",
                "No hay cantidad total para calcular porcentajes.",
            )
            return False

        current_percents = [
            (item.get("amount_g", 0.0) or 0.0) * 100.0 / base_total
            for item in self.formulation_items
        ]

        locked_sum = sum(
            current_percents[idx]
            for idx, item in enumerate(self.formulation_items)
            if item.get("locked") and idx != target_idx
        )
        if locked_sum > 100.0:
            QMessageBox.warning(
                self,
                "Porcentaje inválido",
                "Los ingredientes fijados ya superan el 100%. Libera uno para continuar.",
            )
            return False

        remaining = 100.0 - locked_sum - target_percent
        free_indices = [
            idx
            for idx, item in enumerate(self.formulation_items)
            if not item.get("locked") and idx != target_idx
        ]

        if remaining < -0.0001:
            QMessageBox.warning(
                self,
                "Sin grado de libertad",
                "No hay porcentaje libre suficiente. Libera un ingrediente o reduce el valor.",
            )
            return False

        if not free_indices and abs(remaining) > 0.0001:
            QMessageBox.warning(
                self,
                "Sin grado de libertad",
                "Debe quedar al menos un ingrediente sin fijar para ajustar porcentajes.",
            )
            return False

        cur_free_sum = sum(current_percents[idx] for idx in free_indices)
        new_percents = [0.0 for _ in self.formulation_items]

        for idx, item in enumerate(self.formulation_items):
            if item.get("locked") and idx != target_idx:
                new_percents[idx] = current_percents[idx]
            elif idx == target_idx:
                new_percents[idx] = target_percent
            else:
                if cur_free_sum > 0:
                    new_percents[idx] = (
                        current_percents[idx] * remaining / cur_free_sum
                    )
                elif free_indices:
                    new_percents[idx] = remaining if idx == free_indices[0] else 0.0

        if any(pct < -0.001 for pct in new_percents):
            QMessageBox.warning(
                self,
                "Sin grado de libertad",
                "No hay porcentaje libre suficiente. Ajusta los valores o libera un ingrediente.",
            )
            return False

        for idx, pct in enumerate(new_percents):
            safe_pct = max(pct, 0.0)
            self.formulation_items[idx]["amount_g"] = safe_pct * base_total / 100.0
        return True

    def _remove_selected_from_formulation(self, table: QTableWidget) -> None:
        indexes = table.selectionModel().selectedRows()
        if not indexes:
            self.status_label.setText("Selecciona un ingrediente para eliminar.")
            return
        row = indexes[0].row()
        if 0 <= row < len(self.formulation_items):
            removed = self.formulation_items.pop(row)
            if self.formulation_items and self._locked_count() == len(
                self.formulation_items
            ):
                self.formulation_items[0]["locked"] = False
            self._refresh_formulation_views()
            self.status_label.setText(
                f"Eliminado {removed.get('fdc_id', '')} - {removed.get('description', '')}"
            )
        else:
            self.status_label.setText("Fila seleccionada inválida.")

    def _run_in_thread(self, fn, args, on_success, on_error) -> None:
        """Start a worker thread for API calls and handle cleanup."""
        thread = QThread(self)
        worker = ApiWorker(fn, *args)
        worker.moveToThread(thread)
        self._workers.append(worker)

        thread.started.connect(worker.run)
        worker.finished.connect(on_success)
        worker.error.connect(on_error)
        worker.finished.connect(thread.quit)
        worker.error.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        worker.error.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)

        def _cleanup() -> None:
            if thread in self._threads:
                self._threads.remove(thread)
            if worker in self._workers:
                self._workers.remove(worker)

        thread.finished.connect(_cleanup)

        self._threads.append(thread)
        thread.start()

    # ---- Callbacks for async ops ----
    def _on_search_success(self, foods) -> None:
        foods_sorted = self._sort_search_results(foods)
        filtered = self._filter_results_by_query(foods_sorted, self.last_query)
        self.search_results = filtered
        self._last_results_count = len(filtered)
        self.search_page = 1
        self._show_current_search_page()
        total_pages = max(1, (len(filtered) + self.search_page_size - 1) // self.search_page_size)
        self.status_label.setText(
            f"Se encontraron {len(filtered)} resultados (pagina {self.search_page}/{total_pages})."
        )
        self.search_button.setEnabled(True)
        self._update_paging_buttons(len(filtered))

    def _on_search_error(self, message: str) -> None:
        self.status_label.setText(f"Error: {message}")
        self.search_button.setEnabled(True)
        self._update_paging_buttons(self._last_results_count)

    def _on_details_success(self, details) -> None:
        nutrients = normalize_nutrients(
            details.get("foodNutrients", []) or [], details.get("dataType")
        )
        self._populate_details_table(nutrients)

        desc = details.get("description", "") or ""
        fdc_id = details.get("fdcId", "")
        self.status_label.setText(
            f"Detalles de {fdc_id} - {desc} ({len(nutrients)} nutrientes)"
        )
        self.fdc_id_button.setEnabled(True)

    def _on_details_error(self, message: str) -> None:
        self.status_label.setText(f"Error al cargar detalles: {message}")
        self.fdc_id_button.setEnabled(True)

    def _on_add_details_loaded(self, details, mode: str, value: float) -> None:
        logging.debug(
            f"_on_add_details_loaded fdc_id={details.get('fdcId', '?')} "
            f"mode={mode} value={value}"
        )
        nutrients = normalize_nutrients(details.get("foodNutrients", []) or [], details.get("dataType"))
        self._update_reference_from_details(details)
        desc = details.get("description", "") or ""
        brand = details.get("brandOwner", "") or ""
        data_type = details.get("dataType", "") or ""
        fdc_id = details.get("fdcId", "")

        new_item = {
            "fdc_id": fdc_id,
            "description": desc,
            "brand": brand,
            "data_type": data_type,
            "amount_g": value if mode == "g" else 0.0,
            "nutrients": nutrients,
            "locked": False,
        }
        self.formulation_items.append(new_item)

        if mode == "percent":
            success = self._apply_percent_edit(len(self.formulation_items) - 1, value)
            if not success:
                self.formulation_items.pop()
                self.add_button.setEnabled(True)
                return

        self._populate_details_table(nutrients)
        self._refresh_formulation_views()
        self._select_preview_row(len(self.formulation_items) - 1)
        msg_value = (
            self._format_amount_for_status(new_item.get("amount_g", 0.0))
            if mode == "g"
            else f"{value:.2f} %"
        )
        self.status_label.setText(
            f"Agregado {fdc_id} - {desc} ({msg_value})"
        )
        self.add_button.setEnabled(True)
        self._upgrade_item_to_full(len(self.formulation_items) - 1, int(fdc_id))

    def _on_add_error(self, message: str) -> None:
        self._reset_add_ui_state()
        self.status_label.setText(f"Error al agregar: {message}")
        logging.error(f"_on_add_error: {message}")

    def _upgrade_item_to_full(self, index: int, fdc_id: int) -> None:
        """Upgrade no-op: abridged es la única fuente ahora."""
        return
    def on_totals_checkbox_changed(self, item: QTableWidgetItem) -> None:
        """Sync export checkbox state into memory and update toggle label."""
        if item.column() != 3:
            return
        nut_key = item.data(Qt.UserRole)
        if not nut_key:
            return
        self.nutrient_export_flags[nut_key] = item.checkState() == Qt.Checked
        self._update_toggle_export_button()

    def on_toggle_export_clicked(self) -> None:
        """Toggle all nutrient export checkboxes on/off."""
        if not self.nutrient_export_flags:
            return
        all_checked = all(self.nutrient_export_flags.values())
        new_state = not all_checked
        for key in list(self.nutrient_export_flags.keys()):
            self.nutrient_export_flags[key] = new_state
        self._populate_totals_table()
