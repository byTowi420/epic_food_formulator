"""Formulation presenter - orchestrates formulation use cases for UI.

Handles all formulation-related operations:
- Add/remove ingredients
- Calculate totals
- Adjust amounts/locks
- Generate label data
"""

from decimal import Decimal
import logging
import re
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

from config.container import Container
from config.constants import DATA_TYPE_PRIORITY
from domain.models import CurrencyRate, Formulation, Food, Ingredient, Nutrient, PackagingItem, ProcessCost
from domain.exceptions import FormulationImportError
from domain.services.unit_normalizer import convert_mass, normalize_mass_unit
from domain.services.nutrient_normalizer import augment_fat_nutrients, normalize_nutrients
from ui.adapters.formulation_mapper import FormulationMapper, NutrientDisplayMapper


class FormulationPresenter:
    """Presenter for formulation operations.

    Coordinates use cases and prepares data for UI display.
    UI should only call presenter methods, not use cases directly.
    """

    def __init__(self, container: Optional[Container] = None) -> None:
        """Initialize presenter.

        Args:
            container: DI container (creates new one if not provided)
        """
        self._container = container if container is not None else Container()
        self._formulation: Formulation = Formulation(name="New Formulation")

    @property
    def formulation(self) -> Formulation:
        """Expose the current formulation."""
        return self._formulation

    def set_yield_percent(self, value: Decimal) -> None:
        self._formulation.yield_percent = value

    def get_yield_percent(self) -> Decimal:
        return self._formulation.yield_percent

    def set_process_costs(self, processes: list[Any]) -> None:
        self._formulation.process_costs = list(processes)

    def set_packaging_items(self, items: list[Any]) -> None:
        self._formulation.packaging_items = list(items)

    def apply_cost_meta(self, meta: Dict[str, Any]) -> None:
        def _to_decimal(value: Any) -> Decimal | None:
            if value is None:
                return None
            if isinstance(value, Decimal):
                return value
            if isinstance(value, str):
                cleaned = value.strip().replace(",", ".")
                if cleaned == "":
                    return None
                return Decimal(cleaned)
            return Decimal(str(value))

        yield_raw = _to_decimal(meta.get("yield_percent"))
        if yield_raw is not None and Decimal("0") < yield_raw <= Decimal("100"):
            self._formulation.yield_percent = yield_raw

        if "process_costs" in meta:
            process_items = meta.get("process_costs") or []
            processes: list[ProcessCost] = []
            if isinstance(process_items, list):
                for entry in process_items:
                    if not isinstance(entry, dict):
                        continue
                    processes.append(
                        ProcessCost(
                            name=entry.get("name", "") or "",
                            scale_type=entry.get("scale_type", "") or "",
                            time_value=_to_decimal(entry.get("time_value")),
                            time_unit=entry.get("time_unit"),
                            cost_per_hour_mn=_to_decimal(entry.get("cost_per_hour_mn")),
                            total_cost_mn=_to_decimal(entry.get("total_cost_mn")),
                            setup_time_value=_to_decimal(entry.get("setup_time_value")),
                            setup_time_unit=entry.get("setup_time_unit"),
                            time_per_kg_value=_to_decimal(entry.get("time_per_kg_value")),
                            notes=entry.get("notes"),
                        )
                    )
            self._formulation.process_costs = processes

        if "packaging_items" in meta:
            packaging_items = meta.get("packaging_items") or []
            packaging: list[PackagingItem] = []
            if isinstance(packaging_items, list):
                for entry in packaging_items:
                    if not isinstance(entry, dict):
                        continue
                    qty = _to_decimal(entry.get("quantity_per_pack")) or Decimal("0")
                    unit_cost = _to_decimal(entry.get("unit_cost_mn")) or Decimal("0")
                    packaging.append(
                        PackagingItem(
                            name=entry.get("name", "") or "",
                            quantity_per_pack=qty,
                            unit_cost_mn=unit_cost,
                            notes=entry.get("notes"),
                        )
                    )
            self._formulation.packaging_items = packaging

        if "currency_rates" in meta:
            rate_items = meta.get("currency_rates") or []
            rates: list[CurrencyRate] = []
            if isinstance(rate_items, list):
                for entry in rate_items:
                    if not isinstance(entry, dict):
                        continue
                    symbol = str(entry.get("symbol", "") or "").strip()
                    name = str(entry.get("name", "") or "").strip()
                    rate_value = _to_decimal(entry.get("rate_to_mn"))
                    if not symbol or rate_value is None:
                        continue
                    rates.append(
                        CurrencyRate(
                            name=name or symbol,
                            symbol=symbol,
                            rate_to_mn=rate_value,
                        )
                    )
            if rates:
                self._formulation.currency_rates = rates
                ensure = getattr(self._formulation, "_ensure_currency_rates", None)
                if callable(ensure):
                    ensure()

    @property
    def formulation_name(self) -> str:
        """Get current formulation name."""
        return self._formulation.name

    @formulation_name.setter
    def formulation_name(self, value: str) -> None:
        """Set formulation name."""
        # Since Formulation is mutable, we can just update
        # (Or recreate if we want immutability - decision: keep simple)
        object.__setattr__(self._formulation, "name", value)

    def get_ui_items(self) -> List[Dict[str, Any]]:
        """Get current formulation as list of UI items.

        Returns:
            List of dicts for UI tables
        """
        return FormulationMapper.formulation_to_ui_items(self._formulation)

    def add_ingredient_from_details(
        self,
        details: Dict[str, Any],
        amount_g: float,
    ) -> List[Dict[str, Any]]:
        """Add ingredient using pre-fetched USDA details."""
        nutrients = normalize_nutrients(
            details.get("foodNutrients", []) or [],
            details.get("dataType"),
        )
        domain_nutrients = tuple(
            Nutrient(
                name=n["nutrient"]["name"],
                unit=n["nutrient"].get("unitName", ""),
                amount=Decimal(str(n["amount"])) if n.get("amount") is not None else Decimal("0"),
                nutrient_id=n["nutrient"].get("id"),
                nutrient_number=n["nutrient"].get("number"),
            )
            for n in nutrients
        )

        food = Food(
            fdc_id=details.get("fdcId", 0),
            description=details.get("description", "") or "",
            data_type=details.get("dataType", "") or "",
            brand_owner=details.get("brandOwner", "") or "",
            nutrients=domain_nutrients,
        )
        ingredient = Ingredient(food=food, amount_g=Decimal(str(amount_g)))
        self._formulation.add_ingredient(ingredient)
        return nutrients

    def add_manual_ingredient(
        self,
        *,
        description: str,
        amount_g: float,
        nutrients: List[Dict[str, Any]],
        brand: str = "",
    ) -> List[Dict[str, Any]]:
        """Add a manual ingredient with user-provided nutrients."""
        normalized = normalize_nutrients(nutrients or [], data_type="Manual")
        domain_nutrients = tuple(
            Nutrient(
                name=n["nutrient"]["name"],
                unit=n["nutrient"].get("unitName", ""),
                amount=Decimal(str(n["amount"])) if n.get("amount") is not None else Decimal("0"),
                nutrient_id=n["nutrient"].get("id"),
                nutrient_number=n["nutrient"].get("number"),
            )
            for n in normalized
        )
        food = Food(
            fdc_id=0,
            description=description or "",
            data_type="Manual",
            brand_owner=brand or "",
            nutrients=domain_nutrients,
        )
        ingredient = Ingredient(food=food, amount_g=Decimal(str(amount_g)))
        self._formulation.add_ingredient(ingredient)
        return normalized

    def add_ingredient_from_details_safe(
        self,
        details: Dict[str, Any],
        amount_g: float,
    ) -> List[Dict[str, Any]]:
        """Add ingredient using details with a safe nutrient fallback."""
        try:
            return self.add_ingredient_from_details(details, amount_g)
        except Exception as exc:  # noqa: BLE001
            logging.error("Error adding ingredient from details: %s", exc)
            return normalize_nutrients(
                details.get("foodNutrients", []) or [],
                details.get("dataType"),
            )

    def remove_ingredient_safe(self, index: int) -> tuple[bool, str | None]:
        """Remove ingredient and ensure at least one unlocked remains."""
        if index < 0 or index >= self.get_ingredient_count():
            return False, "row_invalid"
        self._formulation.remove_ingredient(index)
        if self.has_ingredients() and self.get_locked_count() == self.get_ingredient_count():
            self._formulation.get_ingredient(0).locked = False
        return True, None

    def update_ingredient_amount(
        self,
        index: int,
        amount_g: float,
        maintain_total: bool = False,
    ) -> None:
        """Update ingredient amount.

        Args:
            index: Ingredient index
            amount_g: New amount in grams
            maintain_total: Whether to adjust other ingredients
        """
        if maintain_total:
            self._container.formulation_service.set_ingredient_amount(
                self._formulation,
                index,
                Decimal(str(amount_g)),
                maintain_total_weight=True,
            )
        else:
            ingredient = self._formulation.get_ingredient(index)
            ingredient.amount_g = Decimal(str(amount_g))

    def set_lock_state(self, index: int, locked: bool) -> tuple[bool, str | None]:
        """Set lock state with validation (keeps at least one unlocked)."""
        if index < 0 or index >= self.get_ingredient_count():
            return False, "row_invalid"
        if locked and self.get_locked_count(exclude_index=index) >= (
            self.get_ingredient_count() - 1
        ):
            return False, "need_unlocked"
        ingredient = self._formulation.get_ingredient(index)
        ingredient.locked = locked
        return True, None

    def calculate_totals(self) -> Dict[str, Dict[str, Any]]:
        """Calculate nutrient totals per 100g.

        Returns:
            Dict of nutrient name -> display dict
        """
        totals = self._container.calculate_totals.execute(self._formulation)
        return NutrientDisplayMapper.totals_to_display_dict(totals, self._formulation)

    def normalize_to_target_weight(self, target_g: float) -> None:
        """Normalize formulation to a target total weight in grams."""
        self._container.formulation_service.scale_to_target_weight(
            self._formulation,
            Decimal(str(target_g)),
        )

    def export_to_excel(
        self,
        output_path: str,
        export_flags: Dict[str, bool] | None = None,
        mass_unit: str | None = None,
    ) -> None:
        """Export formulation to Excel.

        Args:
            output_path: Output file path
        """
        self._container.export_formulation.execute(
            self._formulation,
            output_path,
            export_flags=export_flags,
            mass_unit=mass_unit,
        )

    def export_to_excel_safe(
        self,
        output_path: str,
        *,
        export_flags: Dict[str, bool] | None = None,
        mass_unit: str | None = None,
        nutrient_ordering: Any | None = None,
    ) -> None:
        """Export formulation to Excel with a legacy fallback when needed."""
        try:
            self.export_to_excel(
                output_path,
                export_flags=export_flags,
                mass_unit=mass_unit,
            )
            return
        except Exception as exc:  # noqa: BLE001
            logging.error(
                "Export via use case failed (%s). Attempting legacy fallback.",
                exc,
            )
            if nutrient_ordering is None:
                raise
        self._export_to_excel_legacy(
            output_path,
            export_flags=export_flags,
            mass_unit=mass_unit,
            nutrient_ordering=nutrient_ordering,
        )

    def _export_to_excel_legacy(
        self,
        output_path: str,
        *,
        export_flags: Dict[str, bool] | None,
        mass_unit: str | None,
        nutrient_ordering: Any,
    ) -> None:
        export_flags = export_flags or {}
        mass_unit = normalize_mass_unit(mass_unit) or "g"
        self.normalize_items_nutrients()

        wb = Workbook()
        ws = wb.active
        ws.title = "Ingredientes"
        totals_sheet = wb.create_sheet("Totales")

        unit_decimals = self.mass_decimals(mass_unit)
        unit_format = "0" if unit_decimals <= 0 else f"0.{'0' * unit_decimals}"

        base_headers = [
            "FDC ID",
            "Ingrediente",
            "Marca / Origen",
            "Tipo de dato",
            f"Cantidad ({mass_unit})",
            "Cantidad (%)",
        ]

        nutrient_headers, header_categories, header_key_map = (
            self.collect_nutrient_columns(
                self.get_ui_items(),
                nutrient_ordering,
                export_flags,
            )
        )
        header_by_key = {v: k for k, v in header_key_map.items()}

        header_fill = PatternFill("solid", fgColor="D9D9D9")
        total_fill = PatternFill("solid", fgColor="FFF2CC")
        category_fills = {
            "Proximates": PatternFill("solid", fgColor="DAEEF3"),
            "Carbohydrates": PatternFill("solid", fgColor="E6B8B7"),
            "Minerals": PatternFill("solid", fgColor="C4D79B"),
            "Vitamins and Other Components": PatternFill("solid", fgColor="FFF2CC"),
            "Lipids": PatternFill("solid", fgColor="D9E1F2"),
            "Amino acids": PatternFill("solid", fgColor="E4DFEC"),
        }
        center = Alignment(horizontal="center", vertical="center")

        # Row 1: group titles (base + nutrient categories)
        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=len(base_headers),
        )
        ws.cell(row=1, column=1, value="Detalles de formulacion").alignment = center

        if nutrient_headers:
            start_col = len(base_headers) + 1
            col = start_col
            while col < start_col + len(nutrient_headers):
                idx = col - start_col
                category = header_categories.get(nutrient_headers[idx], "Nutrientes")
                run_start = col
                while (
                    col < start_col + len(nutrient_headers)
                    and header_categories.get(
                        nutrient_headers[col - start_col], "Nutrientes"
                    )
                    == category
                ):
                    col += 1
                run_end = col - 1
                ws.merge_cells(
                    start_row=1, start_column=run_start, end_row=1, end_column=run_end
                )
                cat_cell = ws.cell(row=1, column=run_start, value=category)
                cat_cell.alignment = center
                cat_cell.fill = category_fills.get(category, header_fill)

        # Row 2: headers
        headers = base_headers + nutrient_headers
        for col_idx, name in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=name)
            cell.fill = (
                category_fills.get(header_categories.get(name, ""), header_fill)
                if col_idx > len(base_headers)
                else header_fill
            )
            cell.alignment = center

        start_row = 3
        grams_col = base_headers.index(f"Cantidad ({mass_unit})") + 1
        percent_col = base_headers.index("Cantidad (%)") + 1
        data_rows = self.get_ingredient_count()
        end_row = start_row + data_rows - 1

        # Write ingredient rows
        for idx, item in enumerate(self.get_ui_items()):
            row = start_row + idx
            amount_g = float(item.get("amount_g", 0.0) or 0.0)
            amount_val = convert_mass(amount_g, "g", mass_unit)
            amount_display = float(amount_val) if amount_val is not None else amount_g
            values = [
                item.get("fdc_id", ""),
                item.get("description", ""),
                item.get("brand", ""),
                item.get("data_type", ""),
                amount_display,
                None,  # placeholder for percent formula
            ]
            for col_idx, val in enumerate(values, start=1):
                ws.cell(row=row, column=col_idx, value=val)

            gram_cell = f"{get_column_letter(grams_col)}{row}"
            total_range = (
                f"${get_column_letter(grams_col)}${start_row}:${get_column_letter(grams_col)}${end_row}"
            )
            ws.cell(
                row=row,
                column=percent_col,
                value=f"={gram_cell}/SUM({total_range})",
            ).number_format = "0.00%"

            nut_map = self.nutrients_by_header(
                item.get("nutrients", []),
                header_by_key,
                nutrient_ordering,
            )
            for offset, header in enumerate(
                nutrient_headers, start=len(base_headers) + 1
            ):
                if header in nut_map:
                    ws.cell(row=row, column=offset, value=nut_map[header])

        total_row = end_row + 1
        ws.cell(row=total_row, column=1, value="Total")
        ws.cell(row=total_row, column=2, value="Formulado")
        ws.cell(row=total_row, column=3, value="Formulado")
        ws.cell(row=total_row, column=4, value="Formulado")

        gram_total_cell = ws.cell(
            row=total_row,
            column=grams_col,
            value=(
                f"=SUBTOTAL(9,{get_column_letter(grams_col)}{start_row}:"
                f"{get_column_letter(grams_col)}{end_row})"
            ),
        )
        gram_total_cell.fill = total_fill
        percent_total_cell = ws.cell(row=total_row, column=percent_col, value="100%")
        percent_total_cell.number_format = "0.00%"
        percent_total_cell.fill = total_fill

        for offset, header in enumerate(nutrient_headers, start=len(base_headers) + 1):
            col_letter = get_column_letter(offset)
            formula = (
                f"=SUMPRODUCT(${get_column_letter(percent_col)}${start_row}:"
                f"${get_column_letter(percent_col)}${end_row},"
                f"${col_letter}${start_row}:${col_letter}${end_row})"
            )
            cell = ws.cell(row=total_row, column=offset, value=formula)
            cell.fill = total_fill

        for row in range(start_row, total_row + 1):
            ws.cell(row=row, column=grams_col).number_format = unit_format

        freeze_col = len(base_headers) + 1 if nutrient_headers else 1
        ws.freeze_panes = f"{get_column_letter(freeze_col)}3"

        widths = {
            "A": 12,
            "B": 35,
            "C": 18,
            "D": 14,
            "E": 12,
            "F": 12,
        }
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

        totals_headers = ["Nutriente", "Total", "Unidad"]
        for col_idx, name in enumerate(totals_headers, start=1):
            cell = totals_sheet.cell(row=1, column=col_idx, value=name)
            cell.fill = header_fill
            cell.alignment = center

        for idx, header in enumerate(nutrient_headers, start=1):
            name_part, unit = self.split_header_unit(header)
            totals_sheet.cell(row=idx + 1, column=1, value=name_part)
            totals_sheet.cell(row=idx + 1, column=3, value=unit or "")
            source_cell = (
                f"Ingredientes!{get_column_letter(len(base_headers) + idx)}{total_row}"
            )
            totals_sheet.cell(row=idx + 1, column=2, value=f"={source_cell}")

        wb.save(output_path)

    def parse_import_file(
        self,
        path: str,
        *,
        current_formula_name: str = "",
    ) -> tuple[list[Dict[str, Any]], Dict[str, Any]]:
        """Parse a formulation file into base items + metadata."""
        ext = Path(path).suffix.lower()
        importer = self._container.formulation_importer
        if ext == ".json":
            return importer.load_state_from_json(path)
        if ext in (".xlsx", ".xls"):
            return importer.load_state_from_excel(
                path,
                default_formula_name=current_formula_name,
            )
        raise FormulationImportError(
            "Formato no soportado",
            "Selecciona un archivo .json o .xlsx",
        )

    def build_import_preview_items(
        self,
        base_items: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """Prepare UI items for import preview (without USDA nutrients)."""
        ui_items: List[Dict[str, Any]] = []
        for item in base_items:
            fdc_id = item.get("fdc_id")
            description = (item.get("description") or "").strip()
            if not description:
                description = "Manual" if item.get("manual") else f"FDC {fdc_id}"
            data_type = (item.get("data_type") or "").strip() or "Imported"
            if item.get("manual"):
                data_type = "Manual"
            ui_items.append(
                {
                    "fdc_id": fdc_id,
                    "description": description,
                    "brand": item.get("brand", "") or "",
                    "data_type": data_type,
                    "amount_g": float(item.get("amount_g", 0.0) or 0.0),
                    "locked": bool(item.get("locked", False)),
                    "nutrients": [],
                    "cost_pack_amount": item.get("cost_pack_amount"),
                    "cost_pack_unit": item.get("cost_pack_unit"),
                    "cost_value": item.get("cost_value"),
                    "cost_currency_symbol": item.get("cost_currency_symbol"),
                    "cost_per_g_mn": item.get("cost_per_g_mn"),
                }
            )
        return ui_items

    def build_hydrated_items_from_payload(
        self,
        payload: List[Dict[str, Any]],
        *,
        on_details: Callable[[Dict[str, Any]], None] | None = None,
    ) -> List[Dict[str, Any]]:
        """Convert ImportWorker payload into hydrated UI items."""
        hydrated: List[Dict[str, Any]] = []
        for entry in payload:
            base = entry.get("base") or {}
            details = entry.get("details") or {}
            fdc_id = base.get("fdc_id") or details.get("fdcId")
            try:
                fdc_id_int = int(fdc_id) if fdc_id is not None else None
            except Exception:
                fdc_id_int = fdc_id

            if on_details:
                on_details(details)

            nutrients = augment_fat_nutrients(details.get("foodNutrients", []) or [])
            hydrated.append(
                {
                    "fdc_id": fdc_id_int,
                    "description": details.get("description", "")
                    or base.get("description", ""),
                    "brand": details.get("brandOwner", "") or base.get("brand", ""),
                    "data_type": details.get("dataType", "") or base.get("data_type", ""),
                    "amount_g": float(base.get("amount_g", 0.0) or 0.0),
                    "nutrients": normalize_nutrients(
                        nutrients, details.get("dataType")
                    ),
                    "locked": bool(base.get("locked", False)),
                    "cost_pack_amount": base.get("cost_pack_amount"),
                    "cost_pack_unit": base.get("cost_pack_unit"),
                    "cost_value": base.get("cost_value"),
                    "cost_currency_symbol": base.get("cost_currency_symbol"),
                    "cost_per_g_mn": base.get("cost_per_g_mn"),
                }
            )
        return hydrated

    def resolve_legacy_export_flags(
        self,
        legacy_flags: Dict[str, bool],
        hydrated_items: list[Dict[str, Any]],
    ) -> Dict[str, bool]:
        """Map legacy export flag keys to header-key flags."""
        return self._container.formulation_importer.resolve_legacy_export_flags(
            legacy_flags, hydrated_items
        )

    def get_total_weight(self) -> float:
        """Get total weight of formulation.

        Returns:
            Total weight in grams
        """
        return float(self._formulation.total_weight)

    def get_ingredient_count(self) -> int:
        """Get number of ingredients.

        Returns:
            Count of ingredients
        """
        return self._formulation.ingredient_count

    def has_ingredients(self) -> bool:
        """Check if formulation has any ingredients.

        Returns:
            True if formulation has ingredients
        """
        return self._formulation.ingredient_count > 0

    def get_ui_item(self, index: int) -> Dict[str, Any]:
        """Get single UI item by index.

        Args:
            index: Ingredient index

        Returns:
            UI item dict

        Raises:
            IndexError: If index out of range
        """
        ingredient = self._formulation.get_ingredient(index)
        return FormulationMapper.ingredient_to_ui_item(ingredient, index)

    def get_locked_count(self, exclude_index: Optional[int] = None) -> int:
        """Count locked ingredients.

        Args:
            exclude_index: Optional index to exclude from count

        Returns:
            Number of locked ingredients
        """
        count = 0
        for idx, ingredient in enumerate(self._formulation.ingredients):
            if exclude_index is not None and idx == exclude_index:
                continue
            if ingredient.locked:
                count += 1
        return count

    def load_from_ui_items(
        self,
        ui_items: List[Dict[str, Any]],
        name: str = "Loaded Formulation",
    ) -> None:
        """Load formulation from UI item list (for compatibility).

        Args:
            ui_items: List of UI item dicts
            name: Formulation name
        """
        self._formulation = FormulationMapper.ui_items_to_formulation(ui_items, name)

    def normalize_export_flags(self, flags: Dict[str, bool]) -> Dict[str, bool]:
        """Normalize export flag keys to stable lowercase strings."""
        normalized: Dict[str, bool] = {}
        for key, value in (flags or {}).items():
            if key is None:
                continue
            key_text = str(key).strip().lower()
            if not key_text:
                continue
            normalized[key_text] = bool(value)
        return normalized

    def normalize_items_nutrients(self, items: List[Dict[str, Any]] | None = None) -> None:
        """Normalize nutrient lists in-place for UI items."""
        items = items if items is not None else self.get_ui_items()
        for item in items:
            original = item.get("nutrients", []) or []
            normalized = normalize_nutrients(original, item.get("data_type"))
            if normalized != original:
                item["nutrients"] = normalized

    def split_header_unit(self, header: str) -> tuple[str, str]:
        if header.endswith(")") and " (" in header:
            name, unit = header.rsplit(" (", 1)
            return name, unit[:-1]
        return header, ""

    def nutrients_by_header(
        self,
        nutrients: List[Dict[str, Any]],
        header_by_key: Dict[str, str],
        nutrient_ordering,
    ) -> Dict[str, float]:
        """Map nutrient amounts to the selected header list."""
        out: Dict[str, float] = {}
        best_priority: Dict[str, int] = {}
        allowed_keys = set(header_by_key.keys())
        alias_priority = {
            "carbohydrate, by difference": 2,
            "carbohydrate, by summation": 1,
            "carbohydrate by summation": 1,
            "sugars, total": 2,
            "total sugars": 1,
        }

        for entry in nutrients:
            amount = entry.get("amount")
            if amount is None:
                continue
            nut = entry.get("nutrient") or {}
            header_key, _, _ = nutrient_ordering.header_key(nut)
            if not header_key or header_key not in allowed_keys:
                continue
            header = header_by_key[header_key]
            priority = alias_priority.get((nut.get("name") or "").strip().lower(), 0)
            current_best = best_priority.get(header, -1)
            if priority < current_best:
                continue

            best_priority[header] = priority
            out[header] = amount

        return out

    def collect_nutrient_columns(
        self,
        items: List[Dict[str, Any]],
        nutrient_ordering,
        export_flags: Dict[str, bool],
    ) -> tuple[List[str], Dict[str, str], Dict[str, str]]:
        """Collect ordered nutrient headers and their categories."""
        candidates: Dict[str, Dict[str, Any]] = {}
        categories_seen_order: Dict[str, int] = {}
        preferred_order = [cat for cat, _ in nutrient_ordering.catalog]
        preferred_count = len(preferred_order)

        for item in items:
            data_priority = DATA_TYPE_PRIORITY.get(
                (item.get("data_type") or "").strip(), len(DATA_TYPE_PRIORITY)
            )
            for entry in nutrient_ordering.sort_nutrients_for_display(
                item.get("nutrients", [])
            ):
                nut = entry.get("nutrient") or {}
                amount = entry.get("amount")
                if amount is None:
                    continue
                header_key, canonical_name, canonical_unit = nutrient_ordering.header_key(nut)
                if header_key and not export_flags.get(header_key, True):
                    continue

                if not header_key or not canonical_name:
                    continue

                category = nutrient_ordering.category_for_nutrient(canonical_name, nut)
                if category not in categories_seen_order:
                    categories_seen_order[category] = len(categories_seen_order)

                order = nutrient_ordering.order_for_name(canonical_name)
                if order is None:
                    order = nutrient_ordering.nutrient_order(nut, len(candidates))
                unit_lower = (canonical_unit or "").strip().lower()
                if canonical_name.strip().lower() == "energy":
                    if unit_lower == "kcal":
                        order = order - 0.1 if isinstance(order, (int, float)) else order
                    elif unit_lower == "kj":
                        order = order + 0.1 if isinstance(order, (int, float)) else order

                header = (
                    f"{canonical_name} ({canonical_unit})"
                    if canonical_unit
                    else canonical_name
                )

                existing = candidates.get(header_key)
                if existing is None or (
                    data_priority < existing["data_priority"]
                    or (
                        data_priority == existing["data_priority"]
                        and order < existing["order"]
                    )
                    or (
                        data_priority == existing["data_priority"]
                        and order == existing["order"]
                        and header < existing["header"]
                    )
                ):
                    candidates[header_key] = {
                        "header_key": header_key,
                        "header": header,
                        "category": category,
                        "order": order,
                        "data_priority": data_priority,
                    }

        def category_rank(cat: str) -> int:
            if cat in preferred_order:
                return preferred_order.index(cat)
            return preferred_count + categories_seen_order.get(cat, preferred_count)

        sorted_candidates = sorted(
            candidates.values(),
            key=lambda c: (
                category_rank(c["category"]),
                c["order"],
                c["header"].lower(),
            ),
        )

        ordered_headers: List[str] = [c["header"] for c in sorted_candidates]
        categories: Dict[str, str] = {c["header"]: c["category"] for c in sorted_candidates}
        header_key_map: Dict[str, str] = {c["header"]: c["header_key"] for c in sorted_candidates}

        return ordered_headers, categories, header_key_map

    def calculate_totals_with_fallback(self, nutrient_ordering) -> Dict[str, Dict[str, Any]]:
        """Calculate totals, falling back to UI item nutrients if needed."""
        logging.debug(
            "calculate_totals_with_fallback start items=%s",
            self.get_ingredient_count(),
        )
        try:
            totals = self.calculate_totals()
            totals = nutrient_ordering.normalize_totals_by_header_key(totals)
            logging.debug(
                "calculate_totals_with_fallback done (via presenter) nutrients=%s",
                len(totals),
            )
            return totals
        except Exception as exc:  # noqa: BLE001
            logging.error("Error calculating totals via presenter: %s", exc)

        self.normalize_items_nutrients()
        totals: Dict[str, Dict[str, Any]] = {}
        total_weight = self.get_total_weight()
        for item in self.get_ui_items():
            qty = item.get("amount_g", 0) or 0
            for nutrient in nutrient_ordering.sort_nutrients_for_display(
                item.get("nutrients", [])
            ):
                amount = nutrient.get("amount")
                if amount is None:
                    continue
                nut = nutrient.get("nutrient") or {}
                header_key, canonical_name, canonical_unit = nutrient_ordering.header_key(nut)
                if not header_key:
                    continue
                entry = totals.setdefault(
                    header_key,
                    {
                        "name": canonical_name or nut.get("name", ""),
                        "unit": canonical_unit or "",
                        "amount": 0.0,
                        "order": nutrient_ordering.nutrient_order(
                            nut, len(totals)
                        ),
                    },
                )
                if canonical_name and not entry["name"]:
                    entry["name"] = canonical_name
                if canonical_unit and not entry["unit"]:
                    entry["unit"] = canonical_unit
                inferred_unit = nutrient_ordering.infer_unit(nut)
                if inferred_unit and not entry["unit"]:
                    entry["unit"] = inferred_unit
                entry["order"] = min(
                    entry.get("order", float("inf")),
                    nutrient_ordering.nutrient_order(nut, len(totals)),
                )
                entry["amount"] += amount * qty / 100.0

        if total_weight > 0:
            factor = 100.0 / total_weight
            for entry in totals.values():
                entry["amount"] *= factor
        logging.debug(
            "calculate_totals_with_fallback done (fallback) nutrients=%s total_weight=%s",
            len(totals),
            total_weight,
        )
        return totals

    def build_totals_rows(
        self,
        totals: Dict[str, Dict[str, Any]],
        nutrient_ordering,
        export_flags: Dict[str, bool],
    ) -> tuple[List[Dict[str, Any]], Dict[str, bool]]:
        """Order totals for display and sync export flags."""
        category_order = [cat for cat, _ in nutrient_ordering.catalog]

        def _cat_rank(name: str) -> int:
            cat = nutrient_ordering.category_for_nutrient(name)
            if cat in category_order:
                return category_order.index(cat)
            return len(category_order) + 1

        def _order_val(name: str) -> float:
            order = nutrient_ordering.order_for_name(name)
            return float(order if order is not None else float("inf"))

        sorted_totals = sorted(
            totals.items(),
            key=lambda item: (
                _cat_rank(item[1].get("name", "")),
                _order_val(item[1].get("name", "")),
                item[1].get("name", "").lower(),
            ),
        )

        rows: List[Dict[str, Any]] = []
        new_flags: Dict[str, bool] = {}
        for nut_key, entry in sorted_totals:
            current_checked = export_flags.get(nut_key, True)
            new_flags[nut_key] = current_checked
            rows.append(
                {
                    "key": nut_key,
                    "name": entry.get("name", ""),
                    "amount": float(entry.get("amount", 0.0) or 0.0),
                    "unit": entry.get("unit", ""),
                    "checked": current_checked,
                }
            )

        return rows, new_flags

    def normalize_quantity_mode(self, mode_raw: str) -> str:
        mode_lower = str(mode_raw or "g").strip().lower()
        if mode_lower in ("%", "percent", "percentage"):
            return "%"
        return normalize_mass_unit(mode_lower) or "g"

    def is_percent_mode(self, quantity_mode: str) -> bool:
        return quantity_mode == "%"

    def current_mass_unit(self, quantity_mode: str) -> str:
        if self.is_percent_mode(quantity_mode):
            return "g"
        return normalize_mass_unit(quantity_mode) or "g"

    def quantity_mode_label(self, mode: str) -> str:
        labels = {
            "g": "gramos (g)",
            "kg": "kilogramos (kg)",
            "ton": "toneladas (ton)",
            "lb": "libras (lb)",
            "oz": "onzas (oz)",
            "%": "porcentaje (%)",
        }
        return labels.get(mode, mode)

    def mass_decimals(self, unit: str) -> int:
        return {
            "g": 1,
            "kg": 3,
            "ton": 6,
            "lb": 3,
            "oz": 3,
        }.get(unit, 2)

    def normalization_dialog_values(
        self,
        total_g: float,
        unit: str,
    ) -> tuple[float, float, float, int]:
        """Return suggested bounds for the normalization dialog."""
        unit = normalize_mass_unit(unit) or unit or "g"
        start_value = convert_mass(total_g, "g", unit) or total_g
        min_value = convert_mass(0.1, "g", unit) or 0.1
        max_value = convert_mass(1_000_000.0, "g", unit) or 1_000_000.0
        decimals = self.mass_decimals(unit)
        return float(start_value), float(min_value), float(max_value), decimals

    def convert_to_grams(self, value: float, unit: str) -> float | None:
        unit = normalize_mass_unit(unit) or unit or "g"
        converted = convert_mass(value, unit, "g")
        return float(converted) if converted is not None else None

    def display_amount_for_unit(self, amount_g: float, quantity_mode: str) -> float:
        unit = self.current_mass_unit(quantity_mode)
        converted = convert_mass(amount_g, "g", unit)
        return float(converted) if converted is not None else amount_g

    def format_mass_amount(self, amount_g: float, quantity_mode: str) -> str:
        unit = self.current_mass_unit(quantity_mode)
        converted = self.display_amount_for_unit(amount_g, quantity_mode)
        decimals = self.mass_decimals(unit)
        return f"{converted:.{decimals}f} {unit}"

    def amount_to_percent(self, amount_g: float, total_weight: float) -> float:
        if total_weight <= 0:
            return 0.0
        return (amount_g / total_weight) * 100.0

    def format_amount_for_status(
        self,
        amount_g: float,
        *,
        quantity_mode: str,
        total_weight: float,
        include_new: bool = False,
    ) -> str:
        if not self.is_percent_mode(quantity_mode):
            return self.format_mass_amount(amount_g, quantity_mode)
        total = total_weight
        if include_new:
            total += amount_g
        percent = self.amount_to_percent(amount_g, total)
        return f"{percent:.2f} %"

    def apply_percent_edit(self, target_idx: int, target_percent: float) -> tuple[bool, str | None]:
        if target_percent < 0 or target_percent > 100:
            return False, "percent_range"

        count = self.get_ingredient_count()
        if target_idx < 0 or target_idx >= count:
            return False, "row_invalid"

        total_weight = float(self._formulation.total_weight)
        base_total = total_weight if total_weight > 0 else 100.0
        if base_total <= 0:
            return False, "no_total"

        current_percents = [
            float(ingredient.amount_g) * 100.0 / base_total
            for ingredient in self._formulation.ingredients
        ]

        locked_sum = sum(
            current_percents[idx]
            for idx, ingredient in enumerate(self._formulation.ingredients)
            if ingredient.locked and idx != target_idx
        )
        if locked_sum > 100.0:
            return False, "locked_over"

        remaining = 100.0 - locked_sum - target_percent
        free_indices = [
            idx
            for idx, ingredient in enumerate(self._formulation.ingredients)
            if not ingredient.locked and idx != target_idx
        ]

        if remaining < -0.0001:
            return False, "insufficient_free"

        if not free_indices and abs(remaining) > 0.0001:
            return False, "need_free"

        cur_free_sum = sum(current_percents[idx] for idx in free_indices)
        new_percents = [0.0 for _ in range(count)]

        for idx, ingredient in enumerate(self._formulation.ingredients):
            if ingredient.locked and idx != target_idx:
                new_percents[idx] = current_percents[idx]
            elif idx == target_idx:
                new_percents[idx] = target_percent
            else:
                if cur_free_sum > 0:
                    new_percents[idx] = (
                        current_percents[idx] * remaining / cur_free_sum
                    )
                elif free_indices:
                    new_percents[idx] = remaining if idx == free_indices[0] else 0.0

        if any(pct < -0.001 for pct in new_percents):
            return False, "negative_percent"

        for idx, pct in enumerate(new_percents):
            safe_pct = max(pct, 0.0)
            amount_g = safe_pct * base_total / 100.0
            self.update_ingredient_amount(idx, amount_g)
        return True, None

    def build_export_payload(
        self,
        *,
        formula_name: str,
        quantity_mode: str,
        export_flags: Dict[str, bool],
        label_settings: Dict[str, Any],
    ) -> Dict[str, Any]:
        """Assemble formulation + UI config into a JSON-friendly payload."""
        def _as_float(value: Any) -> float | None:
            if value is None:
                return None
            try:
                return float(value)
            except Exception:
                return None

        items: list[Dict[str, Any]] = []
        for item in self.get_ui_items():
            fdc_raw = item.get("fdc_id")
            try:
                fdc_id = int(fdc_raw)
            except Exception:
                fdc_id = fdc_raw

            amount_raw = item.get("amount_g", 0.0)
            try:
                amount_g = float(amount_raw) if amount_raw is not None else 0.0
            except Exception:
                amount_g = 0.0

            data_type = item.get("data_type", "") or ""
            nutrient_payload: list[Dict[str, Any]] = []
            if data_type.strip().lower() == "manual":
                for entry in item.get("nutrients", []) or []:
                    nut = entry.get("nutrient") or {}
                    nutrient_payload.append(
                        {
                            "name": nut.get("name", ""),
                            "unit": nut.get("unitName", ""),
                            "amount": entry.get("amount"),
                            "nutrient_id": nut.get("id"),
                            "nutrient_number": nut.get("number"),
                        }
                    )

            payload_item = {
                "fdc_id": fdc_id,
                "description": item.get("description", "") or "",
                "brand": item.get("brand") or item.get("brand_owner") or "",
                "data_type": data_type,
                "amount_g": amount_g,
                "locked": bool(item.get("locked", False)),
                "cost_pack_amount": _as_float(item.get("cost_pack_amount")),
                "cost_pack_unit": item.get("cost_pack_unit"),
                "cost_value": _as_float(item.get("cost_value")),
                "cost_currency_symbol": item.get("cost_currency_symbol"),
                "cost_per_g_mn": _as_float(item.get("cost_per_g_mn")),
            }
            if nutrient_payload:
                payload_item["nutrients"] = nutrient_payload

            items.append(payload_item)

        return {
            "version": 3,
            "formula_name": formula_name or "Current Formulation",
            "quantity_mode": quantity_mode,
            "yield_percent": _as_float(self._formulation.yield_percent),
            "process_costs": [
                {
                    "name": process.name,
                    "scale_type": process.scale_type,
                    "time_value": _as_float(process.time_value),
                    "time_unit": process.time_unit,
                    "cost_per_hour_mn": _as_float(process.cost_per_hour_mn),
                    "total_cost_mn": _as_float(process.total_cost_mn),
                    "setup_time_value": _as_float(process.setup_time_value),
                    "setup_time_unit": process.setup_time_unit,
                    "time_per_kg_value": _as_float(process.time_per_kg_value),
                    "notes": process.notes,
                }
                for process in self._formulation.process_costs
            ],
            "packaging_items": [
                {
                    "name": item.name,
                    "quantity_per_pack": _as_float(item.quantity_per_pack),
                    "unit_cost_mn": _as_float(item.unit_cost_mn),
                    "notes": item.notes,
                }
                for item in self._formulation.packaging_items
            ],
            "currency_rates": [
                {
                    "name": rate.name,
                    "symbol": rate.symbol,
                    "rate_to_mn": _as_float(rate.rate_to_mn),
                }
                for rate in self._formulation.currency_rates
            ],
            "items": items,
            "nutrient_export_flags": self.normalize_export_flags(export_flags),
            "label_settings": label_settings,
        }

    def safe_base_name(self, name: str, fallback: str = "formulacion") -> str:
        """Return a filesystem-safe base name using provided text or fallback."""
        raw = (name or "").strip()
        clean = re.sub(r'[\\/:*?"<>|]+', "_", raw).strip(". ")
        return clean or fallback
