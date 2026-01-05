from __future__ import annotations

import json
import logging
import re
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

from config.constants import DATA_TYPE_PRIORITY
from PySide6.QtCore import QCoreApplication, QItemSelectionModel, QThread, Qt, QTimer
from PySide6.QtGui import QColor, QFont, QIcon, QPainter, QPixmap
from PySide6.QtWidgets import (
    QComboBox,
    QFileDialog,
    QHBoxLayout,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
)
from ui.tabs.table_utils import (
    apply_selection_bar,
    attach_copy_shortcut,
    set_formulation_column_widths,
)

from domain.exceptions import FormulationImportError
from domain.services.nutrient_ordering import NutrientOrdering
from domain.services.unit_normalizer import convert_mass, normalize_mass_unit
from services.nutrient_normalizer import (
    augment_fat_nutrients,
    normalize_nutrients,
)
from ui.workers import AddWorker, ApiWorker, ImportWorker


class FormulationTabMixin:
    """Formulation tab UI and behavior."""

    # ---- State ----
    def _init_formulation_state(self) -> None:
        """Initialize formulation tab state."""
        # Async worker management.
        self._threads: list[QThread] = []
        self._workers: list[object] = []
        self._current_import_worker: ImportWorker | None = None
        self._current_add_worker: AddWorker | None = None

        # Import/add defaults.
        self.import_max_attempts = 4
        self.import_read_timeout = 8.0

        # Formulation UI state.
        self.quantity_mode = "g"
        self.quantity_mode_options = [
            ("g", "Gramos (g)"),
            ("%", "Porcentaje (%)"),
            ("kg", "Kilogramos (kg)"),
            ("ton", "Toneladas (ton)"),
            ("lb", "Libras (lb)"),
            ("oz", "Onzas (oz)"),
        ]
        self.amount_g_column_index = 2
        self.percent_column_index = 3
        self.lock_column_index = 4
        self.nutrient_export_flags: Dict[str, bool] = {}

        # Nutrient ordering helpers.
        self.nutrient_ordering = NutrientOrdering()

        # Cache for totals used by label preview.
        self._last_totals: Dict[str, Dict[str, Any]] = {}

    # ---- UI build ----
    def _build_formulation_tab_ui(self) -> None:
        """Build the Formulation tab UI."""
        # Layout base del tab de formulacion.
        layout = QVBoxLayout(self.formulation_tab)
    
        # Encabezado: export/import, nombre y unidad.
        header_layout = QHBoxLayout()
        self.export_state_button = QPushButton("Exportar")
        self.import_state_button = QPushButton("Importar")
        header_layout.addWidget(self.export_state_button)
        header_layout.addWidget(self.import_state_button)
        header_layout.addStretch()
        self.formula_name_input = QLineEdit()
        self.formula_name_input.setPlaceholderText(
            "Nombre Fórmula Ej.: Pan dulce con chocolate"
        )
        self.formula_name_input.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(self.formula_name_input, 1)
        header_layout.addStretch()
        header_layout.addWidget(QLabel("Unidad de formulación:"))
        self.quantity_mode_selector = QComboBox()
        self.quantity_mode_selector.addItems(
            [label for _, label in self.quantity_mode_options]
        )
        header_layout.addWidget(self.quantity_mode_selector)
        self.normalize_total_button = QPushButton("Normalizar masa")
        header_layout.addWidget(self.normalize_total_button)
        layout.addLayout(header_layout)
    
        # Tabla principal de ingredientes.
        self.formulation_table = QTableWidget(0, 6)
        self.formulation_table.setHorizontalHeaderLabels(
            [
                "FDC ID",
                "Ingrediente",
                "Cantidad (g)",
                "Cantidad (%)",
                "Fijar %",
                "Marca / Origen",
            ]
        )
        self.formulation_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.formulation_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.formulation_table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.formulation_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.formulation_table.horizontalHeader().setStretchLastSection(True)
        apply_selection_bar(self.formulation_table)
        layout.addWidget(self.formulation_table)
    
        # Acciones sobre ingredientes.
        buttons_layout = QHBoxLayout()
        self.edit_quantity_button = QPushButton("Editar cantidad seleccionada")
        self.remove_formulation_button = QPushButton(
            "Eliminar ingrediente seleccionado"
        )
        buttons_layout.addWidget(self.edit_quantity_button)
        buttons_layout.addWidget(self.remove_formulation_button)
        buttons_layout.addStretch()
        layout.addLayout(buttons_layout)
    
        # Totales nutricionales.
        layout.addWidget(
            QLabel(
                "Totales nutricionales (asumiendo valores por 100 g del alimento origen)"
            )
        )
        self.totals_table = QTableWidget(0, 4)
        self.totals_table.setHorizontalHeaderLabels(
            ["Nutriente", "Total", "Unidad", "Exportar"]
        )
        export_header = QTableWidgetItem("Exportar")
        export_header.setIcon(self._create_question_icon())
        export_header.setToolTip(
            "Los nutrientes seleccionados seran exportados al excel"
        )
        self.totals_table.setHorizontalHeaderItem(3, export_header)
        self.totals_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.totals_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.totals_table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.totals_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.totals_table.horizontalHeader().setStretchLastSection(True)
        apply_selection_bar(self.totals_table)
        layout.addWidget(self.totals_table)
    
        # Acciones de exportacion.
        export_buttons_layout = QHBoxLayout()
        self.toggle_export_button = QPushButton("Deseleccionar todos")
        self.export_excel_button = QPushButton("Exportar a Excel")
        export_buttons_layout.addStretch()
        export_buttons_layout.addWidget(self.toggle_export_button)
        export_buttons_layout.addWidget(self.export_excel_button)
        export_buttons_layout.addStretch()
        layout.addLayout(export_buttons_layout)
    
        # Conexiones de eventos.
        self.remove_formulation_button.clicked.connect(
            self.on_remove_formulation_clicked
        )
        self.export_excel_button.clicked.connect(self.on_export_to_excel_clicked)
        self.quantity_mode_selector.currentIndexChanged.connect(
            self.on_quantity_mode_changed
        )
        self.normalize_total_button.clicked.connect(
            self.on_normalize_total_clicked
        )
        self.formulation_table.cellDoubleClicked.connect(
            self.on_formulation_cell_double_clicked
        )
        self.edit_quantity_button.clicked.connect(self.on_edit_quantity_clicked)
        self.formulation_table.itemChanged.connect(self.on_lock_toggled_from_table)
        self.totals_table.itemChanged.connect(self.on_totals_checkbox_changed)
        self.toggle_export_button.clicked.connect(self.on_toggle_export_clicked)
        self.export_state_button.clicked.connect(self.on_export_state_clicked)
        self.import_state_button.clicked.connect(self.on_import_state_clicked)
    
        # Atajos de copia para tablas (Ctrl+C).
        # Copy shortcuts (Ctrl+C) for all tables
        for table in (
            self.table,
            self.details_table,
            self.formulation_preview,
            self.formulation_table,
            self.totals_table,
        ):
            attach_copy_shortcut(table)
    
        # Ajuste inicial de columnas.
        set_formulation_column_widths(self.formulation_table, self.totals_table)


    def _is_importing(self) -> bool:
        """Return True if an import hydration worker is running."""
        return self._current_import_worker is not None


    # ---- Formulation actions ----
    def on_remove_formulation_clicked(self) -> None:
        if self._is_importing():
            self.status_label.setText("Importacion en curso. Espera para editar.")
            return
        self._remove_selected_from_formulation(self.formulation_table)


    def on_formulation_cell_double_clicked(self, row: int, column: int) -> None:
        """Double click on formulation row -> edit its quantity."""
        if self._is_importing():
            self.status_label.setText("Importacion en curso. Espera para editar.")
            return
        if not self._can_edit_column(column):
            return
        self._edit_quantity_for_row(row)


    def on_lock_toggled_from_table(self, item: QTableWidgetItem) -> None:
        """Handle lock/unlock toggles coming from any formulation table."""
        if self._is_importing():
            return
        if item.column() != self.lock_column_index:
            return
        if not self._is_percent_mode():
            return
        table = item.tableWidget()
        if table not in (self.formulation_table, self.formulation_preview):
            return
        row = item.row()
        if row < 0 or row >= self.formulation_presenter.get_ingredient_count():
            return

        desired_locked = item.checkState() == Qt.Checked
        if desired_locked and self.formulation_presenter.get_locked_count(exclude_index=row) >= (
            self.formulation_presenter.get_ingredient_count() - 1
        ):
            # Avoid all items locked: keep one free.
            table.blockSignals(True)
            item.setCheckState(Qt.Unchecked)
            table.blockSignals(False)
            self.status_label.setText("Debe quedar al menos un ingrediente sin fijar.")
            return

        # Toggle lock via presenter
        try:
            self.formulation_presenter.toggle_lock(row)
        except Exception as e:
            logging.error(f"Error toggling lock via presenter: {e}")

        self._refresh_formulation_views()


    def on_edit_quantity_clicked(self) -> None:
        if self._is_importing():
            self.status_label.setText("Importacion en curso. Espera para editar.")
            return
        indexes = self.formulation_table.selectionModel().selectedRows()
        if not indexes:
            self.status_label.setText("Selecciona un ingrediente para editar.")
            return
        self._edit_quantity_for_row(indexes[0].row())


    def on_quantity_mode_changed(self) -> None:
        """Switch between quantity modes for formulation."""
        idx = self.quantity_mode_selector.currentIndex()
        if 0 <= idx < len(self.quantity_mode_options):
            self.quantity_mode = self.quantity_mode_options[idx][0]
        else:
            self.quantity_mode = "g"
        self._refresh_formulation_views()
        mode_text = self._quantity_mode_label(self.quantity_mode)
        self.status_label.setText(f"Modo de cantidad cambiado a {mode_text}.")

    def on_normalize_total_clicked(self) -> None:
        """Scale formulation to a target total mass."""
        if self._is_importing():
            self.status_label.setText("Importacion en curso. Espera para editar.")
            return
        if not self.formulation_presenter.has_ingredients():
            self.status_label.setText("No hay ingredientes para normalizar.")
            return

        unit = self._current_mass_unit()
        total_g = self._total_weight()
        start_value = convert_mass(total_g, "g", unit) or total_g
        decimals = self._mass_decimals(unit)
        min_value = convert_mass(0.1, "g", unit) or 0.1
        max_value = convert_mass(1_000_000.0, "g", unit) or 1_000_000.0

        target_value, ok = QInputDialog.getDouble(
            self,
            "Normalizar masa",
            f"Masa total objetivo ({unit}):",
            float(start_value),
            float(min_value),
            float(max_value),
            decimals,
        )
        if not ok:
            return

        target_g = convert_mass(target_value, unit, "g")
        if target_g is None or target_g <= 0:
            QMessageBox.warning(
                self,
                "Valor invalido",
                "Ingresa una masa total valida.",
            )
            return

        try:
            self.formulation_presenter.normalize_to_target_weight(float(target_g))
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(
                self,
                "Error al normalizar",
                f"No se pudo normalizar la formulacion:\n{exc}",
            )
            return

        self._refresh_formulation_views()
        self.status_label.setText(
            f"Formulacion normalizada a {target_value:.{decimals}f} {unit}."
        )


    def on_export_to_excel_clicked(self) -> None:
        """Export current formulation and totals to an Excel file."""
        if not self.formulation_presenter.has_ingredients():
            QMessageBox.information(
                self,
                "Exportar a Excel",
                "No hay ingredientes en la formulacion para exportar.",
            )
            return

        default_name = f"{self._safe_base_name()}.xlsx"
        initial_path = (
            str(Path(self.last_path or "").with_name(default_name))
            if self.last_path
            else default_name
        )
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar formulacion a Excel",
            initial_path,
            "Archivos de Excel (*.xlsx)",
        )
        if not path:
            return

        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        self._save_last_path(path)
        try:
            self._export_formulation_to_excel(path)
        except Exception as exc:  # noqa: BLE001 - surface the error to the user
            QMessageBox.critical(
                self,
                "Error al exportar",
                f"No se pudo exportar el archivo:\n{exc}",
            )
        else:
            QMessageBox.information(
                self,
                "Exportado",
                f"Archivo guardado en:\n{path}",
            )


    # ---- Import/export ----
    def on_export_state_clicked(self) -> None:
        """Export formulation state (ingredientes + cantidades + flags) a JSON."""
        if not self.formulation_presenter.has_ingredients():
            QMessageBox.information(
                self,
                "Exportar formulación",
                "No hay ingredientes en la formulación para exportar.",
            )
            return

        default_name = f"{self._safe_base_name()}.json"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar formulación",
            str(Path(self.last_path or "").with_name(default_name)) if self.last_path else default_name,
            "Archivos JSON (*.json)",
        )
        if not path:
            return
        if not path.lower().endswith(".json"):
            path += ".json"

        self._save_last_path(path)

        try:
            # Update formulation name from UI input and build JSON payload.
            formulation_name = self.formula_name_input.text() or "Current Formulation"
            self.formulation_presenter.formulation_name = formulation_name
            data = self._build_formulation_export_payload()
            Path(path).write_text(
                json.dumps(data, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(
                self,
                "Error al exportar",
                f"No se pudo exportar la formulación:\n{exc}",
            )
            return

        self.status_label.setText(f"Formulación exportada en {path}")


    def on_import_state_clicked(self) -> None:
        """Import formulation state from JSON or Excel and refresh UI."""
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Importar formulación",
            self.last_path or "",
            "Archivos JSON (*.json);;Archivos Excel (*.xlsx)",
        )
        if not path:
            return
        self._save_last_path(path)

        try:
            base_items, meta = self.formulation_presenter.parse_import_file(
                path,
                current_formula_name=self.formula_name_input.text(),
            )
        except FormulationImportError as exc:
            if exc.severity == "critical":
                QMessageBox.critical(self, exc.title, str(exc))
            else:
                QMessageBox.warning(self, exc.title, str(exc))
            return
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(
                self,
                "Error al importar",
                f"No se pudo leer el archivo:\n{exc}",
            )
            return

        meta.setdefault("path", path)
        self._start_import_hydration(base_items, meta)


    def _build_formulation_export_payload(self) -> Dict[str, Any]:
        """Assemble formulation + UI config into a JSON-friendly payload."""
        items: list[Dict[str, Any]] = []
        for item in self.formulation_presenter.get_ui_items():
            fdc_raw = item.get("fdc_id")
            try:
                fdc_id = int(fdc_raw)
            except Exception:
                fdc_id = fdc_raw

            amount_raw = item.get("amount_g", 0.0)
            try:
                amount_g = float(amount_raw) if amount_raw is not None else 0.0
            except Exception:
                amount_g = 0.0

            items.append(
                {
                    "fdc_id": fdc_id,
                    "description": item.get("description", "") or "",
                    "brand": item.get("brand") or item.get("brand_owner") or "",
                    "data_type": item.get("data_type", "") or "",
                    "amount_g": amount_g,
                    "locked": bool(item.get("locked", False)),
                }
            )

        return {
            "version": 3,
            "formula_name": self.formula_name_input.text() or "Current Formulation",
            "quantity_mode": self.quantity_mode,
            "items": items,
            "nutrient_export_flags": self._snapshot_export_flags(),
            "label_settings": self._snapshot_label_settings(),
        }


    def _snapshot_export_flags(self) -> Dict[str, bool]:
        """Return export flag map using stable header keys."""
        flags: Dict[str, bool] = {}
        for key, value in (self.nutrient_export_flags or {}).items():
            if key is None:
                continue
            key_text = str(key).strip().lower()
            if not key_text:
                continue
            flags[key_text] = bool(value)
        return flags


    def _set_import_controls_enabled(self, enabled: bool) -> None:
        """Toggle formulation controls while import is running."""
        for widget in (
            self.edit_quantity_button,
            self.remove_formulation_button,
            self.export_excel_button,
            self.toggle_export_button,
            self.normalize_total_button,
            self.quantity_mode_selector,
        ):
            if widget is not None:
                widget.setEnabled(enabled)


    def _prefill_import_state(
        self, base_items: list[Dict[str, Any]], meta: Dict[str, Any]
    ) -> None:
        """Populate UI with base items before USDA hydration."""
        mode = meta.get("quantity_mode", "g")
        self.quantity_mode_selector.blockSignals(True)
        self._set_quantity_mode(mode)
        self.quantity_mode_selector.blockSignals(False)

        formula_name = meta.get("formula_name", "")
        respect_existing = meta.get("respect_existing_formula_name", False)
        if not respect_existing or not self.formula_name_input.text().strip():
            self.formula_name_input.setText(formula_name)

        label_settings = meta.get("label_settings") or {}
        if label_settings:
            self._apply_label_settings(label_settings, defer_preview=True)

        ui_items: list[Dict[str, Any]] = []
        for item in base_items:
            fdc_id = item.get("fdc_id")
            description = (item.get("description") or "").strip()
            if not description:
                description = f"FDC {fdc_id}"
            data_type = (item.get("data_type") or "").strip() or "Imported"
            ui_items.append(
                {
                    "fdc_id": fdc_id,
                    "description": description,
                    "brand": item.get("brand", "") or "",
                    "data_type": data_type,
                    "amount_g": float(item.get("amount_g", 0.0) or 0.0),
                    "locked": bool(item.get("locked", False)),
                    "nutrients": [],
                }
            )

        self.formulation_presenter.load_from_ui_items(
            ui_items,
            self.formula_name_input.text() or "Imported",
        )
        self._last_totals = {}
        self.details_table.setRowCount(0)
        self.totals_table.setRowCount(0)
        self._populate_formulation_tables()
        self._ensure_preview_selection()
        self._update_label_preview(force_recalc_totals=True)


    def _start_import_hydration(
        self, base_items: list[Dict[str, Any]], meta: Dict[str, Any]
    ) -> None:
        if not base_items:
            QMessageBox.warning(
                self,
                "Sin ingredientes",
                "El archivo no contiene ingredientes para importar.",
            )
            return

        self.import_state_button.setEnabled(False)
        self.export_state_button.setEnabled(False)
        self._set_import_controls_enabled(False)
        self._suspend_no_sig_update = True
        self.status_label.setText("Importando ingredientes...")
        self._set_window_progress("Importando ingredientes")

        self._pending_import_meta = meta
        self._prefill_import_state(base_items, meta)

        thread = QThread(self)
        worker = ImportWorker(
            lambda: self.container.food_repository,
            base_items,
            max_attempts=self.import_max_attempts,
            read_timeout=self.import_read_timeout,
        )
        worker.moveToThread(thread)
        self._workers.append(worker)
        self._threads.append(thread)
        self._current_import_worker = worker

        thread.started.connect(worker.run)
        worker.progress.connect(self._on_import_progress)
        worker.finished.connect(self._on_import_finished)
        worker.error.connect(self._on_import_error)
        worker.finished.connect(thread.quit)
        worker.error.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        worker.error.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)

        def _cleanup() -> None:
            if thread in self._threads:
                self._threads.remove(thread)
            if worker in self._workers:
                self._workers.remove(worker)
            if self._current_import_worker is worker:
                self._current_import_worker = None

        thread.finished.connect(_cleanup)
        thread.start()


    def _on_import_progress(self, message: str) -> None:
        self._set_window_progress(message)
        self.status_label.setText(f"Importando ingredientes: {message}")


    def _on_import_finished(self, payload: list[Dict[str, Any]]) -> None:
        meta = getattr(self, "_pending_import_meta", {}) or {}
        self._pending_import_meta = {}
        if QThread.currentThread() is not self.thread():
            QTimer.singleShot(0, lambda p=payload, m=meta: self._on_import_finished(p))
            return
        self._reset_import_ui_state()
        hydrated: list[Dict[str, Any]] = []
        for entry in payload:
            base = entry.get("base") or {}
            details = entry.get("details") or {}
            fdc_id = base.get("fdc_id") or details.get("fdcId")
            try:
                fdc_id_int = int(fdc_id) if fdc_id is not None else None
            except Exception:
                fdc_id_int = fdc_id

            nutrients = augment_fat_nutrients(details.get("foodNutrients", []) or [])
            self.nutrient_ordering.update_reference_from_details(details)
            hydrated.append(
                {
                    "fdc_id": fdc_id_int,
                    "description": details.get("description", "") or base.get("description", ""),
                    "brand": details.get("brandOwner", "") or base.get("brand", ""),
                    "data_type": details.get("dataType", "") or base.get("data_type", ""),
                    "amount_g": float(base.get("amount_g", 0.0) or 0.0),
                    "nutrients": normalize_nutrients(
                        nutrients, details.get("dataType")
                    ),
                    "locked": bool(base.get("locked", False)),
                }
            )

        self.formulation_presenter.load_from_ui_items(
            hydrated, self.formula_name_input.text() or "Imported"
        )
        resolved_flags: Dict[str, bool] = {}
        resolved_flags.update(
            self.formulation_presenter.resolve_legacy_export_flags(
                meta.get("legacy_nutrient_export_flags", {}),
                hydrated,
            )
        )
        resolved_flags.update(meta.get("nutrient_export_flags", {}))
        self.nutrient_export_flags = resolved_flags
        mode = meta.get("quantity_mode", "g")
        self.quantity_mode_selector.blockSignals(True)
        self._set_quantity_mode(mode)
        self.quantity_mode_selector.blockSignals(False)

        formula_name = meta.get("formula_name", "")
        respect_existing = meta.get("respect_existing_formula_name", False)
        if not respect_existing or not self.formula_name_input.text().strip():
            self.formula_name_input.setText(formula_name)

        self._refresh_formulation_views()
        source = meta.get("path", "archivo")
        self.status_label.setText(f"Formulación importada desde {source}")


    def _on_import_error(self, message: str) -> None:
        self._reset_import_ui_state()
        self.status_label.setText("Error al importar ingredientes.")
        QMessageBox.critical(self, "Error al cargar ingrediente", message)


    def _reset_import_ui_state(self) -> None:
        self.import_state_button.setEnabled(True)
        self.export_state_button.setEnabled(True)
        self._set_import_controls_enabled(True)
        self._suspend_no_sig_update = False
        self._set_window_progress(None)
        self._current_import_worker = None


    def _total_weight(self) -> float:
        """Total weight of current formulation in grams."""
        return self.formulation_presenter.get_total_weight()

    def _is_percent_mode(self) -> bool:
        return self.quantity_mode == "%"

    def _current_mass_unit(self) -> str:
        if self._is_percent_mode():
            return "g"
        return normalize_mass_unit(self.quantity_mode) or "g"

    def _quantity_mode_label(self, mode: str) -> str:
        labels = {
            "g": "gramos (g)",
            "kg": "kilogramos (kg)",
            "ton": "toneladas (ton)",
            "lb": "libras (lb)",
            "oz": "onzas (oz)",
            "%": "porcentaje (%)",
        }
        return labels.get(mode, mode)

    def _set_quantity_mode(self, mode_raw: str) -> None:
        mode_lower = str(mode_raw or "g").strip().lower()
        if mode_lower in ("%", "percent", "percentage"):
            mode = "%"
        else:
            mode = normalize_mass_unit(mode_lower) or "g"
        self.quantity_mode = mode
        idx = next(
            (
                i
                for i, (value, _) in enumerate(self.quantity_mode_options)
                if value == mode
            ),
            0,
        )
        self.quantity_mode_selector.setCurrentIndex(idx)

    def _mass_decimals(self, unit: str) -> int:
        return {
            "g": 1,
            "kg": 3,
            "ton": 6,
            "lb": 3,
            "oz": 3,
        }.get(unit, 2)

    def _display_amount_for_unit(self, amount_g: float) -> float:
        unit = self._current_mass_unit()
        converted = convert_mass(amount_g, "g", unit)
        return float(converted) if converted is not None else amount_g

    def _format_mass_amount(self, amount_g: float) -> str:
        unit = self._current_mass_unit()
        converted = self._display_amount_for_unit(amount_g)
        decimals = self._mass_decimals(unit)
        return f"{converted:.{decimals}f} {unit}"


    def _amount_to_percent(self, amount_g: float, total: float | None = None) -> float:
        total_weight = self._total_weight() if total is None else total
        if total_weight <= 0:
            return 0.0
        return (amount_g / total_weight) * 100.0


    def _update_quantity_headers(self) -> None:
        mass_unit = self._current_mass_unit()
        self.formulation_preview.setHorizontalHeaderLabels(
            ["FDC ID", "Ingrediente"]
        )
        self.formulation_table.setHorizontalHeaderLabels(
            [
                "FDC ID",
                "Ingrediente",
                f"Cantidad ({mass_unit})",
                "Cantidad (%)",
                "Fijar %",
                "Marca / Origen",
            ]
        )


    def _set_item_enabled(self, item: QTableWidgetItem | None, enabled: bool) -> None:
        if item is None:
            return
        flags = item.flags()
        if enabled:
            flags |= Qt.ItemIsEnabled
        else:
            flags &= ~Qt.ItemIsEnabled
        item.setFlags(flags)
        item.setBackground(QColor("#f0f0f0") if not enabled else QColor("white"))


    def _apply_column_state(self, table: QTableWidget, row: int) -> None:
        grams_enabled = not self._is_percent_mode()
        percent_enabled = self._is_percent_mode()

        self._set_item_enabled(
            table.item(row, self.amount_g_column_index), grams_enabled
        )
        self._set_item_enabled(
            table.item(row, self.percent_column_index), percent_enabled
        )
        self._set_item_enabled(
            table.item(row, self.lock_column_index), percent_enabled
        )


    def _can_edit_column(self, column: int | None) -> bool:
        if column is None:
            return True
        if not self._is_percent_mode():
            return column == self.amount_g_column_index
        return column == self.percent_column_index


    def _nutrients_by_header(
        self, nutrients: List[Dict[str, Any]], header_by_key: Dict[str, str]
    ) -> Dict[str, float]:
        """
        Build a mapping of template header -> nutrient amount (per 100 g),
        aligned to a precomputed header key map to avoid duplicate columns.
        """
        out: Dict[str, float] = {}
        best_priority: Dict[str, int] = {}
        allowed_keys = set(header_by_key.keys())
        alias_priority = {
            "carbohydrate, by difference": 2,
            "carbohydrate, by summation": 1,
            "carbohydrate by summation": 1,
            "sugars, total": 2,
            "total sugars": 1,
        }

        for entry in nutrients:
            amount = entry.get("amount")
            if amount is None:
                continue
            nut = entry.get("nutrient") or {}
            header_key, name, unit = self.nutrient_ordering.header_key(nut)
            if not header_key or header_key not in allowed_keys:
                continue
            header = header_by_key[header_key]
            priority = alias_priority.get((nut.get("name") or "").strip().lower(), 0)
            current_best = best_priority.get(header, -1)
            if priority < current_best:
                continue

            best_priority[header] = priority
            out[header] = amount

        return out


    def _collect_nutrient_columns(self) -> tuple[list[str], Dict[str, str], Dict[str, str]]:
        """
        Collect ordered nutrient headers and their categories.
        Uses the static catalog to force a default ordering/grouping and only keeps
        nutrients that appear with a value in any ingredient.
        """
        candidates: Dict[str, Dict[str, Any]] = {}
        categories_seen_order: Dict[str, int] = {}
        preferred_order = [cat for cat, _ in self.nutrient_ordering.catalog]
        preferred_count = len(preferred_order)

        for item in self.formulation_presenter.get_ui_items():
            data_priority = DATA_TYPE_PRIORITY.get(
                (item.get("data_type") or "").strip(), len(DATA_TYPE_PRIORITY)
            )
            for entry in self.nutrient_ordering.sort_nutrients_for_display(
                item.get("nutrients", [])
            ):
                nut = entry.get("nutrient") or {}
                amount = entry.get("amount")
                if amount is None:
                    continue
                header_key, canonical_name, canonical_unit = self.nutrient_ordering.header_key(nut)
                if header_key and not self.nutrient_export_flags.get(header_key, True):
                    continue

                if not header_key or not canonical_name:
                    continue

                category = self.nutrient_ordering.category_for_nutrient(canonical_name, nut)
                if category not in categories_seen_order:
                    categories_seen_order[category] = len(categories_seen_order)

                order = self.nutrient_ordering.order_for_name(canonical_name)
                if order is None:
                    order = self.nutrient_ordering.nutrient_order(nut, len(candidates))
                # Keep kcal ahead of kJ when both present
                unit_lower = (canonical_unit or "").strip().lower()
                if canonical_name.strip().lower() == "energy":
                    if unit_lower == "kcal":
                        order = order - 0.1 if isinstance(order, (int, float)) else order
                    elif unit_lower == "kj":
                        order = order + 0.1 if isinstance(order, (int, float)) else order

                header = (
                    f"{canonical_name} ({canonical_unit})"
                    if canonical_unit
                    else canonical_name
                )

                existing = candidates.get(header_key)
                if existing is None or (
                    data_priority < existing["data_priority"]
                    or (
                        data_priority == existing["data_priority"]
                        and order < existing["order"]
                    )
                    or (
                        data_priority == existing["data_priority"]
                        and order == existing["order"]
                        and header < existing["header"]
                    )
                ):
                    candidates[header_key] = {
                        "header_key": header_key,
                        "header": header,
                        "category": category,
                        "order": order,
                        "data_priority": data_priority,
                    }

        def category_rank(cat: str) -> int:
            if cat in preferred_order:
                return preferred_order.index(cat)
            return preferred_count + categories_seen_order.get(cat, preferred_count)

        sorted_candidates = sorted(
            candidates.values(),
            key=lambda c: (
                category_rank(c["category"]),
                c["order"],
                c["header"].lower(),
            ),
        )

        ordered_headers: list[str] = [c["header"] for c in sorted_candidates]
        categories: Dict[str, str] = {c["header"]: c["category"] for c in sorted_candidates}
        header_key_map: Dict[str, str] = {c["header"]: c["header_key"] for c in sorted_candidates}

        return ordered_headers, categories, header_key_map


    def _ensure_normalized_items(self) -> None:
        """Normalize all ingredients' nutrients in-place (fat + energy)."""
        for idx, item in enumerate(self.formulation_presenter.get_ui_items()):
            original = item.get("nutrients", []) or []
            normalized = normalize_nutrients(original, item.get("data_type"))
            if normalized != original:
                # preserve reference to allow downstream updates
                item["nutrients"] = normalized


    def _split_header_unit(self, header: str) -> tuple[str, str]:
        if header.endswith(")") and " (" in header:
            name, unit = header.rsplit(" (", 1)
            return name, unit[:-1]
        return header, ""


    def _hydrate_items(self, items: list[Dict[str, Any]]) -> list[Dict[str, Any]] | None:
        """Fetch USDA details for items to populate description/nutrients."""
        hydrated: list[Dict[str, Any]] = []
        total = len(items)
        for item in items:
            fdc_id = item.get("fdc_id") or item.get("fdcId")
            if fdc_id is None:
                QMessageBox.warning(
                    self,
                    "FDC ID faltante",
                    "Uno de los ingredientes no tiene FDC ID.",
                )
                return None
            try:
                fdc_id_int = int(fdc_id)
            except ValueError:
                QMessageBox.warning(
                    self,
                    "FDC ID inválido",
                    f"FDC ID no numérico: {fdc_id}",
                )
                return None

            self.status_label.setText(
                f"Importando ingrediente {len(hydrated)+1}/{total} (FDC {fdc_id_int})..."
            )
            QCoreApplication.processEvents()
            try:
                details = self.food_repository.get_by_id(
                    fdc_id_int,
                    detail_format="abridged",
                )
            except Exception as exc:  # noqa: BLE001 - surface to user
                QMessageBox.critical(
                    self,
                    "Error al cargar ingrediente",
                    f"No se pudo cargar el FDC {fdc_id_int}:\n{exc}",
                )
                return None
            self.nutrient_ordering.update_reference_from_details(details)

            hydrated.append(
                {
                    "fdc_id": fdc_id_int,
                    "description": details.get("description", "")
                    or item.get("description", ""),
                    "brand": details.get("brandOwner", "") or item.get("brand", ""),
                    "data_type": details.get("dataType", "") or item.get("data_type", ""),
                    "amount_g": float(item.get("amount_g", 0.0) or 0.0),
                    "nutrients": normalize_nutrients(
                        details.get("foodNutrients", []) or [], details.get("dataType")
                    ),
                    "locked": bool(item.get("locked", False)),
                }
            )

        return hydrated


    def _populate_formulation_tables(self) -> None:
        """Refresh formulation tables with current items."""
        logging.debug(f"_populate_formulation_tables rows={self.formulation_presenter.get_ingredient_count()}")
        self._update_quantity_headers()
        total_weight = self._total_weight()

        # Left preview (ID + Ingrediente)
        self.formulation_preview.blockSignals(True)
        self.formulation_preview.setRowCount(self.formulation_presenter.get_ingredient_count())
        for idx, item in enumerate(self.formulation_presenter.get_ui_items()):
            # Locked default handled by domain model
            self.formulation_preview.setItem(
                idx, 0, QTableWidgetItem(str(item.get("fdc_id", "")))
            )
            self.formulation_preview.setItem(
                idx, 1, QTableWidgetItem(item.get("description", ""))
            )
        self.formulation_preview.blockSignals(False)

        # Main formulation table
        self.formulation_table.blockSignals(True)
        self.formulation_table.setRowCount(self.formulation_presenter.get_ingredient_count())
        for idx, item in enumerate(self.formulation_presenter.get_ui_items()):
            amount_g = float(item.get("amount_g", 0.0) or 0.0)
            percent = self._amount_to_percent(amount_g, total_weight)
            amount_display = self._display_amount_for_unit(amount_g)
            amount_decimals = self._mass_decimals(self._current_mass_unit())

            cells: list[QTableWidgetItem] = [
                QTableWidgetItem(str(item.get("fdc_id", ""))),
                QTableWidgetItem(item.get("description", "")),
                QTableWidgetItem(f"{amount_display:.{amount_decimals}f}"),
                QTableWidgetItem(f"{percent:.2f}"),
            ]

            lock_item = QTableWidgetItem("")
            lock_item.setFlags(
                Qt.ItemIsSelectable | Qt.ItemIsUserCheckable | Qt.ItemIsEnabled
            )
            lock_item.setCheckState(
                Qt.Checked if item.get("locked") else Qt.Unchecked
            )
            cells.append(lock_item)

            cells.append(QTableWidgetItem(item.get("brand", "")))

            for col, cell in enumerate(cells):
                self.formulation_table.setItem(idx, col, cell)

            self._apply_column_state(self.formulation_table, idx)
        self.formulation_table.blockSignals(False)
        logging.debug("_populate_formulation_tables done")


    def _populate_totals_table(self) -> None:
        logging.debug("_populate_totals_table start")
        totals = self._calculate_totals()
        self._last_totals = totals

        category_order = [cat for cat, _ in self.nutrient_ordering.catalog]

        def _cat_rank(name: str) -> int:
            cat = self.nutrient_ordering.category_for_nutrient(name)
            if cat in category_order:
                return category_order.index(cat)
            return len(category_order) + 1

        def _order_val(name: str) -> float:
            order = self.nutrient_ordering.order_for_name(name)
            return float(order if order is not None else float("inf"))

        sorted_totals = sorted(
            totals.items(),
            key=lambda item: (
                _cat_rank(item[1].get("name", "")),
                _order_val(item[1].get("name", "")),
                item[1].get("name", "").lower(),
            ),
        )

        self.totals_table.blockSignals(True)
        self.totals_table.setRowCount(len(sorted_totals))
        new_flags: Dict[str, bool] = {}

        for row_idx, (nut_key, entry) in enumerate(sorted_totals):
            name_item = QTableWidgetItem(entry["name"])
            name_item.setData(Qt.UserRole, nut_key)
            self.totals_table.setItem(row_idx, 0, name_item)

            self.totals_table.setItem(
                row_idx, 1, QTableWidgetItem(f"{entry['amount']:.2f}")
            )
            self.totals_table.setItem(row_idx, 2, QTableWidgetItem(entry["unit"]))

            export_item = QTableWidgetItem("")
            export_item.setFlags(
                Qt.ItemIsUserCheckable | Qt.ItemIsEnabled | Qt.ItemIsSelectable
            )
            current_checked = self.nutrient_export_flags.get(nut_key, True)
            export_item.setCheckState(Qt.Checked if current_checked else Qt.Unchecked)
            export_item.setData(Qt.UserRole, nut_key)
            self.totals_table.setItem(row_idx, 3, export_item)
            new_flags[nut_key] = current_checked

        self.nutrient_export_flags = new_flags
        self.totals_table.blockSignals(False)
        self._update_toggle_export_button()
        logging.debug("_populate_totals_table done")


    def _update_toggle_export_button(self) -> None:
        if not hasattr(self, "toggle_export_button"):
            return
        all_checked = self.nutrient_export_flags and all(
            self.nutrient_export_flags.values()
        )
        text = "Deseleccionar todos" if all_checked else "Seleccionar todos"
        self.toggle_export_button.setText(text)


    def _create_question_icon(self) -> QIcon:
        """Create a small yellow square icon with a '?' centered."""
        size = 16
        pixmap = QPixmap(size, size)
        pixmap.fill(Qt.transparent)
        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.fillRect(0, 0, size, size, QColor(255, 236, 179))
        painter.setPen(Qt.black)
        font: QFont = painter.font()
        font.setBold(True)
        font.setPointSize(10)
        painter.setFont(font)
        painter.drawText(pixmap.rect(), Qt.AlignCenter, "?")
        painter.end()
        return QIcon(pixmap)


    def _refresh_formulation_views(self) -> None:
        logging.debug(f"_refresh_formulation_views count={self.formulation_presenter.get_ingredient_count()}")
        if QThread.currentThread() is not self.thread():
            QTimer.singleShot(0, self._refresh_formulation_views)
            return
        self._ensure_normalized_items()
        self._populate_formulation_tables()
        self._populate_totals_table()
        self._update_label_preview(force_recalc_totals=True)
        logging.debug("_refresh_formulation_views done")
        self._ensure_preview_selection()


    def _select_preview_row(self, row: int) -> None:
        if row < 0 or row >= self.formulation_preview.rowCount():
            return
        sel_model = self.formulation_preview.selectionModel()
        if sel_model is None:
            return
        sel_model.clearSelection()
        index = self.formulation_preview.model().index(row, 0)
        sel_model.select(
            index, QItemSelectionModel.ClearAndSelect | QItemSelectionModel.Rows
        )
        self.formulation_preview.setCurrentIndex(index)


    def _ensure_preview_selection(self) -> None:
        if not self.formulation_presenter.has_ingredients():
            self.details_table.setRowCount(0)
            return
        self._ensure_normalized_items()
        sel_model = self.formulation_preview.selectionModel()
        has_sel = sel_model and sel_model.hasSelection()
        if not has_sel:
            self._select_preview_row(self.formulation_presenter.get_ingredient_count() - 1)
        self._show_nutrients_for_selected_preview()


    def _show_nutrients_for_row(self, row: int) -> None:
        if row < 0 or row >= self.formulation_presenter.get_ingredient_count():
            self.details_table.setRowCount(0)
            return
        nutrients = self.formulation_presenter.get_ui_item(row).get("nutrients", []) or []
        self._populate_details_table(nutrients)


    def _show_nutrients_for_selected_preview(self) -> None:
        sel_model = self.formulation_preview.selectionModel()
        if not sel_model or not sel_model.hasSelection():
            self._show_nutrients_for_row(-1)
            return
        row = sel_model.selectedRows()[0].row()
        self._show_nutrients_for_row(row)


    def _export_formulation_to_excel(self, filepath: str) -> None:
        """Export formulation to Excel via presenter."""
        # Update formulation name from UI input
        formulation_name = self.formula_name_input.text() or "Current Formulation"
        self.formulation_presenter.formulation_name = formulation_name

        try:
            self.formulation_presenter.export_to_excel(
                filepath,
                export_flags=self.nutrient_export_flags,
                mass_unit=self._current_mass_unit(),
            )
            return
        except Exception as e:
            logging.error(
                f"Error exporting via presenter: {e}, falling back to old method"
            )

        # Fallback to old implementation
        self._ensure_normalized_items()
        wb = Workbook()
        ws = wb.active
        ws.title = "Ingredientes"
        totals_sheet = wb.create_sheet("Totales")

        mass_unit = self._current_mass_unit()
        unit_decimals = self._mass_decimals(mass_unit)
        unit_format = "0" if unit_decimals <= 0 else f"0.{'0' * unit_decimals}"

        base_headers = [
            "FDC ID",
            "Ingrediente",
            "Marca / Origen",
            "Tipo de dato",
            f"Cantidad ({mass_unit})",
            "Cantidad (%)",
        ]

        nutrient_headers, header_categories, header_key_map = self._collect_nutrient_columns()
        header_by_key = {v: k for k, v in header_key_map.items()}

        header_fill = PatternFill("solid", fgColor="D9D9D9")
        total_fill = PatternFill("solid", fgColor="FFF2CC")
        category_fills = {
            "Proximates": PatternFill("solid", fgColor="DAEEF3"),
            "Carbohydrates": PatternFill("solid", fgColor="E6B8B7"),
            "Minerals": PatternFill("solid", fgColor="C4D79B"),
            "Vitamins and Other Components": PatternFill("solid", fgColor="FFF2CC"),
            "Lipids": PatternFill("solid", fgColor="D9E1F2"),
            "Amino acids": PatternFill("solid", fgColor="E4DFEC"),
        }
        center = Alignment(horizontal="center", vertical="center")

        # Row 1: group titles (base + nutrient categories)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(base_headers))
        ws.cell(row=1, column=1, value="Detalles de formulación").alignment = center

        if nutrient_headers:
            start_col = len(base_headers) + 1
            col = start_col
            while col < start_col + len(nutrient_headers):
                idx = col - start_col
                category = header_categories.get(nutrient_headers[idx], "Nutrientes")
                run_start = col
                while (
                    col < start_col + len(nutrient_headers)
                    and header_categories.get(nutrient_headers[col - start_col], "Nutrientes")
                    == category
                ):
                    col += 1
                run_end = col - 1
                ws.merge_cells(start_row=1, start_column=run_start, end_row=1, end_column=run_end)
                cat_cell = ws.cell(row=1, column=run_start, value=category)
                cat_cell.alignment = center
                cat_cell.fill = category_fills.get(category, header_fill)

        # Row 2: headers
        headers = base_headers + nutrient_headers
        for col_idx, name in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=name)
            cell.fill = category_fills.get(
                header_categories.get(name, ""), header_fill
            ) if col_idx > len(base_headers) else header_fill
            cell.alignment = center

        start_row = 3
        grams_col = base_headers.index(f"Cantidad ({mass_unit})") + 1
        percent_col = base_headers.index("Cantidad (%)") + 1
        data_rows = self.formulation_presenter.get_ingredient_count()
        end_row = start_row + data_rows - 1

        # Write ingredient rows
        for idx, item in enumerate(self.formulation_presenter.get_ui_items()):
            row = start_row + idx
            amount_g = float(item.get("amount_g", 0.0) or 0.0)
            amount_val = convert_mass(amount_g, "g", mass_unit)
            amount_display = float(amount_val) if amount_val is not None else amount_g
            values = [
                item.get("fdc_id", ""),
                item.get("description", ""),
                item.get("brand", ""),
                item.get("data_type", ""),
                amount_display,
                None,  # placeholder for percent formula
            ]
            for col_idx, val in enumerate(values, start=1):
                ws.cell(row=row, column=col_idx, value=val)

            gram_cell = f"{get_column_letter(grams_col)}{row}"
            total_range = f"${get_column_letter(grams_col)}${start_row}:${get_column_letter(grams_col)}${end_row}"
            ws.cell(
                row=row,
                column=percent_col,
                value=f"={gram_cell}/SUM({total_range})",
            ).number_format = "0.00%"

            nut_map = self._nutrients_by_header(item.get("nutrients", []), header_by_key)
            for offset, header in enumerate(nutrient_headers, start=len(base_headers) + 1):
                if header in nut_map:
                    ws.cell(row=row, column=offset, value=nut_map[header])

        total_row = end_row + 1
        ws.cell(row=total_row, column=1, value="Total")
        ws.cell(row=total_row, column=2, value="Formulado")
        ws.cell(row=total_row, column=3, value="Formulado")
        ws.cell(row=total_row, column=4, value="Formulado")

        gram_total_cell = ws.cell(
            row=total_row,
            column=grams_col,
            value=f"=SUBTOTAL(9,{get_column_letter(grams_col)}{start_row}:{get_column_letter(grams_col)}{end_row})",
        )
        gram_total_cell.fill = total_fill
        percent_total_cell = ws.cell(row=total_row, column=percent_col, value="100%")
        percent_total_cell.number_format = "0.00%"
        percent_total_cell.fill = total_fill

        for offset, header in enumerate(nutrient_headers, start=len(base_headers) + 1):
            col_letter = get_column_letter(offset)
            formula = (
                f"=SUMPRODUCT(${get_column_letter(percent_col)}${start_row}:${get_column_letter(percent_col)}${end_row},"
                f"${col_letter}${start_row}:${col_letter}${end_row})"
            )
            cell = ws.cell(row=total_row, column=offset, value=formula)
            cell.fill = total_fill

        # Styles for data rows
        for row in range(start_row, total_row + 1):
            ws.cell(row=row, column=grams_col).number_format = unit_format

        # Freeze panes to keep headers/base columns visible
        freeze_col = len(base_headers) + 1 if nutrient_headers else 1
        ws.freeze_panes = f"{get_column_letter(freeze_col)}3"

        # Adjust widths
        widths = {
            "A": 12,
            "B": 35,
            "C": 18,
            "D": 14,
            "E": 12,
            "F": 12,
        }
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

        # Totales sheet: simple reference to totals row
        totals_headers = ["Nutriente", "Total", "Unidad"]
        for col_idx, name in enumerate(totals_headers, start=1):
            cell = totals_sheet.cell(row=1, column=col_idx, value=name)
            cell.fill = header_fill
            cell.alignment = center

        for idx, header in enumerate(nutrient_headers, start=1):
            name_part, unit = self._split_header_unit(header)
            totals_sheet.cell(row=idx + 1, column=1, value=name_part)
            totals_sheet.cell(row=idx + 1, column=3, value=unit or "")
            source_cell = f"Ingredientes!{get_column_letter(len(base_headers) + idx)}{total_row}"
            totals_sheet.cell(row=idx + 1, column=2, value=f"={source_cell}")

        wb.save(filepath)


    def _calculate_totals(self) -> Dict[str, Dict[str, Any]]:
        """Calculate nutrient totals via presenter."""
        logging.debug(f"_calculate_totals start items={self.formulation_presenter.get_ingredient_count()}")

        try:
            # Calculate totals via presenter
            totals = self.formulation_presenter.calculate_totals()
            totals = self.nutrient_ordering.normalize_totals_by_header_key(totals)
            logging.debug(f"_calculate_totals done (via presenter) nutrients={len(totals)}")
            return totals

        except Exception as e:
            logging.error(f"Error calculating totals via presenter: {e}")
            # Fallback to old implementation
            self._ensure_normalized_items()
            totals: Dict[str, Dict[str, Any]] = {}
            total_weight = self._total_weight()
            for item in self.formulation_presenter.get_ui_items():
                qty = item.get("amount_g", 0) or 0
                for nutrient in self.nutrient_ordering.sort_nutrients_for_display(
                    item.get("nutrients", [])
                ):
                    amount = nutrient.get("amount")
                    if amount is None:
                        continue
                    nut = nutrient.get("nutrient") or {}
                    header_key, canonical_name, canonical_unit = self.nutrient_ordering.header_key(nut)
                    if not header_key:
                        continue
                    entry = totals.setdefault(
                        header_key,
                        {
                            "name": canonical_name or nut.get("name", ""),
                            "unit": canonical_unit or "",
                            "amount": 0.0,
                            "order": self.nutrient_ordering.nutrient_order(
                                nut, len(totals)
                            ),
                        },
                    )
                    if canonical_name and not entry["name"]:
                        entry["name"] = canonical_name
                    if canonical_unit and not entry["unit"]:
                        entry["unit"] = canonical_unit
                    inferred_unit = self.nutrient_ordering.infer_unit(nut)
                    if inferred_unit and not entry["unit"]:
                        entry["unit"] = inferred_unit
                    entry["order"] = min(
                        entry.get("order", float("inf")),
                        self.nutrient_ordering.nutrient_order(nut, len(totals)),
                    )
                    entry["amount"] += amount * qty / 100.0

            if total_weight > 0:
                factor = 100.0 / total_weight
                for entry in totals.values():
                    entry["amount"] *= factor
            logging.debug(f"_calculate_totals done (fallback) nutrients={len(totals)} total_weight={total_weight}")
            return totals


    # ---- Add/edit flows ----
    def _add_row_to_formulation(self, row: int | None = None) -> None:
        logging.debug(f"_add_row_to_formulation row={row}")
        if row is None:
            indexes = self.table.selectionModel().selectedRows()
            if not indexes:
                self.status_label.setText("Selecciona un alimento para agregar.")
                return
            row = indexes[0].row()

        fdc_item = self.table.item(row, 0)
        if not fdc_item:
            return
        fdc_id_text = fdc_item.text().strip()
        if not fdc_id_text or not fdc_id_text.isdigit():
            self.status_label.setText("FDC ID inválido.")
            return

        mode, value = self._prompt_quantity()
        logging.debug(f"_prompt_quantity returned mode={mode} value={value}")
        if mode is None:
            return

        display_amount = (
            self._format_amount_for_status(value, include_new=True)
            if mode == "g"
            else f"{value:.2f} %"
        )
        self.status_label.setText(
            f"Agregando {fdc_id_text} ({display_amount})..."
        )
        self.add_button.setEnabled(False)
        self._start_add_fetch(int(fdc_id_text), mode, value)


    def _start_add_fetch(self, fdc_id: int, mode: str, value: float) -> None:
        logging.debug(f"_start_add_fetch fdc_id={fdc_id} mode={mode} value={value}")
        self._set_window_progress(f"1/1 ID #{fdc_id}")
        thread = QThread(self)
        worker = AddWorker(
            lambda: self.container.food_repository,
            fdc_id,
            max_attempts=self.import_max_attempts,
            read_timeout=self.import_read_timeout,
            mode=mode,
            value=value,
        )
        worker.moveToThread(thread)
        self._workers.append(worker)
        self._threads.append(thread)
        self._current_add_worker = worker

        thread.started.connect(worker.run)
        worker.progress.connect(self._on_add_progress)
        worker.finished.connect(self._on_add_finished)
        worker.error.connect(self._on_add_error)
        worker.finished.connect(thread.quit)
        worker.error.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        worker.error.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)

        def _cleanup() -> None:
            if thread in self._threads:
                self._threads.remove(thread)
            if worker in self._workers:
                self._workers.remove(worker)
            if self._current_add_worker is worker:
                self._current_add_worker = None

        thread.finished.connect(_cleanup)
        thread.start()


    def _on_add_progress(self, message: str) -> None:
        self._set_window_progress(message)
        self.status_label.setText(f"Agregando ingrediente: {message}")


    def _on_add_finished(self, details, mode: str, value: float) -> None:
        logging.debug(
            f"_on_add_finished fdc_id={details.get('fdcId', '?')} "
            f"mode={mode} value={value} nutrients={len(details.get('foodNutrients', []) or [])}"
        )
        self._reset_add_ui_state()
        self._on_add_details_loaded(details, mode, value)


    def _reset_add_ui_state(self) -> None:
        self.add_button.setEnabled(True)
        self._set_window_progress(None)
        if self._current_add_worker:
            self._current_add_worker = None


    def _format_amount_for_status(self, amount_g: float, include_new: bool = False) -> str:
        """Return a user-facing label for a quantity using the active mode."""
        if not self._is_percent_mode():
            return self._format_mass_amount(amount_g)

        total = self._total_weight()
        if include_new:
            total += amount_g
        percent = 0.0 if total <= 0 else (amount_g / total) * 100.0
        return f"{percent:.2f} %"


    def _safe_base_name(self, fallback: str = "formulacion") -> str:
        """Return a filesystem-safe base name using the formula input or fallback."""
        name = (self.formula_name_input.text() or "").strip()
        clean = re.sub(r'[\\/:*?"<>|]+', "_", name).strip(". ")
        return clean or fallback


    def _prompt_quantity(
        self, default_amount: float | None = None, editing_index: int | None = None
    ) -> tuple[str | None, float]:
        if not self._is_percent_mode():
            unit = self._current_mass_unit()
            start_amount_g = default_amount if default_amount is not None else 100.0
            start_value = convert_mass(start_amount_g, "g", unit)
            if start_value is None:
                start_value = start_amount_g
            title = "Cantidad"
            label = f"Cantidad del ingrediente ({unit}):"
            min_value = convert_mass(0.1, "g", unit)
            if min_value is None:
                min_value = 0.1
            max_value = convert_mass(1_000_000.0, "g", unit)
            if max_value is None:
                max_value = 1_000_000.0
            decimals = self._mass_decimals(unit)
        else:
            start_value = (
                self._amount_to_percent(default_amount or 0.0)
                if default_amount is not None
                else 10.0
            )
            title = "Porcentaje"
            label = "Porcentaje del ingrediente sobre la formulación (%):"
            min_value, max_value, decimals = 0.01, 100.0, 2

        raw_value, ok = QInputDialog.getDouble(
            self,
            title,
            label,
            start_value,
            min_value,
            max_value,
            decimals,
        )
        if not ok:
            return None, 0.0

        if self._is_percent_mode():
            return "percent", raw_value

        unit = self._current_mass_unit()
        amount_g = convert_mass(raw_value, unit, "g")
        return "g", float(amount_g) if amount_g is not None else raw_value


    def _edit_quantity_for_row(self, row: int) -> None:
        """Prompt the user to update the quantity of a specific row."""
        if row < 0 or row >= self.formulation_presenter.get_ingredient_count():
            self.status_label.setText("Fila seleccionada inválida.")
            return

        item = self.formulation_presenter.get_ui_item(row)
        default_amount = item.get("amount_g", 0.0)
        mode, value = self._prompt_quantity(default_amount, editing_index=row)
        if mode is None:
            return

        if mode == "g":
            # Update the domain model directly so the change persists.
            self.formulation_presenter.update_ingredient_amount(
                row,
                value,
                maintain_total=False,
            )
        else:
            if not self._apply_percent_edit(row, value):
                return

        self._refresh_formulation_views()
        if mode == "g":
            msg_value = self._format_amount_for_status(value)
        else:
            msg_value = f"{value:.2f} %"
        self.status_label.setText(
            f"Actualizado {item.get('fdc_id', '')} a {msg_value}"
        )


    def _apply_percent_edit(self, target_idx: int, target_percent: float) -> bool:
        """Redistribute quantities so locked percentages stay fixed."""
        if target_percent < 0 or target_percent > 100:
            QMessageBox.warning(self, "Porcentaje inválido", "Ingresa un valor entre 0 y 100.")
            return False

        if target_idx < 0 or target_idx >= self.formulation_presenter.get_ingredient_count():
            self.status_label.setText("Fila seleccionada inválida.")
            return False

        total_weight = self._total_weight()
        base_total = total_weight if total_weight > 0 else 100.0
        if base_total <= 0:
            QMessageBox.warning(
                self,
                "Porcentaje inválido",
                "No hay cantidad total para calcular porcentajes.",
            )
            return False

        current_percents = [
            (item.get("amount_g", 0.0) or 0.0) * 100.0 / base_total
            for item in self.formulation_presenter.get_ui_items()
        ]

        locked_sum = sum(
            current_percents[idx]
            for idx, item in enumerate(self.formulation_presenter.get_ui_items())
            if item.get("locked") and idx != target_idx
        )
        if locked_sum > 100.0:
            QMessageBox.warning(
                self,
                "Porcentaje inválido",
                "Los ingredientes fijados ya superan el 100%. Libera uno para continuar.",
            )
            return False

        remaining = 100.0 - locked_sum - target_percent
        free_indices = [
            idx
            for idx, item in enumerate(self.formulation_presenter.get_ui_items())
            if not item.get("locked") and idx != target_idx
        ]

        if remaining < -0.0001:
            QMessageBox.warning(
                self,
                "Sin grado de libertad",
                "No hay porcentaje libre suficiente. Libera un ingrediente o reduce el valor.",
            )
            return False

        if not free_indices and abs(remaining) > 0.0001:
            QMessageBox.warning(
                self,
                "Sin grado de libertad",
                "Debe quedar al menos un ingrediente sin fijar para ajustar porcentajes.",
            )
            return False

        cur_free_sum = sum(current_percents[idx] for idx in free_indices)
        new_percents = [0.0 for _ in range(self.formulation_presenter.get_ingredient_count())]

        for idx, item in enumerate(self.formulation_presenter.get_ui_items()):
            if item.get("locked") and idx != target_idx:
                new_percents[idx] = current_percents[idx]
            elif idx == target_idx:
                new_percents[idx] = target_percent
            else:
                if cur_free_sum > 0:
                    new_percents[idx] = (
                        current_percents[idx] * remaining / cur_free_sum
                    )
                elif free_indices:
                    new_percents[idx] = remaining if idx == free_indices[0] else 0.0

        if any(pct < -0.001 for pct in new_percents):
            QMessageBox.warning(
                self,
                "Sin grado de libertad",
                "No hay porcentaje libre suficiente. Ajusta los valores o libera un ingrediente.",
            )
            return False

        for idx, pct in enumerate(new_percents):
            safe_pct = max(pct, 0.0)
            # Amount updated via presenter.update_ingredient_amount(idx, safe_pct * base_total / 100.0)

            self.formulation_presenter.update_ingredient_amount(idx, safe_pct * base_total / 100.0)
        return True


    def _remove_selected_from_formulation(self, table: QTableWidget) -> None:
        indexes = table.selectionModel().selectedRows()
        if not indexes:
            self.status_label.setText("Selecciona un ingrediente para eliminar.")
            return
        row = indexes[0].row()
        if 0 <= row < self.formulation_presenter.get_ingredient_count():
            # Ingredient removed via presenter

            removed = {"fdc_id": "", "description": "Removed"}

            # Remove from presenter
            try:
                self.formulation_presenter.remove_ingredient(row)
            except Exception as e:
                logging.error(f"Error removing ingredient via presenter: {e}")

            # If all remaining ingredients are locked, unlock the first one
            if (self.formulation_presenter.has_ingredients() and
                self.formulation_presenter.get_locked_count() == self.formulation_presenter.get_ingredient_count()):
                self.formulation_presenter.toggle_lock(0)
            self._refresh_formulation_views()
            self.status_label.setText(
                f"Eliminado {removed.get('fdc_id', '')} - {removed.get('description', '')}"
            )
        else:
            self.status_label.setText("Fila seleccionada inválida.")


    # ---- Async helpers ----
    def _run_in_thread(self, fn, args, on_success, on_error) -> None:
        """Start a worker thread for API calls and handle cleanup."""
        thread = QThread(self)
        worker = ApiWorker(fn, *args)
        worker.moveToThread(thread)
        self._workers.append(worker)

        thread.started.connect(worker.run)
        worker.finished.connect(on_success)
        worker.error.connect(on_error)
        worker.finished.connect(thread.quit)
        worker.error.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        worker.error.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)

        def _cleanup() -> None:
            if thread in self._threads:
                self._threads.remove(thread)
            if worker in self._workers:
                self._workers.remove(worker)

        thread.finished.connect(_cleanup)

        self._threads.append(thread)
        thread.start()

    # ---- Callbacks for async ops ----

    def _on_add_details_loaded(self, details, mode: str, value: float) -> None:
        logging.debug(
            f"_on_add_details_loaded fdc_id={details.get('fdcId', '?')} "
            f"mode={mode} value={value} "
            f"nutrients_count={len(details.get('foodNutrients', []) or [])}"
        )

        # Use details that AddWorker already fetched (avoid double API call)
        fdc_id = details.get("fdcId", "")
        amount_g = value if mode == "g" else 100.0

        # Normalize nutrients from the details we already have
        nutrients = normalize_nutrients(
            details.get("foodNutrients", []) or [],
            details.get("dataType")
        )
        self.nutrient_ordering.update_reference_from_details(details)

        desc = details.get("description", "") or ""
        brand = details.get("brandOwner", "") or ""
        data_type = details.get("dataType", "") or ""

        # Add ingredient directly to presenter using fetched data
        # (avoiding duplicate API call since AddWorker already fetched details)
        try:
            from domain.models import Food, Ingredient, Nutrient
            from decimal import Decimal

            domain_nutrients = tuple(
                Nutrient(
                    name=n["nutrient"]["name"],
                    unit=n["nutrient"].get("unitName", ""),
                    amount=Decimal(str(n["amount"])) if n.get("amount") is not None else Decimal("0"),
                    nutrient_id=n["nutrient"].get("id"),
                    nutrient_number=n["nutrient"].get("number"),
                )
                for n in nutrients
            )

            food = Food(
                fdc_id=fdc_id,
                description=desc,
                data_type=data_type,
                brand_owner=brand,
                nutrients=domain_nutrients,
            )

            ingredient = Ingredient(food=food, amount_g=Decimal(str(amount_g)))
            self.formulation_presenter._formulation.add_ingredient(ingredient)

        except Exception as e:
            logging.error(f"Error adding ingredient to presenter: {e}")

        self.nutrient_ordering.update_reference_from_details(details)

        if mode == "percent":
            success = self._apply_percent_edit(self.formulation_presenter.get_ingredient_count() - 1, value)
            if not success:
                # Percentage adjustment failed, remove last ingredient
                self.formulation_presenter.remove_ingredient(
                    self.formulation_presenter.get_ingredient_count() - 1
                )
                self.add_button.setEnabled(True)
                return

        self._populate_details_table(nutrients)
        self._refresh_formulation_views()
        self._select_preview_row(self.formulation_presenter.get_ingredient_count() - 1)
        msg_value = (
            self._format_amount_for_status(amount_g)
            if mode == "g"
            else f"{value:.2f} %"
        )
        self.status_label.setText(
            f"Agregado {fdc_id} - {desc} ({msg_value})"
        )
        self.add_button.setEnabled(True)
        self._upgrade_item_to_full(self.formulation_presenter.get_ingredient_count() - 1, int(fdc_id))


    def _on_add_error(self, message: str) -> None:
        self._reset_add_ui_state()
        self.status_label.setText(f"Error al agregar: {message}")
        logging.error(f"_on_add_error: {message}")


    def _upgrade_item_to_full(self, index: int, fdc_id: int) -> None:
        """Upgrade no-op: abridged es la única fuente ahora."""
        return

    def on_totals_checkbox_changed(self, item: QTableWidgetItem) -> None:
        """Sync export checkbox state into memory and update toggle label."""
        if self._is_importing():
            return
        if item.column() != 3:
            return
        nut_key = item.data(Qt.UserRole)
        if not nut_key:
            return
        self.nutrient_export_flags[nut_key] = item.checkState() == Qt.Checked
        self._update_toggle_export_button()


    def on_toggle_export_clicked(self) -> None:
        """Toggle all nutrient export checkboxes on/off."""
        if not self.nutrient_export_flags:
            return
        all_checked = all(self.nutrient_export_flags.values())
        new_state = not all_checked
        for key in list(self.nutrient_export_flags.keys()):
            self.nutrient_export_flags[key] = new_state
        self._populate_totals_table()
