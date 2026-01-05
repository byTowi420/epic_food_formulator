"""Excel export functionality with full USDA-style formatting.

Exports formulations and nutrient data to Excel with proper formatting,
categories, formulas, and colors matching the original implementation.
"""

from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

from config.constants import DATA_TYPE_PRIORITY
from domain.exceptions import ExportError
from domain.models import Formulation, Ingredient
from domain.services.unit_normalizer import canonical_unit, convert_mass, normalize_mass_unit
from services.nutrient_normalizer import canonical_alias_name


class ExcelExporter:
    """Export formulations to Excel format with full USDA formatting."""

    def __init__(self):
        """Initialize exporter with nutrient catalog."""
        self._nutrient_catalog = self._build_nutrient_catalog()
        self._nutrient_category_map = self._build_category_map()
        self._nutrient_order_map = self._build_order_map()
        self.data_type_priority = DATA_TYPE_PRIORITY

    def export_formulation(
        self,
        formulation: Formulation,
        nutrient_totals: Dict[str, Decimal] | None = None,
        output_path: Path | str = "",
        *,
        export_flags: Dict[str, bool] | None = None,
        mass_unit: str | None = None,
    ) -> None:
        """Export formulation with nutrient breakdown to Excel.

        Args:
            formulation: Formulation to export
            nutrient_totals: Optional nutrient totals (not used, formulas calculate in Excel)
            output_path: Path to save Excel file

        Raises:
            ExportError: If export fails
        """
        # Handle both signatures: (formulation, totals, path) and (formulation, path)
        if isinstance(nutrient_totals, (str, Path)):
            output_path = nutrient_totals
            nutrient_totals = None
        try:
            unit = normalize_mass_unit(mass_unit) or "g"
            unit_decimals = {
                "g": 1,
                "kg": 3,
                "ton": 6,
                "lb": 3,
                "oz": 3,
            }.get(unit, 2)
            unit_format = "0" if unit_decimals <= 0 else f"0.{'0' * unit_decimals}"

            wb = Workbook()
            ws = wb.active
            ws.title = "Ingredientes"
            totals_sheet = wb.create_sheet("Totales")

            # Base headers
            base_headers = [
                "FDC ID",
                "Ingrediente",
                "Marca / Origen",
                "Tipo de dato",
                f"Cantidad ({unit})",
                "Cantidad (%)",
            ]

            # Collect nutrient columns and categories
            nutrient_headers, header_categories, header_key_map = (
                self._collect_nutrient_columns(formulation, export_flags=export_flags)
            )
            header_by_key = {v: k for k, v in header_key_map.items()}

            # Define fills and styles
            header_fill = PatternFill("solid", fgColor="D9D9D9")
            total_fill = PatternFill("solid", fgColor="FFF2CC")
            category_fills = {
                "Proximates": PatternFill("solid", fgColor="DAEEF3"),
                "Carbohydrates": PatternFill("solid", fgColor="E6B8B7"),
                "Minerals": PatternFill("solid", fgColor="C4D79B"),
                "Vitamins and Other Components": PatternFill("solid", fgColor="FFF2CC"),
                "Lipids": PatternFill("solid", fgColor="D9E1F2"),
                "Amino acids": PatternFill("solid", fgColor="E4DFEC"),
            }
            center = Alignment(horizontal="center", vertical="center")

            # Row 1: group titles (base + nutrient categories)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(base_headers))
            ws.cell(row=1, column=1, value="Detalles de formulación").alignment = center

            if nutrient_headers:
                start_col = len(base_headers) + 1
                col = start_col
                while col < start_col + len(nutrient_headers):
                    idx = col - start_col
                    category = header_categories.get(nutrient_headers[idx], "Nutrientes")
                    run_start = col
                    while (
                        col < start_col + len(nutrient_headers)
                        and header_categories.get(
                            nutrient_headers[col - start_col], "Nutrientes"
                        )
                        == category
                    ):
                        col += 1
                    run_end = col - 1
                    ws.merge_cells(start_row=1, start_column=run_start, end_row=1, end_column=run_end)
                    cat_cell = ws.cell(row=1, column=run_start, value=category)
                    cat_cell.alignment = center
                    cat_cell.fill = category_fills.get(category, header_fill)

            # Row 2: headers
            headers = base_headers + nutrient_headers
            for col_idx, name in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col_idx, value=name)
                cell.fill = (
                    category_fills.get(header_categories.get(name, ""), header_fill)
                    if col_idx > len(base_headers)
                    else header_fill
                )
                cell.alignment = center

            start_row = 3
            grams_col = base_headers.index(f"Cantidad ({unit})") + 1
            percent_col = base_headers.index("Cantidad (%)") + 1
            data_rows = len(formulation.ingredients)
            end_row = start_row + data_rows - 1

            # Write ingredient rows
            for idx, ingredient in enumerate(formulation.ingredients):
                row = start_row + idx
                amount_val = convert_mass(ingredient.amount_g, "g", unit)
                values = [
                    ingredient.food.fdc_id,
                    ingredient.food.description,
                    ingredient.food.brand_owner,
                    ingredient.food.data_type,
                    float(amount_val) if amount_val is not None else float(ingredient.amount_g),
                    None,  # placeholder for percent formula
                ]
                for col_idx, val in enumerate(values, start=1):
                    ws.cell(row=row, column=col_idx, value=val)

                # Percentage formula
                gram_cell = f"{get_column_letter(grams_col)}{row}"
                total_range = f"${get_column_letter(grams_col)}${start_row}:${get_column_letter(grams_col)}${end_row}"
                ws.cell(
                    row=row,
                    column=percent_col,
                    value=f"={gram_cell}/SUM({total_range})",
                ).number_format = "0.00%"

                # Nutrient values
                nut_map = self._nutrients_by_header(ingredient, header_by_key)
                for offset, header in enumerate(nutrient_headers, start=len(base_headers) + 1):
                    if header in nut_map:
                        ws.cell(row=row, column=offset, value=nut_map[header])

            # Total row
            total_row = end_row + 1
            ws.cell(row=total_row, column=1, value="Total")
            ws.cell(row=total_row, column=2, value="Formulado")
            ws.cell(row=total_row, column=3, value="Formulado")
            ws.cell(row=total_row, column=4, value="Formulado")

            # Total grams with SUBTOTAL formula
            gram_total_cell = ws.cell(
                row=total_row,
                column=grams_col,
                value=f"=SUBTOTAL(9,{get_column_letter(grams_col)}{start_row}:{get_column_letter(grams_col)}{end_row})",
            )
            gram_total_cell.fill = total_fill

            # Total percentage
            percent_total_cell = ws.cell(row=total_row, column=percent_col, value="100%")
            percent_total_cell.number_format = "0.00%"
            percent_total_cell.fill = total_fill

            # Nutrient totals with SUMPRODUCT formula
            for offset, header in enumerate(nutrient_headers, start=len(base_headers) + 1):
                col_letter = get_column_letter(offset)
                formula = (
                    f"=SUMPRODUCT(${get_column_letter(percent_col)}${start_row}:${get_column_letter(percent_col)}${end_row},"
                    f"${col_letter}${start_row}:${col_letter}${end_row})"
                )
                cell = ws.cell(row=total_row, column=offset, value=formula)
                cell.fill = total_fill

            # Number formats for data rows
            for row in range(start_row, total_row + 1):
                ws.cell(row=row, column=grams_col).number_format = unit_format

            # Freeze panes to keep headers/base columns visible
            freeze_col = len(base_headers) + 1 if nutrient_headers else 1
            ws.freeze_panes = f"{get_column_letter(freeze_col)}3"

            # Adjust column widths
            widths = {
                "A": 12,
                "B": 35,
                "C": 18,
                "D": 14,
                "E": 12,
                "F": 12,
            }
            for col_letter, width in widths.items():
                ws.column_dimensions[col_letter].width = width

            # Set default width for nutrient columns
            for col_idx in range(len(base_headers) + 1, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 13

            # Totales sheet: simple reference to totals row
            totals_headers = ["Nutriente", "Total", "Unidad"]
            for col_idx, name in enumerate(totals_headers, start=1):
                cell = totals_sheet.cell(row=1, column=col_idx, value=name)
                cell.fill = header_fill
                cell.alignment = center

            for idx, header in enumerate(nutrient_headers, start=1):
                name_part, unit = self._split_header_unit(header)
                totals_sheet.cell(row=idx + 1, column=1, value=name_part)
                totals_sheet.cell(row=idx + 1, column=3, value=unit or "")
                source_cell = f"Ingredientes!{get_column_letter(len(base_headers) + idx)}{total_row}"
                totals_sheet.cell(row=idx + 1, column=2, value=f"={source_cell}")

            wb.save(output_path)

        except Exception as exc:
            raise ExportError(f"Failed to export to Excel: {exc}") from exc

    def _collect_nutrient_columns(
        self,
        formulation: Formulation,
        export_flags: Dict[str, bool] | None = None,
    ) -> Tuple[List[str], Dict[str, str], Dict[str, str]]:
        """Collect ordered nutrient headers and their categories."""
        candidates: Dict[str, Dict[str, Any]] = {}
        categories_seen_order: Dict[str, int] = {}
        preferred_order = [cat for cat, _ in self._nutrient_catalog]
        preferred_count = len(preferred_order)

        for ingredient in formulation.ingredients:
            data_priority = self.data_type_priority.get(
                ingredient.food.data_type.strip(), len(self.data_type_priority)
            )

            for nutrient in sorted(ingredient.food.nutrients, key=lambda n: self._nutrient_order(n, 0)):
                header_key, canonical_name, canonical_unit_str = self._header_key(nutrient)

                if export_flags and not export_flags.get(header_key, True):
                    continue

                if not header_key or not canonical_name:
                    continue

                category = self._category_for_nutrient(canonical_name)
                if category not in categories_seen_order:
                    categories_seen_order[category] = len(categories_seen_order)

                order = self._nutrient_order_map.get(canonical_name.strip().lower())
                if order is None:
                    order = self._nutrient_order(nutrient, len(candidates))

                # Keep kcal ahead of kJ when both present
                unit_lower = canonical_unit_str.strip().lower()
                if canonical_name.strip().lower() == "energy":
                    if unit_lower == "kcal":
                        order = order - 0.1 if isinstance(order, (int, float)) else order
                    elif unit_lower == "kj":
                        order = order + 0.1 if isinstance(order, (int, float)) else order

                if header_key not in candidates:
                    header = self._make_header(canonical_name, canonical_unit_str)
                    candidates[header_key] = {
                        "header": header,
                        "category": category,
                        "order": order,
                        "data_priority": data_priority,
                        "header_key": header_key,
                    }
                else:
                    current = candidates[header_key]
                    if data_priority < current["data_priority"]:
                        current["data_priority"] = data_priority
                    if order < current["order"]:
                        current["order"] = order

        # Sort by category order then nutrient order
        def category_rank(cat: str) -> int:
            if cat in preferred_order:
                return preferred_order.index(cat)
            return preferred_count + categories_seen_order.get(cat, 999)

        sorted_candidates = sorted(
            candidates.values(),
            key=lambda c: (
                category_rank(c["category"]),
                c["order"],
                c["header"].lower(),
            ),
        )

        ordered_headers: List[str] = [c["header"] for c in sorted_candidates]
        categories: Dict[str, str] = {c["header"]: c["category"] for c in sorted_candidates}
        header_key_map: Dict[str, str] = {c["header"]: c["header_key"] for c in sorted_candidates}

        return ordered_headers, categories, header_key_map

    def _nutrients_by_header(
        self, ingredient: Ingredient, header_by_key: Dict[str, str]
    ) -> Dict[str, float]:
        """Build a mapping of template header -> nutrient amount (per 100 g)."""
        out: Dict[str, float] = {}
        best_priority: Dict[str, int] = {}
        allowed_keys = set(header_by_key.keys())
        alias_priority = {
            "carbohydrate, by difference": 2,
            "carbohydrate, by summation": 1,
            "carbohydrate by summation": 1,
            "sugars, total": 2,
            "total sugars": 1,
        }

        for nutrient in ingredient.food.nutrients:
            header_key, name, unit = self._header_key(nutrient)
            if not header_key or header_key not in allowed_keys:
                continue
            header = header_by_key[header_key]
            priority = alias_priority.get(nutrient.name.strip().lower(), 0)
            current_best = best_priority.get(header, -1)
            if priority < current_best:
                continue

            best_priority[header] = priority
            out[header] = float(nutrient.amount)

        return out

    def _category_for_nutrient(self, name: str) -> str:
        """Resolve a nutrient category using the static catalog."""
        lower = (name or "").strip().lower()
        if lower in self._nutrient_category_map:
            return self._nutrient_category_map[lower]

        # Fallback heuristics
        amino_acids = {
            "tryptophan",
            "threonine",
            "isoleucine",
            "leucine",
            "lysine",
            "methionine",
            "phenylalanine",
            "tyrosine",
            "valine",
            "arginine",
            "histidine",
            "alanine",
            "aspartic acid",
            "glutamic acid",
            "glycine",
            "proline",
            "serine",
            "hydroxyproline",
            "cysteine",
        }
        oligosaccharides = {"verbascose", "raffinose", "stachyose"}
        isoflavones = {"daidzin", "genistin", "glycitin", "daidzein", "genistein"}
        organic_acids = {"citric acid", "malic acid", "oxalic acid", "quinic acid"}

        # Proximates
        if lower in {
            "water",
            "energy",
            "nitrogen",
            "protein",
            "total fat (nlea)",
            "total lipid (fat)",
            "ash",
            "carbohydrate, by difference",
        }:
            return "Proximates"

        # Carbohydrates
        if any(
            kw in lower
            for kw in [
                "fiber",
                "sugar",
                "glucose",
                "fructose",
                "lactose",
                "sucrose",
                "maltose",
                "galactose",
                "starch",
            ]
        ) and "fatty" not in lower:
            return "Carbohydrates"

        # Minerals
        if any(
            mineral in lower
            for mineral in [
                "calcium",
                "iron",
                "magnesium",
                "phosphorus",
                "potassium",
                "sodium",
                "zinc",
                "copper",
                "manganese",
                "iodine",
                "selenium",
                "molybdenum",
                "fluoride",
            ]
        ):
            return "Minerals"

        # Vitamins
        vitamin_like = (
            lower.startswith("vitamin")
            or "thiamin" in lower
            or "riboflavin" in lower
            or "niacin" in lower
            or "pantothenic" in lower
            or "biotin" in lower
            or "tocopherol" in lower
            or "tocotrienol" in lower
            or "carotene" in lower
            or "lycopene" in lower
            or "lutein" in lower
            or "zeaxanthin" in lower
            or "retinol" in lower
            or "folate" in lower
            or "folic acid" in lower
            or "betaine" in lower
            or "choline" in lower
            or "caffeine" in lower
            or "theobromine" in lower
        )
        if vitamin_like:
            return "Vitamins and Other Components"

        # Amino acids
        if lower in amino_acids:
            return "Amino acids"

        # Lipids
        if (
            "fatty acids" in lower
            or lower.startswith(("sfa ", "mufa ", "pufa ", "tfa "))
            or lower in {"cholesterol", "total lipid (fat)", "total fat (nlea)"}
        ):
            return "Lipids"

        # Phytosterols
        if "sterol" in lower:
            return "Phytosterols"

        # Organic acids
        if lower in organic_acids or (lower.endswith("acid") and lower not in amino_acids):
            return "Organic acids"

        # Oligosaccharides
        if lower in oligosaccharides:
            return "Oligosaccharides"

        # Isoflavones
        if lower in isoflavones:
            return "Isoflavones"

        return "Nutrientes"

    def _header_key(self, nutrient) -> Tuple[str, str, str]:
        """Return a stable header key plus canonical name and unit for a nutrient."""
        name = canonical_alias_name(nutrient.name or "")
        unit = canonical_unit(nutrient.unit or "")
        unit_part = unit.strip().lower()
        name_part = name.strip().lower()

        if name_part:
            header_key = f"{name_part}|{unit_part}"
        else:
            return "", name, unit

        return header_key, name, unit

    def _make_header(self, name: str, unit: str) -> str:
        """Create display header with unit in parentheses."""
        if unit and unit.strip():
            return f"{name} ({unit})"
        return name

    def _nutrient_order(self, nutrient, default: int) -> float:
        """Get nutrient order for sorting."""
        # Try to get from order map first
        name_lower = nutrient.name.strip().lower()
        if name_lower in self._nutrient_order_map:
            return self._nutrient_order_map[name_lower]

        # Use nutrient_id or nutrient_number if available
        if nutrient.nutrient_id:
            return float(nutrient.nutrient_id)
        if nutrient.nutrient_number:
            try:
                return float(nutrient.nutrient_number)
            except (ValueError, TypeError):
                pass

        return float(default + 10000)

    def _split_header_unit(self, header: str) -> Tuple[str, str]:
        """Split header into name and unit parts."""
        if header.endswith(")") and " (" in header:
            name, unit = header.rsplit(" (", 1)
            return name, unit[:-1]
        return header, ""

    def _build_nutrient_catalog(self) -> List[Tuple[str, List[str]]]:
        """Static catalog to enforce ordering/categories."""
        return [
            (
                "Proximates",
                [
                    "Water",
                    "Energy",
                    "Nitrogen",
                    "Protein",
                    "Total fat (NLEA)",
                    "Total lipid (fat)",
                    "Ash",
                    "Carbohydrate, by difference",
                ],
            ),
            (
                "Carbohydrates",
                [
                    "Fiber, total dietary",
                    "Fiber, soluble",
                    "Fiber, insoluble",
                    "Total dietary fiber (AOAC 2011.25)",
                    "High Molecular Weight Dietary Fiber (HMWDF)",
                    "Low Molecular Weight Dietary Fiber (LMWDF)",
                    "Sugars, Total",
                    "Sucrose",
                    "Glucose",
                    "Fructose",
                    "Lactose",
                    "Maltose",
                    "Galactose",
                    "Starch",
                    "Resistant starch",
                    "Sugars, added",
                ],
            ),
            (
                "Minerals",
                [
                    "Calcium, Ca",
                    "Iron, Fe",
                    "Magnesium, Mg",
                    "Phosphorus, P",
                    "Potassium, K",
                    "Sodium, Na",
                    "Zinc, Zn",
                    "Copper, Cu",
                    "Manganese, Mn",
                    "Iodine, I",
                    "Selenium, Se",
                    "Molybdenum, Mo",
                    "Fluoride, F",
                ],
            ),
            (
                "Vitamins and Other Components",
                [
                    "Thiamin",
                    "Riboflavin",
                    "Niacin",
                    "Vitamin B-6",
                    "Folate, total",
                    "Folic acid",
                    "Folate, DFE",
                    "Choline, total",
                    "Vitamin B-12",
                    "Vitamin A, RAE",
                    "Vitamin A, IU",
                    "Vitamin D (D2 + D3)",
                    "Vitamin K (phylloquinone)",
                    "Vitamin E (alpha-tocopherol)",
                    "Vitamin C, total ascorbic acid",
                    "Pantothenic acid",
                    "Biotin",
                    "Caffeine",
                    "Theobromine",
                ],
            ),
            (
                "Lipids",
                [
                    "Fatty acids, total saturated",
                    "Fatty acids, total monounsaturated",
                    "Fatty acids, total polyunsaturated",
                    "Fatty acids, total trans",
                    "Cholesterol",
                ],
            ),
            (
                "Amino acids",
                [
                    "Tryptophan",
                    "Threonine",
                    "Isoleucine",
                    "Leucine",
                    "Lysine",
                    "Methionine",
                    "Phenylalanine",
                    "Tyrosine",
                    "Valine",
                    "Arginine",
                    "Histidine",
                    "Alanine",
                    "Aspartic acid",
                    "Glutamic acid",
                    "Glycine",
                    "Proline",
                    "Serine",
                    "Hydroxyproline",
                    "Cysteine",
                ],
            ),
            ("Phytosterols", ["Phytosterols", "Beta-sitosterol", "Campesterol", "Stigmasterol"]),
            ("Organic acids", ["Citric acid", "Malic acid", "Oxalic acid", "Quinic acid"]),
            ("Oligosaccharides", ["Verbascose", "Raffinose", "Stachyose"]),
            ("Isoflavones", ["Daidzin", "Genistin", "Glycitin", "Daidzein", "Genistein"]),
        ]

    def _build_category_map(self) -> Dict[str, str]:
        """Build a map of nutrient name -> category."""
        category_map = {}
        for category, nutrients in self._nutrient_catalog:
            for nutrient in nutrients:
                category_map[nutrient.lower()] = category
        return category_map

    def _build_order_map(self) -> Dict[str, int]:
        """Build a map of nutrient name -> order."""
        order_map = {}
        order = 0
        for _, nutrients in self._nutrient_catalog:
            for nutrient in nutrients:
                order_map[nutrient.lower()] = order
                order += 1
        return order_map
